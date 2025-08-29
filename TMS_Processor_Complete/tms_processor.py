import pandas as pd
import numpy as np
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class ModernTMSProcessor:
    # Configuration constants
    DEFAULT_HEADER_ROW = 8
    DEFAULT_DATA_START_ROW = 11
    MIN_DATA_COLUMNS = 5
    EXPECTED_COLUMNS = 21
    
    def __init__(self):
        self.raw_data = None
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}
        
    def _extract_title_info(self, df_raw):
        """Extract title and report information from the top rows"""
        title_info = {}
        
        try:
            # Extract report title with bounds checking
            if len(df_raw) > 1 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[1, 1]):
                title_info['report_title'] = str(df_raw.iloc[1, 1])
            
            # Extract company name
            if len(df_raw) > 3 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[3, 1]):
                title_info['company_name'] = str(df_raw.iloc[3, 1])
            
            # Extract date range
            if len(df_raw) > 5 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[5, 1]):
                title_info['date_range'] = str(df_raw.iloc[5, 1])
        except (IndexError, KeyError):
            # If extraction fails, continue with empty title_info
            pass
        
        return title_info
    
    def _detect_data_structure(self, df_raw):
        """Intelligently detect header and data start positions"""
        header_row = self.DEFAULT_HEADER_ROW
        data_start_row = self.DEFAULT_DATA_START_ROW
        
        # Look for header indicators in different rows
        header_indicators = ['Load No.', 'Carrier', 'Service Type', 'Ship Date']
        
        for row_idx in range(5, min(15, len(df_raw))):
            row_data = df_raw.iloc[row_idx].dropna().astype(str).tolist()
            row_str = ' '.join(row_data).lower()
            
            # Check if this row contains header-like content
            matches = sum(1 for indicator in header_indicators if indicator.lower() in row_str)
            if matches >= 2:  # Found at least 2 header indicators
                header_row = row_idx
                data_start_row = row_idx + 2  # Skip potential blank row
                break
        
        return header_row, data_start_row
        
    def _remove_duplicate_headers(self, df):
        """Remove duplicate header rows that appear in the middle of data"""
        # Look for rows that contain header-like text
        header_indicators = ['Load No.', 'Carrier', 'Service Type']
        
        rows_to_drop = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.dropna().tolist()])
            if any(indicator in row_str for indicator in header_indicators):
                # Check if this looks like a header row (not actual data)
                if not any(str(val).startswith('A') and str(val)[1:].isdigit() for val in row.dropna().tolist()):
                    rows_to_drop.append(idx)
        
        return df.drop(rows_to_drop)
    
    def clean_and_process_data(self, file_path):
        """Main function to clean and process the TMS Excel file"""
        try:
            # Read raw Excel file
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
            
            # Extract title information from top rows
            self.title_info = self._extract_title_info(df_raw)
            
            # Intelligently detect data structure
            header_row, data_start_row = self._detect_data_structure(df_raw)
            
            # Get headers
            headers = df_raw.iloc[header_row].dropna().tolist()
            
            # Extract data starting from row 10
            data_df = df_raw.iloc[data_start_row:].copy()
            
            # Remove completely empty rows and duplicate header rows
            data_df = data_df.dropna(how='all')
            data_df = self._remove_duplicate_headers(data_df)
            
            # Reset index after dropping rows
            data_df = data_df.reset_index(drop=True)
            
            # Extract relevant data columns (skip column A, start from column 2)
            # Check how many columns we actually have
            max_cols = min(22, len(data_df.columns))
            relevant_columns = list(range(2, max_cols + 1))
            data_df = data_df.iloc[:, relevant_columns]
            
            # Set proper column names with full descriptive headers
            base_column_names = [
                'Load No.', 'Ship Date', 'Origin City', 'Origin State', 'Origin Postal',
                'Destination City', 'Destination State', 'Destination Postal',
                'Selected Carrier', 'Selected Service Type', 'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
                'Least Cost Carrier', 'Least Cost Service Type', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
                'Potential Savings'
            ]
            
            # Ensure column names match the actual extracted columns
            if len(base_column_names) != len(data_df.columns):
                print(f"Warning: Column count mismatch. Expected {len(base_column_names)}, got {len(data_df.columns)}")
                # Adjust column names to match actual columns
                if len(data_df.columns) < len(base_column_names):
                    column_names = base_column_names[:len(data_df.columns)]
                else:
                    # Add generic names for extra columns
                    extra_columns = [f'Column_{i}' for i in range(len(base_column_names), len(data_df.columns))]
                    column_names = base_column_names + extra_columns
            else:
                column_names = base_column_names
            
            data_df.columns = column_names
            
            # Clean and fix data types
            data_df = self._clean_data_types(data_df)
            
            # Remove rows where Load No. is missing or empty
            data_df = data_df.dropna(subset=['Load No.'])
            data_df = data_df[data_df['Load No.'].astype(str).str.strip() != '']
            data_df = data_df[data_df['Load No.'].astype(str).str.strip() != 'nan']
            
            # Remove any remaining rows that are mostly empty
            data_df = data_df.dropna(thresh=5)  # Keep rows with at least 5 non-null values
            
            # Reset index again after final cleaning
            data_df = data_df.reset_index(drop=True)
            
            # Apply business logic rules
            data_df = self._apply_business_logic(data_df)
            
            # Sort by Destination City
            data_df = data_df.sort_values('Destination City', na_position='last')
            
            # Calculate summary statistics
            self._calculate_summary_stats(data_df)
            
            self.processed_data = data_df
            return data_df
            
        except (FileNotFoundError, PermissionError) as e:
            raise FileNotFoundError(f"Cannot access file: {str(e)}")
        except (pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            raise ValueError(f"Invalid Excel file format: {str(e)}")
        except Exception as e:
            raise RuntimeError(f"Error processing file: {str(e)}")
    
    def _clean_data_types(self, df):
        """Clean and fix data types for each column"""
        df = df.copy()
        
        # Convert numeric columns
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings'
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Ensure PS column is properly numeric and handle any string values
        if 'PS' in df.columns:
            # First try to convert to numeric, handling any string values
            df['PS'] = pd.to_numeric(df['PS'], errors='coerce')
            # Fill any NaN values with 0
            df['PS'] = df['PS'].fillna(0)
        
        # Convert date column and format as MM/DD/YY
        if 'Ship Date' in df.columns:
            df['Ship Date'] = pd.to_datetime(df['Ship Date'], errors='coerce').dt.strftime('%m/%d/%y')
        
        # Clean string columns
        string_columns = [
            'Load No.', 'Origin City', 'Origin State', 'Origin Postal',
            'Destination City', 'Destination State', 'Destination Postal',
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]
        
        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
        
        return df
    
    def _apply_business_logic(self, df):
        """Apply TMS business logic rules"""
        df = df.copy()
        
        try:
            # Ensure PS column is numeric from the start to avoid comparison errors
            if 'PS' in df.columns:
                df['PS'] = pd.to_numeric(df['PS'], errors='coerce').fillna(0)
            else:
                print("Warning: PS column not found in dataframe")
                print(f"Available columns: {df.columns.tolist()}")
        
            # Rule 1: Same Carriers - Set Potential Savings to 0
            same_carrier_mask = (
                (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) & 
                (df['Selected Carrier'].notna()) & 
                (df['Least Cost Carrier'].notna()) &
                (df['Selected Carrier'].astype(str) != '') & 
                (df['Least Cost Carrier'].astype(str) != '') &
                (df['Selected Carrier'].astype(str) != 'nan') & 
                (df['Least Cost Carrier'].astype(str) != 'nan')
            )
            if 'Potential Savings' in df.columns:
                df.loc[same_carrier_mask, 'Potential Savings'] = 0
            
            # Rule 2: Empty Least Cost - Copy Selected data and set savings to 0
            empty_least_cost_mask = (
                df['Least Cost Carrier'].isna() | 
                (df['Least Cost Carrier'].astype(str) == '') |
                (df['Least Cost Carrier'].astype(str) == 'nan')
            )
            column_pairs = [
                ('Selected Carrier', 'Least Cost Carrier'), ('Selected Service Type', 'Least Cost Service Type'), ('Selected Transit Days', 'Least Cost Transit Days'),
                ('Selected Freight Cost', 'Least Cost Freight Cost'), ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), ('Selected Total Cost', 'Least Cost Total Cost')
            ]
            self._copy_selected_to_least_cost(df, empty_least_cost_mask, column_pairs)
            if 'Potential Savings' in df.columns:
                df.loc[empty_least_cost_mask, 'Potential Savings'] = 0

            # Rule 3: Negative Savings - Copy Selected data and set savings to 0
            if 'Potential Savings' in df.columns:
                # Ensure Potential Savings is numeric before comparison
                ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
                negative_savings_mask = ps_numeric < 0
                self._copy_selected_to_least_cost(df, negative_savings_mask, column_pairs)
                df.loc[negative_savings_mask, 'Potential Savings'] = 0
                
        except Exception as e:
            print(f"Error in _apply_business_logic: {e}")
            print(f"DataFrame info:")
            print(f"  Shape: {df.shape}")
            print(f"  Columns: {df.columns.tolist()}")
            if 'PS' in df.columns:
                print(f"  PS column dtype: {df['PS'].dtype}")
            raise e
            
        return df
    
    def _copy_selected_to_least_cost(self, df, mask, column_pairs):
        """Helper method to copy selected carrier data to least cost columns"""
        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]
    
    def _calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'percentage_savings': 0,
                'loads_with_savings': 0,
                'total_savings_opportunity': 0
            }
            return
        
        # Basic stats - ensure numeric columns are properly converted
        total_loads = len(df)
        total_selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        total_least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        total_potential_savings = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0).sum()
        
        # Advanced stats - optimize by filtering once
        ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
        savings_df = df[ps_numeric > 0]
        loads_with_savings = len(savings_df)
        total_savings_opportunity = pd.to_numeric(savings_df['Potential Savings'], errors='coerce').fillna(0).sum()
        
        # Calculate percentages
        if total_selected_cost > 0:
            percentage_savings = (total_potential_savings / total_selected_cost) * 100
        else:
            percentage_savings = 0
        
        if total_loads > 0:
            average_savings_per_load = total_potential_savings / total_loads
        else:
            average_savings_per_load = 0
        
        self.summary_stats = {
            'total_loads': total_loads,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'percentage_savings': percentage_savings,
            'loads_with_savings': loads_with_savings,
            'total_savings_opportunity': total_savings_opportunity
        }
    
    def save_processed_data(self, output_file):
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Processed Data"
        
        # Add company and date info (no big title rows)
        row = 1
        if self.title_info:
            # Style company and date range with expanded width
            header_style_border = Border(
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                bottom=Side(style='medium', color='1F4E79')
            )
            
            if 'company_name' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                company_cell = ws_data[f'A{row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="FFFFFF")
                company_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                company_cell.border = header_style_border
                company_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
                
            if 'date_range' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                date_cell = ws_data[f'A{row}']
                date_cell.value = f"Date Range: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="FFFFFF")
                date_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                date_cell.border = header_style_border
                date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
            
            # Add section headers row with color coding
            row = 4
            # Selected Carrier section (columns I-N, which are 9-14) - Light Blue
            selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
            selected_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            selected_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('I4:N4')
            for col in range(9, 15):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            
            # Least Cost Carrier section (columns O-T, which are 15-20) - Light Orange
            least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
            least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('O4:T4')
            for col in range(15, 21):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            
            row = 5  # Headers will be on row 5
        
        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Color code headers based on section
            if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light Blue
            elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Light Orange
            elif header == 'Potential Savings':  # Potential Savings column - Green
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
            else:
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Default blue
        
        # Add data with alternating row colors
        data_border = Border(
            left=Side(style='thin', color='E8E8E8'),
            right=Side(style='thin', color='E8E8E8'),
            top=Side(style='thin', color='E8E8E8'),
            bottom=Side(style='thin', color='E8E8E8')
        )
        
        # Filter out any remaining empty rows before writing to Excel
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']
        
        # Ensure all data is properly typed before processing
        for col in clean_data.columns:
            if col in ['Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings']:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')
        
        for data_idx, data_row in enumerate(dataframe_to_rows(clean_data, index=False, header=False)):
            # Skip rows that are mostly empty
            # Ensure we're comparing integers by converting the sum result
            try:
                non_empty_count = sum(1 for val in data_row if val is not None and str(val).strip() != '' and str(val) != 'nan')
                if non_empty_count < 3:
                    continue
            except Exception as e:
                print(f"Error processing row {data_idx}: {e}")
                print(f"Row data: {data_row}")
                continue
                
            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"
            
            # First pass: collect all content lengths to determine optimal row height
            max_content_length = 0
            for col_idx, value in enumerate(data_row, 1):
                content_length = len(str(value)) if value else 0
                max_content_length = max(max_content_length, content_length)
            
            # Calculate optimal row height based on longest content in the row
            if max_content_length > 30:  # Very long content (like "CENTRAL TRANSPORT INTERNATIONAL")
                optimal_height = min(40, max(20, max_content_length * 0.7))  # More generous scaling for long names
            elif max_content_length > 20:  # Long content
                optimal_height = min(30, max(18, max_content_length * 0.6))  # Scale height with content
            elif max_content_length > 15:  # Medium content
                optimal_height = 22
            else:
                optimal_height = 18  # Default compact height for short content
            
            # Set the row height once for the entire row
            ws_data.row_dimensions[row].height = optimal_height
            
            # Second pass: apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                # Center all cell contents and enable text wrapping for compactness
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border
                
                # Apply color coding to data cells based on section
                if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(start_color=light_blue_bg, end_color=light_blue_bg, fill_type="solid")
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(start_color=light_orange_bg, end_color=light_orange_bg, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                
                # Format currency columns and apply green color for positive Potential Savings values
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost', 'Selected Freight Cost', 'Least Cost Freight Cost', 'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if headers[col_idx-1] in currency_columns or headers[col_idx-1] == 'Potential Savings':
                    cell.number_format = '"$"#,##0.00'
                    # Apply light green background for positive Potential Savings values
                    if headers[col_idx-1] == 'Potential Savings':
                        try:
                            # Convert value to float for comparison, handle None and string values
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                # Handle different value types safely
                                if isinstance(value, (int, float)):
                                    numeric_value = float(value)
                                else:
                                    numeric_value = float(str(value).replace('$', '').replace(',', ''))
                                if numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        except (ValueError, TypeError, AttributeError):
                            pass  # Skip coloring if value can't be converted
                    cell.font = Font(size=10, bold=False, color="2C3E50")
            else:
                cell.font = Font(size=10, color="495057")
        
        # Enable auto-filter over header and data range (no freeze panes)
        try:
            header_row_idx = 5
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Skip totals row since we have Performance Insights instead
        row += 1
        
        # Legend column widths will be determined by auto-fit below
        
        # Position Performance Insights box below the data table area
        # Add Performance Insights (PI) compact block at columns O-P (15-16)
        insights_row = row + 1  # Place closer to table for compactness
        insights_col = 15  # Column O
        
        thick_border = Border(
            left=Side(style='medium', color='2E86AB'),
            right=Side(style='medium', color='2E86AB'),
            top=Side(style='medium', color='2E86AB'),
            bottom=Side(style='medium', color='2E86AB')
        )
        thin_border = Border(
            left=Side(style='thin', color='B0C4DE'),
            right=Side(style='thin', color='B0C4DE'),
            top=Side(style='thin', color='B0C4DE'),
            bottom=Side(style='thin', color='B0C4DE')
        )
        
        # Two-row PI with labels in W and values in X
        metric_row = insights_row
        value_row = insights_row
        pi_items = [
            ("Total Selected Carrier Costs", f"${self.summary_stats['total_selected_cost']:,.2f}"),
            ("Total Potential Savings", f"${self.summary_stats['total_potential_savings']:,.2f}")
        ]
        current_row = metric_row
        for label, val in pi_items:
            lcell = ws_data.cell(row=current_row, column=insights_col, value=label)
            lcell.font = Font(size=13, bold=True, color="2C3E50")  # Reduced from 14 to 13
            lcell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            lcell.border = thin_border
            lcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            vcell = ws_data.cell(row=current_row, column=insights_col + 1, value=val)
            vcell.font = Font(size=13, bold=True, color="27AE60")  # Reduced from 14 to 13
            vcell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
            vcell.border = thin_border
            vcell.alignment = Alignment(horizontal="left", vertical="center")

            current_row += 1
        
        # Let autofit later determine widths for all ws_data columns (including PI/CAL)
        
        # Create beautiful summary sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Title with modern styling
        title_cell = ws_summary['A1']
        title_cell.value = "📊 TMS REPORT - SUMMARY STATISTICS"
        title_cell.font = Font(size=18, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells('A1:D1')
        
        # Apply styling to merged title cells
        title_border = Border(
            left=Side(style='medium', color='1F4E79'),
            right=Side(style='medium', color='1F4E79'),
            top=Side(style='medium', color='1F4E79'),
            bottom=Side(style='medium', color='1F4E79')
        )
        for col in ['A', 'B', 'C', 'D']:
            cell = ws_summary[f'{col}1']
            cell.border = title_border
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        # Add company and date info if available
        info_row = 3
        if self.title_info:
            if 'company_name' in self.title_info:
                company_cell = ws_summary[f'A{info_row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="2D3748")
                info_row += 1
            if 'date_range' in self.title_info:
                date_cell = ws_summary[f'A{info_row}']
                date_cell.value = f"Report Period: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="2D3748")
                info_row += 1
        
        # Headers with professional styling
        header_row = info_row + 2
        headers = ["📋 Metric", "📊 Value", "💡 Description", "🎯 Impact"]
        header_colors = ["4A90E2", "27AE60", "FF8C42", "9B59B6"]
        
        for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
            cell = ws_summary.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color='FFFFFF'),
                right=Side(style='thin', color='FFFFFF'),
                top=Side(style='medium', color='2C3E50'),
                bottom=Side(style='medium', color='2C3E50')
            )
        
        # Enhanced summary data with impact analysis
        summary_data = [
            ["Total Loads Processed", f"{self.summary_stats['total_loads']:,}", "Number of shipments analyzed", "📦 Volume"],
            ["Total Selected Carrier Costs", f"${self.summary_stats['total_selected_cost']:,.2f}", "Current transportation spend", "💰 Baseline"],
            ["Total Least Cost Alternative", f"${self.summary_stats['total_least_cost']:,.2f}", "Optimal cost if all loads used cheapest option", "🎯 Target"],
            ["Total Potential Savings", f"${self.summary_stats['total_potential_savings']:,.2f}", "Money that could be saved", "💵 Opportunity"],
            ["Average Savings per Load", f"${self.summary_stats['average_savings_per_load']:,.2f}", "Per-shipment savings potential", "📈 Efficiency"],
            ["Savings Percentage", f"{self.summary_stats['percentage_savings']:.2f}%", "Savings as % of total spend", "📊 Rate"],
            ["Loads with Savings Opportunities", f"{self.summary_stats['loads_with_savings']:,}", "Shipments that could be optimized", "🔍 Focus"],
            ["Actionable Savings Total", f"${self.summary_stats['total_savings_opportunity']:,.2f}", "Realistic savings from optimizable loads", "✅ Achievable"]
        ]
        
        # Add data with alternating colors and professional styling
        data_border = Border(
            left=Side(style='thin', color='E8E8E8'),
            right=Side(style='thin', color='E8E8E8'),
            top=Side(style='thin', color='E8E8E8'),
            bottom=Side(style='thin', color='E8E8E8')
        )
        
        for row_idx, (metric, value, description, impact) in enumerate(summary_data, header_row + 1):
            row_color = "F8F9FA" if row_idx % 2 == 0 else "FFFFFF"
            
            # Metric column
            metric_cell = ws_summary.cell(row=row_idx, column=1, value=metric)
            metric_cell.font = Font(size=11, bold=True, color="2D3748")
            metric_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            metric_cell.border = data_border
            metric_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            # Value column with special formatting for currency and percentages
            value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
            if "$" in str(value):
                value_cell.font = Font(size=11, bold=True, color="27AE60")
            elif "%" in str(value):
                value_cell.font = Font(size=11, bold=True, color="3498DB")
            else:
                value_cell.font = Font(size=11, bold=True, color="E74C3C")
            value_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            value_cell.border = data_border
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Description column
            desc_cell = ws_summary.cell(row=row_idx, column=3, value=description)
            desc_cell.font = Font(size=10, color="4A5568")
            desc_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            desc_cell.border = data_border
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            
            # Impact column
            impact_cell = ws_summary.cell(row=row_idx, column=4, value=impact)
            impact_cell.font = Font(size=10, bold=True, color="7C3AED")
            impact_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            impact_cell.border = data_border
            impact_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit columns on Summary sheet (ensure content fits)
        try:
            for col_idx in range(1, ws_summary.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(1, ws_summary.max_row + 1):
                    val = ws_summary.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws_summary.column_dimensions[col_letter].width = max(8, max_len + 1.5)  # Ensure content fits
            ws_summary.sheet_view.showGridLines = True
        except Exception:
            pass
        
        # Add insights section
        insights_start_row = header_row + len(summary_data) + 3
        
        # Insights title
        insights_title = ws_summary.cell(row=insights_start_row, column=1, value="🎯 KEY INSIGHTS & RECOMMENDATIONS")
        insights_title.font = Font(size=14, bold=True, color="FFFFFF")
        insights_title.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        insights_title.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells(f'A{insights_start_row}:D{insights_start_row}')
        
        # Apply styling to merged insights title
        for col in [1, 2, 3, 4]:
            cell = ws_summary.cell(row=insights_start_row, column=col)
            cell.border = title_border
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        # Add key insights
        insights = [
            f"💡 {self.summary_stats['loads_with_savings']} out of {self.summary_stats['total_loads']} loads ({(self.summary_stats['loads_with_savings']/max(self.summary_stats['total_loads'], 1)*100):.1f}%) have savings opportunities",
            f"🎯 Focus on optimizing the {self.summary_stats['loads_with_savings']} loads with savings potential",
            f"📈 Average savings per optimizable load: ${(self.summary_stats['total_savings_opportunity']/max(self.summary_stats['loads_with_savings'], 1)):,.2f}",
            f"⚡ Quick wins: Target loads with highest individual savings first"
        ]
        
        for i, insight in enumerate(insights, insights_start_row + 1):
            insight_cell = ws_summary.cell(row=i, column=1, value=insight)
            insight_cell.font = Font(size=11, color="2D3748")
            insight_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            ws_summary.merge_cells(f'A{i}:D{i}')
            
            # Apply styling to merged insight cells
            for col in [1, 2, 3, 4]:
                cell = ws_summary.cell(row=i, column=col)
                cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='E2E8F0'),
                    right=Side(style='thin', color='E2E8F0'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='thin', color='E2E8F0')
                )

        # === Move CAL (Column Abbreviations Legend) to Summary sheet ===
        cal_start_row = insights_start_row + len(insights) + 3
        cal_title = ws_summary.cell(row=cal_start_row, column=1, value="📋 COLUMN ABBREVIATIONS LEGEND")
        cal_title.font = Font(size=12, bold=True, color="FFFFFF")
        cal_title.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        cal_title.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells(f'A{cal_start_row}:D{cal_start_row}')
        for col in [1, 2, 3, 4]:
            tcell = ws_summary.cell(row=cal_start_row, column=col)
            tcell.border = Border(
                left=Side(style='medium', color='4A5568'),
                right=Side(style='medium', color='4A5568'),
                top=Side(style='medium', color='4A5568'),
                bottom=Side(style='medium', color='4A5568')
            )

        # Headers for CAL
        cal_hdr_row = cal_start_row + 1
        left_hdr = ws_summary.cell(row=cal_hdr_row, column=1, value="SELECTED CARRIER")
        left_hdr.font = Font(size=10, bold=True, color="FFFFFF")
        left_hdr.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        left_hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells(f'A{cal_hdr_row}:B{cal_hdr_row}')

        right_hdr = ws_summary.cell(row=cal_hdr_row, column=3, value="LEAST COST CARRIER")
        right_hdr.font = Font(size=10, bold=True, color="FFFFFF")
        right_hdr.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
        right_hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.merge_cells(f'C{cal_hdr_row}:D{cal_hdr_row}')

        # CAL entries
        selected_entries = [("SC", "Selected Carrier"),("SST", "Selected Service Type"),("STD", "Selected Transit Days"),("SFF", "Selected Fuel/Fees"),("STA", "Selected Accessorials"),("STC", "Selected Total Cost")]
        least_cost_entries = [("LCC", "Least Cost Carrier"),("LCST", "Least Cost Service Type"),("LCTD", "Least Cost Transit Days"),("LCFF", "Least Cost Fuel/Fees"),("LCTA", "Least Cost Accessorials"),("LCTC", "Least Cost Total Cost")]

        cal_row = cal_hdr_row
        for i in range(max(len(selected_entries), len(least_cost_entries))):
            cal_row += 1
            if i < len(selected_entries):
                abbr, desc = selected_entries[i]
                a = ws_summary.cell(row=cal_row, column=1, value=abbr)
                a.font = Font(size=9, bold=True, color="1A365D")
                a.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                a.alignment = Alignment(horizontal="center", vertical="center")
                b = ws_summary.cell(row=cal_row, column=2, value=desc)
                b.font = Font(size=9, color="2D3748")
                b.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                b.alignment = Alignment(horizontal="left", vertical="center")
            if i < len(least_cost_entries):
                abbr, desc = least_cost_entries[i]
                c = ws_summary.cell(row=cal_row, column=3, value=abbr)
                c.font = Font(size=9, bold=True, color="8B4513")
                c.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")
                d = ws_summary.cell(row=cal_row, column=4, value=desc)
                d.font = Font(size=9, color="2D3748")
                d.fill = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")
                d.alignment = Alignment(horizontal="left", vertical="center")

        # Compact auto-fit for Summary (incl. CAL)
        try:
            for col_idx in range(1, ws_summary.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for r in range(1, ws_summary.max_row + 1):
                    val = ws_summary.cell(row=r, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws_summary.column_dimensions[col_letter].width = max(8, max_len + 1.5)  # Ensure content fits
        except Exception:
            pass
        
        # Auto-fit column widths on the Processed Data sheet (compact and consistent for all: table, CAL, PI)
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(5, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value)
                        length = len(text)
                        if length > max_length:
                            max_length = length
                # Aggressive compact auto-sizing: minimize width while ensuring content fits
                # Check if this column contains text content
                has_text_content = False
                for check_row in range(5, ws_data.max_row + 1):
                    check_cell = ws_data.cell(row=check_row, column=col_idx)
                    if check_cell.value and any(c.isalpha() for c in str(check_cell.value)):
                        has_text_content = True
                        break
                
                # Very tight padding for maximum compactness
                if has_text_content:
                    padding = 1.0  # Minimal padding for text
                else:
                    padding = 0.5   # Very tight for numbers
                
                # Calculate width with aggressive compacting
                adjusted_width = max_length + padding
                
                # Apply maximum width constraints for compactness
                if has_text_content:
                    max_width = 25  # Cap text columns at 25 characters
                else:
                    max_width = 15  # Cap number columns at 15 characters
                
                final_width = min(adjusted_width, max_width)
                ws_data.column_dimensions[col_letter].width = max(6, final_width)  # Minimum 6 for readability
            
            # Optimize row heights for compact layout
            for rh in [1, 2, 4, 5]:
                if rh <= ws_data.max_row:
                    ws_data.row_dimensions[rh].height = max(ws_data.row_dimensions[rh].height or 0, 20)  # Reduced from 22 to 20
            
            # Ensure gridlines visible
            ws_data.sheet_view.showGridLines = True
        except Exception:
            pass

        # Add an outside border around the data table area
        try:
            thin_side = Side(style='thin', color='A0AEC0')
            outer_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
            header_row_idx = 5
            first_row = header_row_idx
            last_row = row
            first_col = 1
            last_col = len(headers)
            # Top and bottom edges
            for c in range(first_col, last_col + 1):
                ws_data.cell(row=first_row, column=c).border = outer_border
                ws_data.cell(row=last_row, column=c).border = outer_border
            # Left and right edges
            for r in range(first_row, last_row + 1):
                ws_data.cell(row=r, column=first_col).border = outer_border
                ws_data.cell(row=r, column=last_col).border = outer_border
        except Exception:
            pass

        wb.save(output_file)
        wb.close()

class ModernTMSProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("TMS Data Processor Pro")
        self.root.geometry("1000x750")
        self.root.configure(bg='#f8f9fa')
        self.root.minsize(900, 650)
        
        # Initialize processors
        self.basic_processor = ModernTMSProcessor()
        self.detailed_processor = None
        self.input_file = None
        self.output_file = None
        
        # Configure style
        self.setup_styles()
        
        # Create GUI
        self.create_widgets()
        
        # Center window
        self.center_window()
        
    def setup_styles(self):
        """Setup modern styling for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors with modern palette
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 28, 'bold'), 
                       foreground='#1a365d',
                       background='#f8f9fa')
        style.configure('Subtitle.TLabel', 
                       font=('Segoe UI', 12), 
                       foreground='#4a5568',
                       background='#f8f9fa')
        style.configure('Header.TLabel', 
                       font=('Segoe UI', 13, 'bold'), 
                       foreground='#2d3748',
                       background='#f8f9fa')
        style.configure('Info.TLabel', 
                       font=('Segoe UI', 10), 
                       foreground='#718096',
                       background='#f8f9fa')
        style.configure('Success.TLabel', 
                       font=('Segoe UI', 10, 'bold'), 
                       foreground='#38a169',
                       background='#f8f9fa')
        
        # Configure modern buttons with hover effects
        style.configure('Primary.TButton', 
                       font=('Segoe UI', 11, 'bold'),
                       background='#4299e1',
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(20, 10))
        style.map('Primary.TButton',
                 background=[('active', '#3182ce'), ('pressed', '#2c5aa0')])
        
        style.configure('Success.TButton',
                       font=('Segoe UI', 12, 'bold'),
                       background='#48bb78',
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(25, 12))
        style.map('Success.TButton',
                 background=[('active', '#38a169'), ('pressed', '#2f855a')])
        
        style.configure('Browse.TButton',
                       font=('Segoe UI', 10),
                       background='#667eea',
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(15, 8))
        style.map('Browse.TButton',
                 background=[('active', '#5a67d8'), ('pressed', '#4c51bf')])
        
        # Configure radio buttons
        style.configure('Modern.TRadiobutton',
                       font=('Segoe UI', 11),
                       foreground='#2d3748',
                       background='#f8f9fa',
                       focuscolor='none')
        
        # Configure frames
        style.configure('Card.TFrame',
                       background='#ffffff',
                       relief='flat',
                       borderwidth=1)
        

        
    def create_widgets(self):
        """Create the main GUI widgets"""
        # Main container with modern styling
        main_frame = ttk.Frame(self.root, padding="30", style='Card.TFrame')
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=20, pady=20)
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Header section with gradient-like background
        header_frame = tk.Frame(main_frame, bg='#ffffff', relief='flat')
        header_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 30))
        header_frame.columnconfigure(0, weight=1)
        
        # Title with modern styling
        title_label = ttk.Label(header_frame, text="🚛 TMS Data Processor Pro", style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(10, 5))
        
        # Subtitle with better spacing
        subtitle_label = ttk.Label(header_frame, 
                                 text="Transform your TMS Excel reports into professional, actionable insights",
                                 style='Subtitle.TLabel')
        subtitle_label.grid(row=1, column=0, pady=(0, 15))
        
        # Version badge
        version_label = ttk.Label(header_frame, text="v2.0 Pro", 
                                font=('Segoe UI', 9, 'bold'), 
                                foreground='#667eea',
                                background='#ffffff')
        version_label.grid(row=2, column=0, pady=(0, 10))
        
        # Report Type Selection with modern card design
        report_section = tk.Frame(main_frame, bg='#f7fafc', relief='flat', bd=1)
        report_section.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 25), padx=5)
        report_section.columnconfigure(1, weight=1)
        
        ttk.Label(report_section, text="📋 Report Type:", style='Header.TLabel', background='#f7fafc').grid(row=0, column=0, sticky=tk.W, padx=20, pady=(15, 10))
        
        self.report_type = tk.StringVar(value="basic")
        report_frame = tk.Frame(report_section, bg='#f7fafc')
        report_frame.grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), padx=20, pady=(15, 15))
        
        # Modern radio buttons with better styling
        basic_radio = ttk.Radiobutton(report_frame, text="📊 Basic Report", 
                                     variable=self.report_type, value="basic",
                                     style='Modern.TRadiobutton')
        basic_radio.grid(row=0, column=0, padx=(0, 30), sticky=tk.W)
        
        detailed_radio = ttk.Radiobutton(report_frame, text="📈 Detailed Report", 
                                       variable=self.report_type, value="detailed",
                                       style='Modern.TRadiobutton')
        detailed_radio.grid(row=0, column=1, sticky=tk.W)
        
        # Add description labels with enhanced text
        basic_desc = ttk.Label(report_frame, text="Standard processing with color-coded sections & legend", 
                              font=('Segoe UI', 9), foreground='#718096', background='#f7fafc')
        basic_desc.grid(row=1, column=0, padx=(20, 30), pady=(2, 0), sticky=tk.W)
        
        detailed_desc = ttk.Label(report_frame, text="Enhanced processing with advanced analytics", 
                                 font=('Segoe UI', 9), foreground='#718096', background='#f7fafc')
        detailed_desc.grid(row=1, column=1, padx=(20, 0), pady=(2, 0), sticky=tk.W)
        
        # File Selection with modern card design
        file_section = tk.Frame(main_frame, bg='#f0fff4', relief='flat', bd=1)
        file_section.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 25), padx=5)
        file_section.columnconfigure(1, weight=1)
        
        ttk.Label(file_section, text="📁 Input File:", style='Header.TLabel', background='#f0fff4').grid(row=0, column=0, sticky=tk.W, padx=20, pady=(15, 10))
        
        file_frame = tk.Frame(file_section, bg='#f0fff4')
        file_frame.grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), padx=20, pady=(15, 15))
        file_frame.columnconfigure(0, weight=1)
        
        # File display with better styling
        file_display_frame = tk.Frame(file_frame, bg='#ffffff', relief='solid', bd=1)
        file_display_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 15))
        file_display_frame.columnconfigure(0, weight=1)
        
        self.file_label = ttk.Label(file_display_frame, text="No file selected", 
                                   style='Info.TLabel', background='#ffffff')
        self.file_label.grid(row=0, column=0, sticky=tk.W, padx=15, pady=10)
        
        browse_button = ttk.Button(file_frame, text="📂 Browse", 
                                 command=self.browse_file, style='Browse.TButton')
        browse_button.grid(row=0, column=1)
        
        # Process Button with enhanced styling
        button_frame = tk.Frame(main_frame, bg='#ffffff')
        button_frame.grid(row=3, column=0, columnspan=3, pady=30)
        
        self.process_button = ttk.Button(button_frame, text="🚀 Process File", 
                                       command=self.process_file, style='Success.TButton', state="disabled")
        self.process_button.grid(row=0, column=0)
        
        # Status Display with modern terminal-like design
        status_section = tk.Frame(main_frame, bg='#ffffff')
        status_section.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        status_section.columnconfigure(0, weight=1)
        status_section.rowconfigure(1, weight=1)
        
        # Status header with icon
        status_header = tk.Frame(status_section, bg='#2d3748', height=35)
        status_header.grid(row=0, column=0, sticky=(tk.W, tk.E))
        status_header.grid_propagate(False)
        
        ttk.Label(status_header, text="💻 Processing Status", 
                 font=('Segoe UI', 11, 'bold'), 
                 foreground='#ffffff', background='#2d3748').grid(row=0, column=0, padx=15, pady=8, sticky=tk.W)
        
        # Create status text with scrollbar and modern styling
        status_frame = tk.Frame(status_section, bg='#1a202c', relief='flat')
        status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        status_frame.columnconfigure(0, weight=1)
        status_frame.rowconfigure(0, weight=1)
        
        self.status_text = tk.Text(status_frame, height=16, width=80, 
                                  font=('Consolas', 10), bg='#1a202c', fg='#e2e8f0',
                                  insertbackground='#4299e1', selectbackground='#4a5568',
                                  relief='flat', bd=0, padx=15, pady=10)
        status_scrollbar = ttk.Scrollbar(status_frame, orient=tk.VERTICAL, command=self.status_text.yview)
        self.status_text.configure(yscrollcommand=status_scrollbar.set)
        
        self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        status_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        main_frame.rowconfigure(4, weight=1)
        
        # Welcome message
        self.log_message("🎉 Welcome to TMS Data Processor Pro v2.0!")
        self.log_message("📋 Select your report type and input file to get started")
        self.log_message("")
        
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def browse_file(self):
        """Browse for input file"""
        file_path = filedialog.askopenfilename(
            title="Select TMS Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.input_file = file_path
            filename = os.path.basename(file_path)
            self.file_label.config(text=f"✅ {filename}", foreground='#38a169')
            self.update_process_button_state()
            self.log_message(f"📁 Selected file: {filename}")
    
    def update_process_button_state(self):
        """Enable process button if input file is selected"""
        if self.input_file:
            self.process_button.config(state="normal")
        else:
            self.process_button.config(state="disabled")
    
    def log_message(self, message):
        """Add message to status text with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
    
    def process_file(self):
        """Process the selected file in a separate thread"""
        if not self.input_file:
            return
            
        # Disable UI during processing
        self.process_button.config(state="disabled")
        
        # Start processing in separate thread
        thread = threading.Thread(target=self._process_file_thread)
        thread.daemon = True
        thread.start()
        
    def _process_file_thread(self):
        """Process file in background thread"""
        try:
            # Ask user where to save the file
            output_file = filedialog.asksaveasfilename(
                title="Save Processed File As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not output_file:
                self.root.after(0, self._reset_ui)
                return
            
            self.root.after(0, lambda: self.log_message("🔄 Starting file processing..."))
            
            # Select processor based on report type
            if self.report_type.get() == "basic":
                processor = self.basic_processor
                self.root.after(0, lambda: self.log_message("📊 Processing BASIC report..."))
            else:
                # Import detailed processor when needed
                if self.detailed_processor is None:
                    try:
                        from tms_detailed_processor import TMSDetailedDataProcessor
                        self.detailed_processor = TMSDetailedDataProcessor()
                        self.root.after(0, lambda: self.log_message("📈 Detailed processor loaded successfully"))
                    except ImportError as e:
                        self.root.after(0, lambda: self.log_message(f"❌ ERROR: Failed to load detailed processor: {e}"))
                        self.root.after(0, self._reset_ui)
                        return
                processor = self.detailed_processor
                self.root.after(0, lambda: self.log_message("📈 Processing DETAILED report..."))
            
            # Process the data
            self.root.after(0, lambda: self.log_message("📖 Reading and cleaning raw data..."))
            processed_data = processor.clean_and_process_data(self.input_file)
            
            self.root.after(0, lambda: self.log_message(f"✅ Successfully processed {len(processed_data)} records"))
            
            # Display summary statistics
            stats = processor.summary_stats
            self.root.after(0, lambda: self.log_message("\n📊 === SUMMARY STATISTICS ==="))
            self.root.after(0, lambda: self.log_message(f"📦 Total Loads: {stats['total_loads']}"))
            self.root.after(0, lambda: self.log_message(f"💰 Total Selected Cost: ${stats['total_selected_cost']:,.2f}"))
            self.root.after(0, lambda: self.log_message(f"💡 Total Least Cost: ${stats['total_least_cost']:,.2f}"))
            self.root.after(0, lambda: self.log_message(f"💵 Total Potential Savings: ${stats['total_potential_savings']:,.2f}"))
            self.root.after(0, lambda: self.log_message(f"📈 Average Savings per Load: ${stats['average_savings_per_load']:,.2f}"))
            self.root.after(0, lambda: self.log_message(f"🎯 Percentage Savings: {stats['percentage_savings']:.2f}%"))
            self.root.after(0, lambda: self.log_message(f"🚀 Loads with Savings: {stats['loads_with_savings']}"))
            
            # Save the processed data
            self.root.after(0, lambda: self.log_message("💾 Saving processed data..."))
            processor.save_processed_data(output_file)
            
            self.root.after(0, lambda: self.log_message(f"✅ File successfully saved to: {os.path.basename(output_file)}"))
            self.root.after(0, lambda: self.log_message("🎉 Processing completed successfully!"))
            
            # Show success message
            self.root.after(0, lambda: messagebox.showinfo("Success", 
                f"🎉 File processed successfully!\n\n"
                f"📊 Processed: {stats['total_loads']} loads\n"
                f"💵 Potential Savings: ${stats['total_potential_savings']:,.2f}\n"
                f"📈 Savings Rate: {stats['percentage_savings']:.2f}%\n"
                f"🎯 Optimizable Loads: {stats['loads_with_savings']}\n\n"
                f"📁 File saved to:\n{output_file}"))
            
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda: self.log_message(f"❌ ERROR: {error_msg}"))
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred:\n{error_msg}"))
        
        finally:
            self.root.after(0, self._reset_ui)
            
    def _reset_ui(self):
        """Reset UI to normal state"""
        self.process_button.config(state="normal")

def main():
    root = tk.Tk()
    app = ModernTMSProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()