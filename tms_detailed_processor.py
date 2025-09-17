"""
TMS Detailed Data Processor
Enhanced processor for TMS Detailed reports with 27 columns and special handling for W/Z column quirks
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple, Any
import time
from pathlib import Path

# Import our configuration and logging modules
try:
    from config import tms_config
    from logger_config import main_logger, data_logger, gui_logger, ProgressLogger
    from validators import tms_validator, tms_cleaner
except ImportError as e:
    print(f"Warning: Could not import enhanced modules: {e}")
    # Create mock objects for backward compatibility
    class MockConfig:
        def get(self, key, default=None):
            defaults = {
                'data_structure.default_header_row': 8,
                'data_structure.default_data_start_row': 9,
                'data_structure.min_data_columns': 5,
                'data_structure.expected_columns': 27,  # Detailed reports have 27 columns
                'business_rules.min_non_empty_values': 5,
                'formatting.date_format': '%m/%d/%y'
            }
            return defaults.get(key, default)
    
    class MockLogger:
        def info(self, msg, **kwargs): print(f"INFO: {msg}")
        def error(self, msg, **kwargs): print(f"ERROR: {msg}")
        def warning(self, msg, **kwargs): print(f"WARNING: {msg}")
        def debug(self, msg, **kwargs): print(f"DEBUG: {msg}")
        def log_processing_step(self, step, details=None): self.info(f"Processing: {step}")
        def log_data_stats(self, stats, prefix=""): self.info(f"{prefix}: {stats}")
        def log_performance(self, op, duration, records=None): self.info(f"{op}: {duration:.2f}s")
    
    class MockValidator:
        def run_full_validation(self, file_path): return {'overall_valid': True, 'validation_steps': {}}
    
    tms_config = MockConfig()
    main_logger = data_logger = gui_logger = MockLogger()
    tms_validator = MockValidator()
    
    def ProgressLogger(logger, total, operation):
        class MockProgress:
            def update(self, inc=1): pass
            def complete(self): pass
        return MockProgress()


class TMSDetailedDataProcessor:
    """
    Enhanced TMS Processor specifically designed for Detailed reports
    
    Key Differences from Basic Reports:
    - 27 columns instead of ~21
    - Contains "Origin Name" and "Created By" columns
    - Has empty W (col 22) and Z (col 25) columns that need special handling
    - Different column mapping and processing rules
    """
    
    def __init__(self):
        self.logger = main_logger
        self.data_logger = data_logger
        self.config = tms_config
        
        # Data storage
        self.raw_data = None
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}
        self.validation_results = None
        
        # Performance tracking
        self.processing_start_time = None
        self.processing_stats = {}
        
        # Detailed report specific settings
        self.EXPECTED_COLUMNS = 27
        self.DEFAULT_HEADER_ROW = 8   # Row 9 in Excel (column headers)
        self.DEFAULT_DATA_START_ROW = 9  # Row 10 in Excel (first data row)

        # Carrier lists for special processing
        # TL carriers that require copy-paste and zero-out logic
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'SMARTWAY CORPORATION INC'
        }

        self.logger.info("TMSDetailedDataProcessor initialized for detailed reports with 27 columns")
        
    def _extract_title_info(self, df_raw):
        """Extract title and report information from the top rows"""
        title_info = {}
        
        try:
            # Extract report title with bounds checking
            if len(df_raw) > 1 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[1, 1]):
                title_info['report_title'] = str(df_raw.iloc[1, 1])
            
            # Extract company name
            if len(df_raw) > 3 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[3, 1]):
                title_info['company_name'] = str(df_raw.iloc[3, 1])
            
            # Extract date range
            if len(df_raw) > 5 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[5, 1]):
                title_info['date_range'] = str(df_raw.iloc[5, 1])
        except (IndexError, KeyError):
            # If extraction fails, continue with empty title_info
            pass
        
        return title_info
    
    def _detect_data_structure(self, df_raw):
        """Intelligently detect header and data start positions for detailed reports"""
        # For detailed reports, headers span two rows: 8 (section) + 9 (columns)
        header_row = 8  # Row 9 in Excel (0-based index 8)
        data_start_row = 9  # Row 10 in Excel (0-based index 9)

        # Verify the structure by checking for key indicators
        try:
            # Check row 8 for section headers
            section_row = df_raw.iloc[7]  # Row 8 in Excel
            column_row = df_raw.iloc[8]   # Row 9 in Excel

            # Look for section indicators in row 8
            section_indicators = ['Selected Carrier', 'Least Cost Carrier']
            section_found = any(str(val) in str(section_row.values) for val in section_indicators)

            # Look for column indicators in row 9
            column_indicators = ['Load No.', 'Origin Name', 'Created By', 'Carrier']
            column_found = any(str(val) in str(column_row.values) for val in column_indicators)

            if not (section_found and column_found):
                self.logger.warning("Expected two-row header structure not found, using fallback detection")
                # Fallback to original logic
                for row_idx in range(5, min(15, len(df_raw))):
                    row_data = df_raw.iloc[row_idx].dropna().astype(str).tolist()
                    row_str = ' '.join(row_data).lower()

                    matches = sum(1 for indicator in column_indicators if indicator.lower() in row_str)
                    if matches >= 3:
                        header_row = row_idx
                        data_start_row = row_idx + 1
                        break

        except (IndexError, KeyError):
            self.logger.warning("Error detecting header structure, using defaults")

        return header_row, data_start_row
    
    def _remove_duplicate_headers(self, df):
        """Remove duplicate header rows that appear in the middle of data"""
        # Look for rows that contain header-like text
        header_indicators = ['Load No.', 'Origin Name', 'Created By', 'Carrier', 'Service Type']
        
        rows_to_drop = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.dropna().tolist()])
            if any(indicator in row_str for indicator in header_indicators):
                # Check if this looks like a header row (not actual data)
                if not any(str(val).startswith('A') and str(val)[1:].isdigit() for val in row.dropna().tolist()):
                    rows_to_drop.append(idx)
        
        return df.drop(rows_to_drop)
    
    def clean_and_process_data(self, file_path: str) -> pd.DataFrame:
        """Main function to clean and process the TMS Detailed Excel file"""
        self.processing_start_time = time.time()
        self.logger.log_processing_step("Starting TMS detailed data processing", {'file': Path(file_path).name})
        
        try:
            # Step 1: Comprehensive validation
            self.validation_results = tms_validator.run_full_validation(file_path)
            
            if not self.validation_results['overall_valid']:
                failed_steps = [step for step, result in self.validation_results['validation_steps'].items() 
                               if not result.get('valid', False)]
                self.logger.warning(f"Validation issues detected: {failed_steps}")
            
            # Step 2: Load and validate Excel data
            self.logger.log_processing_step("Loading detailed Excel file")
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
            
            self.logger.log_data_stats({
                'raw_rows': len(df_raw),
                'raw_columns': len(df_raw.columns),
                'expected_columns': self.EXPECTED_COLUMNS,
                'file_size_mb': Path(file_path).stat().st_size / (1024*1024)
            }, "RAW_DETAILED_DATA")
            
            # Verify we have expected number of columns
            if len(df_raw.columns) != self.EXPECTED_COLUMNS:
                self.logger.warning(f"Expected {self.EXPECTED_COLUMNS} columns, got {len(df_raw.columns)}")
            
            # Step 3: Extract metadata
            self.logger.log_processing_step("Extracting title information")
            self.title_info = self._extract_title_info(df_raw)
            
            # Step 4: Detect structure
            header_row, data_start_row = self._detect_data_structure(df_raw)
            
            self.logger.log_processing_step("Data structure detected", {
                'header_row': header_row,
                'data_start_row': data_start_row,
                'report_type': 'detailed'
            })
            
            # Step 5: Extract and clean data
            self.logger.log_processing_step("Extracting data rows from detailed report")
            data_df = df_raw.iloc[data_start_row:].copy()
            
            # Remove completely empty rows and duplicate header rows
            initial_rows = len(data_df)
            data_df = data_df.dropna(how='all')
            empty_rows_removed = initial_rows - len(data_df)
            
            data_df = self._remove_duplicate_headers(data_df)
            duplicate_headers_removed = initial_rows - empty_rows_removed - len(data_df)
            
            self.logger.log_data_stats({
                'initial_data_rows': initial_rows,
                'empty_rows_removed': empty_rows_removed,
                'duplicate_headers_removed': duplicate_headers_removed,
                'remaining_rows': len(data_df)
            }, "DETAILED_DATA_CLEANING")
            
            # Reset index after dropping rows
            data_df = data_df.reset_index(drop=True)
            
            # Step 6: Handle the detailed report column extraction
            # Skip columns A and B (0,1), extract from C onwards
            # Also need to handle the W and Z empty columns (22, 25)
            self.logger.log_processing_step("Extracting relevant columns from detailed report")
            
            # Extract columns C through AA (2 through 26) - skip the empty A and B columns
            relevant_columns = list(range(2, min(self.EXPECTED_COLUMNS, len(data_df.columns))))


            data_df = data_df.iloc[:, relevant_columns]
            
            # Step 7: Set proper column names based on actual Excel structure
            # Map columns based on analysis: C=Load No., D=Ship Date, etc.
            # Columns extracted: 2-26 (C through AA, skipping empty columns)
            detailed_column_names = [
                'Load No.',                    # C (index 2)
                'Ship Date',                   # D (index 3)
                'Origin Name',                 # E (index 4)
                'Origin City',                 # F (index 5)
                'Origin State',               # G (index 6)
                'Origin Postal',              # H (index 7)
                'Destination City',           # I (index 8)
                'Destination State',          # J (index 9)
                'Destination Postal',         # K (index 10)
                'Created By',                 # L (index 11)
                'Selected Carrier',           # M (index 12)
                'Selected Service Type',      # N (index 13)
                'Selected Transit Days',      # O (index 14)
                'Selected Freight Cost',      # P (index 15)
                'Selected Accessorial Cost',  # Q (index 16)
                'Selected Total Cost',        # R (index 17)
                'Least Cost Carrier',         # S (index 18)
                'Least Cost Service Type',    # T (index 19)
                'Least Cost Transit Days',    # U (index 20)
                'Least Cost Freight Cost',    # V (index 21)
                'Empty_W_Column',            # W (index 22) - empty
                'Least Cost Accessorial Cost', # X (index 23)
                'Least Cost Total Cost',      # Y (index 24)
                'Empty_Z_Column',            # Z (index 25) - empty
                'Potential Savings'           # AA (index 26)
            ]

            # Ensure column names match the actual extracted columns
            if len(detailed_column_names) != len(data_df.columns):
                self.logger.warning(f"Column count mismatch. Expected {len(detailed_column_names)}, got {len(data_df.columns)}")
                # Adjust column names to match actual columns
                if len(data_df.columns) < len(detailed_column_names):
                    column_names = detailed_column_names[:len(data_df.columns)]
                else:
                    # Add generic names for extra columns
                    extra_columns = [f'Extra_Column_{i}' for i in range(len(detailed_column_names), len(data_df.columns))]
                    column_names = detailed_column_names + extra_columns
            else:
                column_names = detailed_column_names

            data_df.columns = column_names



            # Step 8: Remove the problematic W and Z columns that are always empty
            self.logger.log_processing_step("Removing empty W and Z columns")
            columns_to_drop = ['Empty_W_Column', 'Empty_Z_Column']
            
            for col in columns_to_drop:
                if col in data_df.columns:
                    # Log what's in these columns before dropping
                    non_null_count = data_df[col].count()
                    if non_null_count > 0:
                        self.logger.warning(f"Column {col} has {non_null_count} non-null values - review before dropping")
                    data_df = data_df.drop(columns=[col])
                    self.logger.info(f"Dropped problematic column: {col}")
            
            self.logger.log_data_stats({
                'columns_after_extraction': len(data_df.columns),
                'columns_dropped': len(columns_to_drop),
                'final_column_count': len(data_df.columns)
            }, "COLUMN_PROCESSING")
            
            # Step 9: Enhanced data type cleaning
            self.logger.log_processing_step("Cleaning and validating data types for detailed report")
            cleaning_start = time.time()
            data_df = self._clean_data_types_detailed(data_df)
            cleaning_time = time.time() - cleaning_start
            self.logger.log_performance("Detailed data type cleaning", cleaning_time, len(data_df))
            
            # Step 10: Enhanced row filtering
            self.logger.log_processing_step("Filtering invalid rows")
            pre_filter_count = len(data_df)
            
            # Remove rows where Load No. is missing or empty
            if 'Load No.' in data_df.columns:
                data_df = data_df.dropna(subset=['Load No.'])
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != '']
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != 'nan']
            else:
                self.logger.warning("Load No. column not found - skipping Load No. validation")
            
            # Remove any remaining rows that are mostly empty
            min_values = self.config.get('business_rules.min_non_empty_values', 5)
            data_df = data_df.dropna(thresh=min_values)
            
            # Reset index and log filtering results
            data_df = data_df.reset_index(drop=True)
            rows_filtered = pre_filter_count - len(data_df)
            
            self.logger.log_data_stats({
                'rows_before_filtering': pre_filter_count,
                'rows_after_filtering': len(data_df),
                'rows_removed': rows_filtered,
                'filter_rate': f"{(rows_filtered/pre_filter_count*100):.1f}%" if pre_filter_count > 0 else "0%"
            }, "DETAILED_ROW_FILTERING")
            
            # Step 11: Apply business logic rules
            self.logger.log_processing_step("Applying business logic rules for detailed report")
            business_start = time.time()
            data_df = self._apply_business_logic_detailed(data_df)
            business_time = time.time() - business_start
            self.logger.log_performance("Detailed business logic application", business_time, len(data_df))
            
            # Step 12: Sort by Destination City
            self.logger.log_processing_step("Sorting data")
            if 'Destination City' in data_df.columns:
                data_df = data_df.sort_values('Destination City', na_position='last')
            else:
                self.logger.warning("Destination City column not found - skipping sort")
            
            # Step 13: Calculate summary statistics
            self.logger.log_processing_step("Calculating summary statistics")
            self._calculate_summary_stats(data_df)
            
            # Step 14: Final processing metrics
            processing_time = time.time() - self.processing_start_time
            self.processing_stats = {
                'total_time': processing_time,
                'records_processed': len(data_df),
                'processing_rate': len(data_df) / processing_time if processing_time > 0 else 0,
                'report_type': 'detailed'
            }
            
            self.logger.log_performance(
                "Total detailed TMS processing", 
                processing_time, 
                len(data_df)
            )
            
            self.processed_data = data_df
            return data_df
            
        except Exception as e:
            self.logger.error("Detailed processing error", exception=e, file_path=file_path)
            raise RuntimeError(f"Error processing detailed file: {str(e)}")
    
    def _clean_data_types_detailed(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhanced data type cleaning for detailed reports"""
        df = df.copy()
        cleaning_stats = {'columns_processed': 0, 'conversion_failures': 0}
        
        # Convert numeric columns with enhanced error tracking
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1
                original_nulls = df[col].isnull().sum()

                # Enhanced cleaning for numeric columns
                col_series = df[col].copy()
                if col_series.dtype == 'object':
                    # Clean string representations
                    col_series = col_series.astype(str)
                    col_series = col_series.str.replace('[$,]', '', regex=True)
                    col_series = col_series.str.strip()
                    col_series = col_series.replace(['nan', 'None', '', 'N/A', '#N/A'], '0')

                df[col] = pd.to_numeric(col_series, errors='coerce').fillna(0)
                new_nulls = df[col].isnull().sum()
                conversion_failures = new_nulls - original_nulls
                if conversion_failures > 0:
                    cleaning_stats['conversion_failures'] += conversion_failures
                    self.data_logger.warning(f"Failed to convert {conversion_failures} values in {col} to numeric")
        
        # Convert date column
        if 'Ship Date' in df.columns:
            cleaning_stats['columns_processed'] += 1
            original_nulls = df['Ship Date'].isnull().sum()
            date_series = pd.to_datetime(df['Ship Date'], errors='coerce')
            new_nulls = date_series.isnull().sum()
            date_failures = new_nulls - original_nulls
            if date_failures > 0:
                cleaning_stats['conversion_failures'] += date_failures
                self.data_logger.warning(f"Failed to convert {date_failures} date values in Ship Date")
            
            date_format = self.config.get('formatting.date_format', '%m/%d/%y')
            df['Ship Date'] = date_series.dt.strftime(date_format)
        
        # Clean string columns (including detailed report specific ones)
        string_columns = [
            'Load No.', 'Origin Name', 'Origin City', 'Origin State', 'Origin Postal',
            'Destination City', 'Destination State', 'Destination Postal', 'Created By',
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]
        
        for col in string_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1

                # Special handling for Least Cost Carrier and Service Type
                if col in ['Least Cost Carrier', 'Least Cost Service Type']:
                    # Check for numeric zeros that should be converted to empty strings
                    before_cleaning = df[col].copy()
                    df[col] = df[col].astype(str).str.strip()

                    # Only convert actual zeros and NaN to empty strings, keep real carrier names
                    df[col] = df[col].replace(['0', '0.0', 'nan', 'None'], '')

                    # Log the cleaning results for debugging
                    zero_count = (before_cleaning == 0).sum()
                    nan_count = before_cleaning.isna().sum()
                    if zero_count > 0 or nan_count > 0:
                        self.data_logger.info(f"Column {col}: Found {zero_count} zeros and {nan_count} NaN values - converted to empty strings")
                else:
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace('nan', '')
        

        # Calculate Potential Savings = Selected Total Cost - Least Cost Total Cost
        if 'Selected Total Cost' in df.columns and 'Least Cost Total Cost' in df.columns:
            selected_costs = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0)
            least_costs = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0)

            # Calculate savings
            calculated_savings = selected_costs - least_costs

            # Only update if Potential Savings column exists
            if 'Potential Savings' in df.columns:
                # Update with calculated values
                df['Potential Savings'] = calculated_savings
                self.data_logger.info(f"Recalculated Potential Savings for {len(df)} rows")
            else:
                # Add the column if it doesn't exist
                df['Potential Savings'] = calculated_savings
                self.data_logger.info(f"Added calculated Potential Savings column for {len(df)} rows")

        self.data_logger.log_data_stats(cleaning_stats, "DETAILED_TYPE_CLEANING")
        return df
    
    def _apply_business_logic_detailed(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply TMS business logic rules for detailed reports"""
        df = df.copy()
        business_stats = {
            'same_carrier_rule_applied': 0,
            'empty_data_rule_applied': 0,
            'negative_savings_rule_applied': 0,
            'tl_carrier_rule_applied': 0,
            'ddi_carrier_rule_applied': 0,
            'total_rows_affected': 0
        }
        
        try:
            # Rule 1: Same Carriers - Set Potential Savings to 0 (but don't copy data)
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                same_carrier_mask = (
                    (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                    (df['Selected Carrier'].notna()) &
                    (df['Least Cost Carrier'].notna()) &
                    (df['Selected Carrier'].astype(str) != '') &
                    (df['Least Cost Carrier'].astype(str) != '') &
                    (df['Selected Carrier'].astype(str) != 'nan') &
                    (df['Least Cost Carrier'].astype(str) != 'nan') &
                    (df['Selected Carrier'].astype(str) != '0') &
                    (df['Least Cost Carrier'].astype(str) != '0')
                )

                same_carrier_count = same_carrier_mask.sum()
                business_stats['same_carrier_rule_applied'] = same_carrier_count

                if 'Potential Savings' in df.columns and same_carrier_count > 0:
                    df.loc[same_carrier_mask, 'Potential Savings'] = 0.0
                    self.data_logger.info(f"Applied same carrier rule to {same_carrier_count} rows")
            
            # Rule 2: Empty Least Cost - Copy Selected data and set savings to 0
            if 'Least Cost Carrier' in df.columns:
                empty_least_cost_mask = (
                    df['Least Cost Carrier'].isna() | 
                    (df['Least Cost Carrier'].astype(str) == '') |
                    (df['Least Cost Carrier'].astype(str) == 'nan')
                )
                
                empty_count = empty_least_cost_mask.sum()
                business_stats['empty_data_rule_applied'] = empty_count
                
                if empty_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, empty_least_cost_mask, column_pairs)
                    
                    if 'Potential Savings' in df.columns:
                        df.loc[empty_least_cost_mask, 'Potential Savings'] = 0
                    
                    self.data_logger.info(f"Applied empty data rule to {empty_count} rows")

            # Rule 3: Negative Savings - Copy Selected data and set savings to 0
            if 'Potential Savings' in df.columns:
                ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
                negative_savings_mask = ps_numeric < 0
                negative_count = negative_savings_mask.sum()
                business_stats['negative_savings_rule_applied'] = negative_count
                
                if negative_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, negative_savings_mask, column_pairs)
                    df.loc[negative_savings_mask, 'Potential Savings'] = 0
                    self.data_logger.info(f"Applied negative savings rule to {negative_count} rows")

            # Rule 4: TL Carriers - Copy selected to least cost and zero out savings
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Find rows where Selected Carrier or Least Cost Carrier is in TL list
                tl_mask = (
                    df['Selected Carrier'].astype(str).str.upper().isin([carrier.upper() for carrier in self.TL_CARRIERS]) |
                    df['Least Cost Carrier'].astype(str).str.upper().isin([carrier.upper() for carrier in self.TL_CARRIERS])
                )

                tl_count = tl_mask.sum()
                business_stats['tl_carrier_rule_applied'] = tl_count

                if tl_count > 0:
                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, tl_mask, column_pairs)

                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        df.loc[tl_mask, 'Potential Savings'] = 0

                    self.data_logger.info(f"Applied TL carrier rule to {tl_count} rows (LANDSTAR/SMARTWAY)")

            # Rule 5: DDI/Carrier Matching - New custom rule for detailed reports
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Create mask for rows where Selected Carrier contains "DDI/" or similar patterns
                # and the part after "/" matches Least Cost Carrier
                ddi_matches = []
                
                for idx, row in df.iterrows():
                    selected = str(row['Selected Carrier']).strip()
                    least_cost = str(row['Least Cost Carrier']).strip()
                    
                    # Skip empty or nan values
                    if selected in ['', 'nan', 'None'] or least_cost in ['', 'nan', 'None']:
                        continue
                    
                    # Check if selected carrier has "/" and extract the part after it
                    if '/' in selected:
                        # Split on "/" and get the part after the last "/"
                        carrier_after_slash = selected.split('/')[-1].strip()
                        
                        # Check if the carrier after "/" matches the least cost carrier
                        # Using case-insensitive comparison and handling common variations
                        if carrier_after_slash.upper() == least_cost.upper():
                            ddi_matches.append(idx)
                        # Also check for R&L Carriers vs R%L Carriers variations
                        elif (carrier_after_slash.upper().replace('&', '%') == least_cost.upper().replace('&', '%') or
                              carrier_after_slash.upper().replace('%', '&') == least_cost.upper().replace('%', '&')):
                            ddi_matches.append(idx)
                
                ddi_match_count = len(ddi_matches)
                business_stats['ddi_carrier_rule_applied'] = ddi_match_count
                
                if ddi_match_count > 0:
                    ddi_mask = df.index.isin(ddi_matches)
                    
                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, ddi_mask, column_pairs)
                    
                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        df.loc[ddi_mask, 'Potential Savings'] = 0
                    
                    self.data_logger.info(f"Applied DDI/carrier matching rule to {ddi_match_count} rows in detailed report")
            else:
                self.data_logger.warning("Cannot apply DDI/carrier matching rule - required columns missing in detailed report")
                    
            # Calculate total affected rows
            business_stats['total_rows_affected'] = (
                business_stats['same_carrier_rule_applied'] +
                business_stats['empty_data_rule_applied'] +
                business_stats['negative_savings_rule_applied'] +
                business_stats['tl_carrier_rule_applied'] +
                business_stats['ddi_carrier_rule_applied']
            )
            
            self.data_logger.log_data_stats(business_stats, "DETAILED_BUSINESS_LOGIC")
                
        except Exception as e:
            self.data_logger.error("Detailed business logic application failed", exception=e, 
                                 df_shape=df.shape, df_columns=df.columns.tolist())
            raise RuntimeError(f"Detailed business logic error: {str(e)}")
            
        return df
    
    def _copy_selected_to_least_cost(self, df, mask, column_pairs):
        """Helper method to copy selected carrier data to least cost columns"""
        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]
    
    def _calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics for detailed reports"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'percentage_savings': 0,
                'loads_with_savings': 0,
                'total_savings_opportunity': 0,
                'report_type': 'detailed'
            }
            return

        # Clean the Potential Savings column for stats calculation
        if 'Potential Savings' in df.columns:
            # Clean the Potential Savings column more thoroughly
            ps_series = df['Potential Savings'].copy()

            # Convert to string first to handle mixed types
            ps_series = ps_series.astype(str)

            # Clean common non-numeric values
            ps_series = ps_series.replace(['nan', 'None', '', 'N/A', '#N/A'], '0')

            # Remove any currency symbols and commas
            ps_series = ps_series.str.replace('$', '', regex=False)
            ps_series = ps_series.str.replace(',', '', regex=False)
            ps_series = ps_series.str.strip()

            # Convert to numeric
            ps_numeric = pd.to_numeric(ps_series, errors='coerce').fillna(0)
        else:
            ps_numeric = pd.Series([0] * len(df))

        # Basic stats with improved numeric conversion
        total_loads = len(df)

        # Clean cost columns the same way
        if 'Selected Total Cost' in df.columns:
            selected_cost_series = df['Selected Total Cost'].astype(str).str.replace('[$,]', '', regex=True).str.strip()
            total_selected_cost = pd.to_numeric(selected_cost_series, errors='coerce').fillna(0).sum()
        else:
            total_selected_cost = 0

        if 'Least Cost Total Cost' in df.columns:
            least_cost_series = df['Least Cost Total Cost'].astype(str).str.replace('[$,]', '', regex=True).str.strip()
            total_least_cost = pd.to_numeric(least_cost_series, errors='coerce').fillna(0).sum()
        else:
            total_least_cost = 0

        total_potential_savings = ps_numeric.sum()

        # Advanced stats
        savings_mask = ps_numeric > 0
        loads_with_savings = int(savings_mask.sum())  # Convert to Python int
        total_savings_opportunity = float(ps_numeric[savings_mask].sum())  # Convert to Python float

        # Calculate percentages
        if total_selected_cost > 0:
            percentage_savings = (total_potential_savings / total_selected_cost) * 100
        else:
            percentage_savings = 0

        if total_loads > 0:
            average_savings_per_load = total_potential_savings / total_loads
        else:
            average_savings_per_load = 0


        self.summary_stats = {
            'total_loads': int(total_loads),
            'total_selected_cost': float(total_selected_cost),
            'total_least_cost': float(total_least_cost),
            'total_potential_savings': float(total_potential_savings),
            'average_savings_per_load': float(average_savings_per_load),
            'percentage_savings': float(percentage_savings),
            'loads_with_savings': loads_with_savings,  # Already converted above
            'total_savings_opportunity': total_savings_opportunity,  # Already converted above
            'report_type': 'detailed'
        }
    
    def save_processed_data(self, output_file):
        """Save processed detailed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Detailed Processed Data"
        
        # Add company and date info
        row = 1
        if self.title_info:
            # Style company and date range headers
            header_style_border = Border(
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                bottom=Side(style='medium', color='1F4E79')
            )
            
            if 'company_name' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                company_cell = ws_data[f'A{row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="FFFFFF")
                company_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                company_cell.border = header_style_border
                company_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
                
            if 'date_range' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                date_cell = ws_data[f'A{row}']
                date_cell.value = f"Date Range: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="FFFFFF")
                date_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                date_cell.border = header_style_border
                date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
            
            # Report type indicator
            type_cell = ws_data[f'A{row}']
            type_cell.value = "Report Type: DETAILED"
            type_cell.font = Font(size=11, bold=True, color="FFFFFF")
            type_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            type_cell.border = header_style_border
            type_cell.alignment = Alignment(horizontal="center", vertical="center")
            last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
            ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
            for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                cell = ws_data.cell(row=row, column=col)
                cell.border = header_style_border
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            ws_data.row_dimensions[row].height = 25
            row += 1
            
            # Add section headers for detailed reports
            row = 5
            # Find column positions for detailed report sections
            headers = self.processed_data.columns.tolist()
            selected_start = next((i for i, h in enumerate(headers) if 'Selected Carrier' in h), 9) + 1
            selected_end = next((i for i, h in enumerate(headers) if 'Selected Total Cost' in h), 13) + 1
            least_cost_start = next((i for i, h in enumerate(headers) if 'Least Cost Carrier' in h), 15) + 1
            least_cost_end = next((i for i, h in enumerate(headers) if 'Least Cost Total Cost' in h), 19) + 1
            
            # Selected Carrier section header
            if selected_start and selected_end:
                selected_header = ws_data.cell(row=row, column=selected_start, value="Selected Carrier")
                selected_header.font = Font(size=10, bold=True, color="FFFFFF")
                selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                selected_header.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(f'{get_column_letter(selected_start)}{row}:{get_column_letter(selected_end)}{row}')
                for col in range(selected_start, selected_end + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            
            # Least Cost Carrier section header
            if least_cost_start and least_cost_end:
                least_cost_header = ws_data.cell(row=row, column=least_cost_start, value="Least Cost Carrier")
                least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")
                least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
                least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
                ws_data.merge_cells(f'{get_column_letter(least_cost_start)}{row}:{get_column_letter(least_cost_end)}{row}')
                for col in range(least_cost_start, least_cost_end + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            
            row = 6  # Headers will be on row 6
        
        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Color code headers based on section and detailed report specifics
            if any(x in header for x in ['Selected Carrier', 'Selected Service', 'Selected Transit', 'Selected Freight', 'Selected Access', 'Selected Total']):
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light Blue
            elif any(x in header for x in ['Least Cost Carrier', 'Least Cost Service', 'Least Cost Transit', 'Least Cost Freight', 'Least Cost Access', 'Least Cost Total']):
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Light Orange
            elif header == 'Potential Savings':
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
            elif header in ['Origin Name', 'Created By']:  # Detailed report specific columns
                cell.fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")  # Light Purple for detailed-specific
            else:
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Default blue
        
        # Add data with alternating row colors
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Filter out any remaining empty rows
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']
        
        # Ensure all data is properly typed
        for col in clean_data.columns:
            # Only convert actual cost/numeric columns, not carrier columns that contain "Cost" in the name
            if any(x in col for x in ['Transit Days', 'Freight Cost', 'Accessorial Cost', 'Total Cost', 'Savings']) and 'Carrier' not in col:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')
        
        for data_idx, data_row in enumerate(dataframe_to_rows(clean_data, index=False, header=False)):
            # Skip mostly empty rows
            try:
                non_empty_count = sum(1 for val in data_row if val is not None and str(val).strip() != '' and str(val) != 'nan')
                if non_empty_count < 3:
                    continue
            except Exception:
                continue
                
            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"
            
            # Dynamic height calculation
            max_content_length = max(len(str(val)) if val else 0 for val in data_row)
            if max_content_length > 25:
                optimal_height = min(50, max(30, max_content_length * 1.2))
            elif max_content_length > 20:
                optimal_height = min(35, max(22, max_content_length * 0.8))
            else:
                optimal_height = 20
            
            ws_data.row_dimensions[row].height = optimal_height
            
            # Apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border
                
                # Apply color coding based on column type
                if any(x in header_name for x in ['Selected Carrier', 'Selected Service', 'Selected Transit', 'Selected Freight', 'Selected Access', 'Selected Total']):
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(start_color=light_blue_bg, end_color=light_blue_bg, fill_type="solid")
                elif any(x in header_name for x in ['Least Cost Carrier', 'Least Cost Service', 'Least Cost Transit', 'Least Cost Freight', 'Least Cost Access', 'Least Cost Total']):
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(start_color=light_orange_bg, end_color=light_orange_bg, fill_type="solid")
                elif header_name in ['Origin Name', 'Created By']:  # Detailed report specific
                    light_purple_bg = "F5F0FF" if data_idx % 2 == 0 else "F8F5FF"
                    cell.fill = PatternFill(start_color=light_purple_bg, end_color=light_purple_bg, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                
                # Format currency columns and highlight positive savings
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost', 'Selected Freight Cost', 'Least Cost Freight Cost', 'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if any(x in header_name for x in currency_columns) or header_name == 'Potential Savings':
                    cell.number_format = '"$"#,##0.00'
                    if header_name == 'Potential Savings':
                        try:
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                numeric_value = float(value) if isinstance(value, (int, float)) else float(str(value).replace('$', '').replace(',', ''))
                                if numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        except (ValueError, TypeError, AttributeError):
                            pass
                    cell.font = Font(size=10, bold=False, color="2C3E50")
                else:
                    cell.font = Font(size=10, color="495057")
        
        # Enable auto-filter
        try:
            header_row_idx = 6
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Add totals row with key metrics
        totals_row = row + 2
        
        # Add "DETAILED TOTALS" label
        totals_label = ws_data.cell(row=totals_row, column=1, value="DETAILED TOTALS")
        totals_label.font = Font(size=12, bold=True, color="FFFFFF")
        totals_label.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        totals_label.alignment = Alignment(horizontal="center", vertical="center")
        totals_label.border = Border(
            left=Side(style='medium', color='2C3E50'),
            right=Side(style='medium', color='2C3E50'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )
        
        # Find and populate totals
        for col_idx, header in enumerate(headers, 1):
            if 'Selected Total Cost' in header:
                cost_cell = ws_data.cell(row=totals_row, column=col_idx, 
                                       value=f"${self.summary_stats['total_selected_cost']:,.2f}")
                cost_cell.font = Font(size=12, bold=True, color="FFFFFF")
                cost_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                cost_cell.alignment = Alignment(horizontal="center", vertical="center")
                cost_cell.number_format = '"$"#,##0.00'
                cost_cell.border = Border(
                    left=Side(style='medium', color='3498DB'),
                    right=Side(style='medium', color='3498DB'),
                    top=Side(style='medium', color='3498DB'),
                    bottom=Side(style='medium', color='3498DB')
                )
            elif header == 'Potential Savings':
                savings_cell = ws_data.cell(row=totals_row, column=col_idx, 
                                          value=f"${self.summary_stats['total_potential_savings']:,.2f}")
                savings_cell.font = Font(size=14, bold=True, color="FFFFFF")
                savings_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                savings_cell.alignment = Alignment(horizontal="center", vertical="center")
                savings_cell.number_format = '"$"#,##0.00'
                savings_cell.border = Border(
                    left=Side(style='thick', color='27AE60'),
                    right=Side(style='thick', color='27AE60'),
                    top=Side(style='thick', color='27AE60'),
                    bottom=Side(style='thick', color='27AE60')
                )
        
        ws_data.row_dimensions[totals_row].height = 25

        # Auto-fit column widths for detailed report
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(6, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                
                # Adjust width based on content type
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                if header_name in ['Origin Name', 'Created By']:  # Detailed specific columns may be longer
                    padding = 1.5
                    max_width = 30
                elif any(c.isalpha() for c in str(ws_data.cell(row=6, column=col_idx).value or "")):
                    padding = 1.0
                    max_width = 25
                else:
                    padding = 0.5
                    max_width = 15
                
                adjusted_width = max_length + padding
                final_width = min(adjusted_width, max_width)
                ws_data.column_dimensions[col_letter].width = max(6, final_width)
        except Exception:
            pass

        wb.save(output_file)
        wb.close()
        
        self.logger.info(f"Detailed report saved successfully to {output_file}")