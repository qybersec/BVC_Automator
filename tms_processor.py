import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Try to import calendar widget - fallback to text entry if not available
try:
    from tkcalendar import DateEntry, Calendar
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False
    print("Note: tkcalendar not available. Install with 'pip install tkcalendar' for calendar widgets.")

# Import UI styles from the new modular UI components
try:
    from ui import COLORS as UI_COLORS
except ImportError:
    # Fallback UI color constants for consistent theming
    UI_COLORS = {
        # Primary colors
        'PRIMARY_BLUE': '#4299e1',
        'PRIMARY_BLUE_HOVER': '#3182ce',
        'PRIMARY_BLUE_PRESSED': '#2c5aa0',
        
        # Background colors
        'BACKGROUND_WHITE': '#ffffff',
        'BACKGROUND_GRAY': '#f8f9fa',
        'BACKGROUND_LIGHT': '#f7fafc',
        'BACKGROUND_BORDER': '#e2e8f0',
        
        # Text colors
        'TEXT_PRIMARY': '#2d3748',
        'TEXT_SECONDARY': '#4a5568',
        'TEXT_MUTED': '#718096',
        'TEXT_DISABLED': '#a0aec0',
        
        # Status colors
        'SUCCESS_GREEN': '#38a169',
        'SUCCESS_LIGHT': '#e6fffa',
        'WARNING_YELLOW': '#ffd700',
        'ERROR_RED': '#e53e3e',
        
        # Navigation colors
        'NAV_INACTIVE': '#ffffff',
        'NAV_ACTIVE': '#4299e1',
        
        # Legacy colors for backward compatibility
        'ACCENT_PURPLE': '#667eea'
    }

# Import our new modules
try:
    from config import tms_config
    from logger_config import main_logger, data_logger, gui_logger, ProgressLogger
    from validators import tms_validator, tms_cleaner
    from src.generators import create_bvc_generator
except ImportError as e:
    print(f"Warning: Could not import enhanced modules: {e}")
    print("Falling back to basic functionality...")
    # Create mock objects for backward compatibility
    class MockConfig:
        def get(self, key, default=None):
            defaults = {
                'data_structure.default_header_row': 8,
                'data_structure.default_data_start_row': 11,
                'data_structure.min_data_columns': 5,
                'data_structure.expected_columns': 21,
                'business_rules.min_non_empty_values': 5
            }
            return defaults.get(key, default)
    
    class MockLogger:
        def info(self, msg, **kwargs): print(f"INFO: {msg}")
        def error(self, msg, **kwargs): print(f"ERROR: {msg}")
        def warning(self, msg, **kwargs): print(f"WARNING: {msg}")
        def debug(self, msg, **kwargs): print(f"DEBUG: {msg}")
        def log_processing_step(self, step, details=None): self.info(f"Processing: {step}")
        def log_data_stats(self, stats, prefix=""): self.info(f"{prefix}: {stats}")
        def log_performance(self, op, duration, records=None): self.info(f"{op}: {duration:.2f}s")
    
    class MockValidator:
        def run_full_validation(self, file_path): return {'overall_valid': True, 'validation_steps': {}}
    
    def create_bvc_generator():
        class MockTemplateGenerator:
            def generate_template(self, date_range, output_file=None): 
                raise RuntimeError("Template generator not available")
            def validate_input(self, date_range): 
                return True
        return MockTemplateGenerator()
    
    tms_config = MockConfig()
    main_logger = data_logger = gui_logger = MockLogger()
    tms_validator = MockValidator()
    
    def ProgressLogger(logger, total, operation):
        class MockProgress:
            def update(self, inc=1): pass
            def complete(self): pass
        return MockProgress()

class ModernTMSProcessor:
    """Enhanced TMS Processor with comprehensive validation and error handling"""
    
    def __init__(self):
        self.logger = main_logger
        self.data_logger = data_logger
        self.config = tms_config
        
        # Data storage
        self.raw_data = None
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}
        self.validation_results = None
        
        # Performance tracking
        self.processing_start_time = None
        self.processing_stats = {}
        
        self.logger.info("ModernTMSProcessor initialized with enhanced features")
        
    def _extract_title_info(self, df_raw):
        """Extract title and report information from the top rows"""
        title_info = {}
        
        try:
            # Extract report title with bounds checking
            if len(df_raw) > 1 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[1, 1]):
                title_info['report_title'] = str(df_raw.iloc[1, 1])
            
            # Extract company name
            if len(df_raw) > 3 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[3, 1]):
                title_info['company_name'] = str(df_raw.iloc[3, 1])
            
            # Extract date range
            if len(df_raw) > 5 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[5, 1]):
                title_info['date_range'] = str(df_raw.iloc[5, 1])
        except (IndexError, KeyError):
            # If extraction fails, continue with empty title_info
            pass
        
        return title_info
    
    def _detect_data_structure(self, df_raw):
        """Intelligently detect header and data start positions"""
        header_row = self.DEFAULT_HEADER_ROW
        data_start_row = self.DEFAULT_DATA_START_ROW
        
        # Look for header indicators in different rows
        header_indicators = ['Load No.', 'Carrier', 'Service Type', 'Ship Date']
        
        for row_idx in range(5, min(15, len(df_raw))):
            row_data = df_raw.iloc[row_idx].dropna().astype(str).tolist()
            row_str = ' '.join(row_data).lower()
            
            # Check if this row contains header-like content
            matches = sum(1 for indicator in header_indicators if indicator.lower() in row_str)
            if matches >= 2:  # Found at least 2 header indicators
                header_row = row_idx
                data_start_row = row_idx + 2  # Skip potential blank row
                break
        
        return header_row, data_start_row
        
    def _remove_duplicate_headers(self, df):
        """Remove duplicate header rows that appear in the middle of data"""
        # Look for rows that contain header-like text
        header_indicators = ['Load No.', 'Carrier', 'Service Type']
        
        rows_to_drop = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.dropna().tolist()])
            if any(indicator in row_str for indicator in header_indicators):
                # Check if this looks like a header row (not actual data)
                if not any(str(val).startswith('A') and str(val)[1:].isdigit() for val in row.dropna().tolist()):
                    rows_to_drop.append(idx)
        
        return df.drop(rows_to_drop)
    
    def clean_and_process_data(self, file_path: str) -> pd.DataFrame:
        """Main function to clean and process the TMS Excel file with comprehensive validation"""
        self.processing_start_time = time.time()
        self.logger.log_processing_step("Starting TMS data processing", {'file': Path(file_path).name})
        
        try:
            # Step 1: Comprehensive validation
            self.validation_results = tms_validator.run_full_validation(file_path)
            
            if not self.validation_results['overall_valid']:
                failed_steps = [step for step, result in self.validation_results['validation_steps'].items() 
                               if not result.get('valid', False)]
                raise ValueError(f"File validation failed. Issues: {failed_steps}")
            
            # Step 2: Load and validate Excel data
            self.logger.log_processing_step("Loading Excel file")
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)
            
            self.logger.log_data_stats({
                'raw_rows': len(df_raw),
                'raw_columns': len(df_raw.columns),
                'file_size_mb': Path(file_path).stat().st_size / (1024*1024)
            }, "RAW_DATA")
            
            # Step 3: Extract metadata
            self.logger.log_processing_step("Extracting title information")
            self.title_info = self._extract_title_info(df_raw)
            
            # Step 4: Use validation results for structure detection
            header_info = self.validation_results['validation_steps']['header_detection']['details']
            header_row = header_info.get('header_row', self.config.get('data_structure.default_header_row', 8))
            data_start_row = header_info.get('data_start_row', self.config.get('data_structure.default_data_start_row', 11))
            
            self.logger.log_processing_step("Data structure detected", {
                'header_row': header_row,
                'data_start_row': data_start_row,
                'confidence': header_info.get('confidence_score', 0)
            })
            
            # Get headers
            headers = df_raw.iloc[header_row].dropna().tolist()
            
            # Step 5: Extract and clean data with progress tracking
            self.logger.log_processing_step("Extracting data rows")
            data_df = df_raw.iloc[data_start_row:].copy()
            
            # Remove completely empty rows and duplicate header rows with logging
            initial_rows = len(data_df)
            data_df = data_df.dropna(how='all')
            empty_rows_removed = initial_rows - len(data_df)
            
            data_df = self._remove_duplicate_headers(data_df)
            duplicate_headers_removed = initial_rows - empty_rows_removed - len(data_df)
            
            self.logger.log_data_stats({
                'initial_data_rows': initial_rows,
                'empty_rows_removed': empty_rows_removed,
                'duplicate_headers_removed': duplicate_headers_removed,
                'remaining_rows': len(data_df)
            }, "DATA_CLEANING")
            
            # Reset index after dropping rows
            data_df = data_df.reset_index(drop=True)
            
            # Step 6: Column extraction with intelligent selection
            self.logger.log_processing_step("Extracting relevant columns")
            max_cols = min(22, len(data_df.columns))
            relevant_columns = list(range(2, max_cols + 1))
            data_df = data_df.iloc[:, relevant_columns]
            
            self.logger.log_data_stats({
                'total_available_columns': len(df_raw.columns),
                'relevant_columns_selected': len(relevant_columns),
                'max_expected_columns': 22
            }, "COLUMN_EXTRACTION")
            
            # Set proper column names with full descriptive headers
            base_column_names = [
                'Load No.', 'Ship Date', 'Origin City', 'Origin State', 'Origin Postal',
                'Destination City', 'Destination State', 'Destination Postal',
                'Selected Carrier', 'Selected Service Type', 'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
                'Least Cost Carrier', 'Least Cost Service Type', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
                'Potential Savings'
            ]
            
            # Ensure column names match the actual extracted columns
            if len(base_column_names) != len(data_df.columns):
                print(f"Warning: Column count mismatch. Expected {len(base_column_names)}, got {len(data_df.columns)}")
                # Adjust column names to match actual columns
                if len(data_df.columns) < len(base_column_names):
                    column_names = base_column_names[:len(data_df.columns)]
                else:
                    # Add generic names for extra columns
                    extra_columns = [f'Column_{i}' for i in range(len(base_column_names), len(data_df.columns))]
                    column_names = base_column_names + extra_columns
            else:
                column_names = base_column_names
            
            data_df.columns = column_names
            
            # Step 7: Enhanced data type cleaning with validation
            self.logger.log_processing_step("Cleaning and validating data types")
            cleaning_start = time.time()
            data_df = self._clean_data_types_enhanced(data_df)
            cleaning_time = time.time() - cleaning_start
            self.logger.log_performance("Data type cleaning", cleaning_time, len(data_df))
            
            # Enhanced row filtering with detailed logging
            self.logger.log_processing_step("Filtering invalid rows")
            pre_filter_count = len(data_df)
            
            # Remove rows where Load No. is missing or empty
            if 'Load No.' in data_df.columns:
                data_df = data_df.dropna(subset=['Load No.'])
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != '']
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != 'nan']
            else:
                self.logger.warning("Load No. column not found - skipping Load No. validation")
            
            # Remove any remaining rows that are mostly empty
            min_values = self.config.get('business_rules.min_non_empty_values', 5)
            data_df = data_df.dropna(thresh=min_values)
            
            # Reset index and log filtering results
            data_df = data_df.reset_index(drop=True)
            rows_filtered = pre_filter_count - len(data_df)
            
            self.logger.log_data_stats({
                'rows_before_filtering': pre_filter_count,
                'rows_after_filtering': len(data_df),
                'rows_removed': rows_filtered,
                'filter_rate': f"{(rows_filtered/pre_filter_count*100):.1f}%" if pre_filter_count > 0 else "0%"
            }, "ROW_FILTERING")
            
            # Apply business logic rules with enhanced tracking
            self.logger.log_processing_step("Applying business logic rules")
            business_start = time.time()
            data_df = self._apply_business_logic_enhanced(data_df)
            business_time = time.time() - business_start
            self.logger.log_performance("Business logic application", business_time, len(data_df))
            
            # Sort by Destination City with error handling
            self.logger.log_processing_step("Sorting data")
            if 'Destination City' in data_df.columns:
                data_df = data_df.sort_values('Destination City', na_position='last')
            else:
                self.logger.warning("Destination City column not found - skipping sort")
                # Try alternative column names
                destination_cols = [col for col in data_df.columns if 'destination' in col.lower() and 'city' in col.lower()]
                if destination_cols:
                    data_df = data_df.sort_values(destination_cols[0], na_position='last')
                    self.logger.info(f"Sorted by alternative column: {destination_cols[0]}")
            
            # Step 8: Calculate summary statistics
            self.logger.log_processing_step("Calculating summary statistics")
            self._calculate_summary_stats(data_df)
            
            # Step 9: Final processing metrics
            processing_time = time.time() - self.processing_start_time
            self.processing_stats = {
                'total_time': processing_time,
                'records_processed': len(data_df),
                'processing_rate': len(data_df) / processing_time if processing_time > 0 else 0
            }
            
            self.logger.log_performance(
                "Total TMS processing", 
                processing_time, 
                len(data_df)
            )
            
            self.processed_data = data_df
            return data_df
            
        except (FileNotFoundError, PermissionError) as e:
            self.logger.error("File access error", exception=e, file_path=file_path)
            raise FileNotFoundError(f"Cannot access file: {str(e)}")
        except (pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            self.logger.error("Excel parsing error", exception=e, file_path=file_path)
            raise ValueError(f"Invalid Excel file format: {str(e)}")
        except ValueError as e:
            # Re-raise validation errors
            self.logger.error("Validation error", exception=e)
            raise
        except Exception as e:
            self.logger.error("Unexpected processing error", exception=e, file_path=file_path)
            raise RuntimeError(f"Error processing file: {str(e)}")
    
    def _clean_data_types_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhanced data type cleaning with comprehensive validation and logging"""
        df = df.copy()
        cleaning_stats = {'columns_processed': 0, 'conversion_failures': 0}
        
        # Convert numeric columns with enhanced error tracking
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings'
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1
                original_nulls = df[col].isnull().sum()
                df[col] = pd.to_numeric(df[col], errors='coerce')
                new_nulls = df[col].isnull().sum()
                conversion_failures = new_nulls - original_nulls
                if conversion_failures > 0:
                    cleaning_stats['conversion_failures'] += conversion_failures
                    self.data_logger.warning(f"Failed to convert {conversion_failures} values in {col} to numeric")
        
        # Ensure PS column is properly numeric and handle any string values
        if 'PS' in df.columns:
            # First try to convert to numeric, handling any string values
            df['PS'] = pd.to_numeric(df['PS'], errors='coerce')
            # Fill any NaN values with 0
            df['PS'] = df['PS'].fillna(0)
        
        # Convert date column with enhanced error handling
        if 'Ship Date' in df.columns:
            cleaning_stats['columns_processed'] += 1
            original_nulls = df['Ship Date'].isnull().sum()
            date_series = pd.to_datetime(df['Ship Date'], errors='coerce')
            new_nulls = date_series.isnull().sum()
            date_failures = new_nulls - original_nulls
            if date_failures > 0:
                cleaning_stats['conversion_failures'] += date_failures
                self.data_logger.warning(f"Failed to convert {date_failures} date values in Ship Date")
            
            date_format = self.config.get('formatting.date_format', '%m/%d/%y')
            df['Ship Date'] = date_series.dt.strftime(date_format)
        
        # Clean string columns with tracking
        string_columns = [
            'Load No.', 'Origin City', 'Origin State', 'Origin Postal',
            'Destination City', 'Destination State', 'Destination Postal',
            'Selected Carrier', 'Selected Service Type', 'Least Cost Carrier', 'Least Cost Service Type'
        ]
        
        for col in string_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
        
        self.data_logger.log_data_stats(cleaning_stats, "TYPE_CLEANING")
        return df
    
    def _apply_business_logic_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply TMS business logic rules with enhanced tracking and validation"""
        df = df.copy()
        business_stats = {
            'same_carrier_rule_applied': 0,
            'empty_data_rule_applied': 0,
            'negative_savings_rule_applied': 0,
            'ddi_carrier_rule_applied': 0,
            'total_rows_affected': 0
        }
        
        try:
            # Ensure PS column is numeric from the start to avoid comparison errors
            if 'PS' in df.columns:
                df['PS'] = pd.to_numeric(df['PS'], errors='coerce').fillna(0)
            else:
                print("Warning: PS column not found in dataframe")
                print(f"Available columns: {df.columns.tolist()}")
        
            # Rule 1: Same Carriers - Set Potential Savings to 0 (Enhanced)
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                same_carrier_mask = (
                    (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) & 
                    (df['Selected Carrier'].notna()) & 
                    (df['Least Cost Carrier'].notna()) &
                    (df['Selected Carrier'].astype(str) != '') & 
                    (df['Least Cost Carrier'].astype(str) != '') &
                    (df['Selected Carrier'].astype(str) != 'nan') & 
                    (df['Least Cost Carrier'].astype(str) != 'nan')
                )
                
                same_carrier_count = same_carrier_mask.sum()
                business_stats['same_carrier_rule_applied'] = same_carrier_count
                
                if 'Potential Savings' in df.columns and same_carrier_count > 0:
                    default_savings = self.config.get('business_rules.same_carrier_savings', 0.0)
                    df.loc[same_carrier_mask, 'Potential Savings'] = default_savings
                    self.data_logger.info(f"Applied same carrier rule to {same_carrier_count} rows")
            else:
                self.data_logger.warning("Cannot apply same carrier rule - required columns missing")
            
            # Rule 2: Empty Least Cost - Copy Selected data and set savings to 0 (Enhanced)
            if 'Least Cost Carrier' in df.columns:
                empty_least_cost_mask = (
                    df['Least Cost Carrier'].isna() | 
                    (df['Least Cost Carrier'].astype(str) == '') |
                    (df['Least Cost Carrier'].astype(str) == 'nan')
                )
                
                empty_count = empty_least_cost_mask.sum()
                business_stats['empty_data_rule_applied'] = empty_count
                
                if empty_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, empty_least_cost_mask, column_pairs)
                    
                    if 'Potential Savings' in df.columns:
                        df.loc[empty_least_cost_mask, 'Potential Savings'] = 0
                    
                    self.data_logger.info(f"Applied empty data rule to {empty_count} rows")
            else:
                self.data_logger.warning("Cannot apply empty data rule - Least Cost Carrier column missing")

            # Rule 3: Negative Savings - Copy Selected data and set savings to 0 (Enhanced)
            if 'Potential Savings' in df.columns:
                # Ensure Potential Savings is numeric before comparison
                ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
                negative_savings_mask = ps_numeric < 0
                negative_count = negative_savings_mask.sum()
                business_stats['negative_savings_rule_applied'] = negative_count
                
                if negative_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, negative_savings_mask, column_pairs)
                    df.loc[negative_savings_mask, 'Potential Savings'] = 0
                    self.data_logger.info(f"Applied negative savings rule to {negative_count} rows")
            else:
                self.data_logger.warning("Cannot apply negative savings rule - Potential Savings column missing")

            # Rule 4: DDI/Carrier Matching - New custom rule
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Create mask for rows where Selected Carrier contains "DDI/" or similar patterns
                # and the part after "/" matches Least Cost Carrier
                ddi_matches = []
                
                for idx, row in df.iterrows():
                    selected = str(row['Selected Carrier']).strip()
                    least_cost = str(row['Least Cost Carrier']).strip()
                    
                    # Skip empty or nan values
                    if selected in ['', 'nan', 'None'] or least_cost in ['', 'nan', 'None']:
                        continue
                    
                    # Check if selected carrier has "/" and extract the part after it
                    if '/' in selected:
                        # Split on "/" and get the part after the last "/"
                        carrier_after_slash = selected.split('/')[-1].strip()
                        
                        # Check if the carrier after "/" matches the least cost carrier
                        # Using case-insensitive comparison and handling common variations
                        if carrier_after_slash.upper() == least_cost.upper():
                            ddi_matches.append(idx)
                        # Also check for R&L Carriers vs R%L Carriers variations
                        elif (carrier_after_slash.upper().replace('&', '%') == least_cost.upper().replace('&', '%') or
                              carrier_after_slash.upper().replace('%', '&') == least_cost.upper().replace('%', '&')):
                            ddi_matches.append(idx)
                
                ddi_match_count = len(ddi_matches)
                business_stats['ddi_carrier_rule_applied'] = ddi_match_count
                
                if ddi_match_count > 0:
                    ddi_mask = df.index.isin(ddi_matches)
                    
                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'), 
                        ('Selected Service Type', 'Least Cost Service Type'), 
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'), 
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'), 
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, ddi_mask, column_pairs)
                    
                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        df.loc[ddi_mask, 'Potential Savings'] = 0
                    
                    self.data_logger.info(f"Applied DDI/carrier matching rule to {ddi_match_count} rows")
            else:
                self.data_logger.warning("Cannot apply DDI/carrier matching rule - required columns missing")
                
            # Calculate total affected rows
            business_stats['total_rows_affected'] = (
                business_stats['same_carrier_rule_applied'] + 
                business_stats['empty_data_rule_applied'] + 
                business_stats['negative_savings_rule_applied'] +
                business_stats['ddi_carrier_rule_applied']
            )
            
            self.data_logger.log_data_stats(business_stats, "BUSINESS_LOGIC")
                
        except Exception as e:
            self.data_logger.error("Business logic application failed", exception=e, 
                                 df_shape=df.shape, df_columns=df.columns.tolist())
            raise RuntimeError(f"Business logic error: {str(e)}")
            
        return df
    
    def _copy_selected_to_least_cost(self, df, mask, column_pairs):
        """Helper method to copy selected carrier data to least cost columns"""
        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]
    
    def _calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'percentage_savings': 0,
                'loads_with_savings': 0,
                'total_savings_opportunity': 0
            }
            return
        
        # Basic stats - ensure numeric columns are properly converted
        total_loads = len(df)
        total_selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        total_least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        total_potential_savings = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0).sum()
        
        # Advanced stats - optimize by filtering once
        ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
        savings_df = df[ps_numeric > 0]
        loads_with_savings = len(savings_df)
        total_savings_opportunity = pd.to_numeric(savings_df['Potential Savings'], errors='coerce').fillna(0).sum()
        
        # Calculate percentages
        if total_selected_cost > 0:
            percentage_savings = (total_potential_savings / total_selected_cost) * 100
        else:
            percentage_savings = 0
        
        if total_loads > 0:
            average_savings_per_load = total_potential_savings / total_loads
        else:
            average_savings_per_load = 0
        
        self.summary_stats = {
            'total_loads': total_loads,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'percentage_savings': percentage_savings,
            'loads_with_savings': loads_with_savings,
            'total_savings_opportunity': total_savings_opportunity
        }
    
    def save_processed_data(self, output_file):
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Processed Data"
        
        # Add company and date info (no big title rows)
        row = 1
        if self.title_info:
            # Style company and date range with expanded width
            header_style_border = Border(
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                bottom=Side(style='medium', color='1F4E79')
            )
            
            if 'company_name' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                company_cell = ws_data[f'A{row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="FFFFFF")
                company_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                company_cell.border = header_style_border
                company_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
                
            if 'date_range' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                date_cell = ws_data[f'A{row}']
                date_cell.value = f"Date Range: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="FFFFFF")
                date_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                date_cell.border = header_style_border
                date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1
            
            # Add section headers row with color coding
            row = 4
            # Selected Carrier section (columns I-N, which are 9-14) - Light Blue
            selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
            selected_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            selected_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('I4:N4')
            for col in range(9, 15):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            
            # Least Cost Carrier section (columns O-T, which are 15-20) - Light Orange
            least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
            least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")  # Reduced from 11 to 10
            least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('O4:T4')
            for col in range(15, 21):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            
            row = 5  # Headers will be on row 5
        
        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Color code headers based on section
            if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light Blue
            elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Light Orange
            elif header == 'Potential Savings':  # Potential Savings column - Green
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
            else:
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Default blue
        
        # Add data with alternating row colors and comprehensive borders
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Filter out any remaining empty rows before writing to Excel
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']
        
        # Ensure all data is properly typed before processing
        for col in clean_data.columns:
            if col in ['Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings']:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')
        
        for data_idx, data_row in enumerate(dataframe_to_rows(clean_data, index=False, header=False)):
            # Skip rows that are mostly empty
            # Ensure we're comparing integers by converting the sum result
            try:
                non_empty_count = sum(1 for val in data_row if val is not None and str(val).strip() != '' and str(val) != 'nan')
                if non_empty_count < 3:
                    continue
            except Exception as e:
                print(f"Error processing row {data_idx}: {e}")
                print(f"Row data: {data_row}")
                continue
                
            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"
            
            # First pass: collect all content lengths to determine optimal row height
            max_content_length = 0
            for col_idx, value in enumerate(data_row, 1):
                content_length = len(str(value)) if value else 0
                max_content_length = max(max_content_length, content_length)
            
            # Enhanced dynamic height calculation for long carrier names
            # Check if this row contains carrier information that might wrap
            has_carrier_data = any('TRANSPORT' in str(val).upper() or 
                                 'LOGISTICS' in str(val).upper() or 
                                 'FREIGHT' in str(val).upper() or
                                 len(str(val)) > 25 for val in data_row if val)
            
            if has_carrier_data and max_content_length > 25:
                # For carrier names, be more generous with height to prevent cutoff
                optimal_height = min(50, max(30, max_content_length * 1.2))
            elif max_content_length > 30:  # Very long content
                optimal_height = min(45, max(25, max_content_length * 1.0))
            elif max_content_length > 20:  # Long content
                optimal_height = min(35, max(22, max_content_length * 0.8))
            elif max_content_length > 15:  # Medium content
                optimal_height = 25
            else:
                optimal_height = 20  # Default height with a bit more room
            
            # Set the row height once for the entire row
            ws_data.row_dimensions[row].height = optimal_height
            
            # Second pass: apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                # Center all cell contents and enable text wrapping for compactness
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border
                
                # Apply color coding to data cells based on section
                if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(start_color=light_blue_bg, end_color=light_blue_bg, fill_type="solid")
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(start_color=light_orange_bg, end_color=light_orange_bg, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                
                # Format currency columns and apply green color for positive Potential Savings values
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost', 'Selected Freight Cost', 'Least Cost Freight Cost', 'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if headers[col_idx-1] in currency_columns or headers[col_idx-1] == 'Potential Savings':
                    cell.number_format = '"$"#,##0.00'
                    # Apply light green background for positive Potential Savings values
                    if headers[col_idx-1] == 'Potential Savings':
                        try:
                            # Convert value to float for comparison, handle None and string values
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                # Handle different value types safely
                                if isinstance(value, (int, float)):
                                    numeric_value = float(value)
                                else:
                                    numeric_value = float(str(value).replace('$', '').replace(',', ''))
                                if numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        except (ValueError, TypeError, AttributeError):
                            pass  # Skip coloring if value can't be converted
                    cell.font = Font(size=10, bold=False, color="2C3E50")
            else:
                cell.font = Font(size=10, color="495057")
        
        # Enable auto-filter over header and data range (no freeze panes)
        try:
            header_row_idx = 5
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Add totals row with key financial metrics
        totals_row = row + 2
        
        # Add "TOTALS" label
        totals_label = ws_data.cell(row=totals_row, column=1, value="TOTALS")
        totals_label.font = Font(size=12, bold=True, color="FFFFFF")
        totals_label.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        totals_label.alignment = Alignment(horizontal="center", vertical="center")
        totals_label.border = Border(
            left=Side(style='medium', color='2C3E50'),
            right=Side(style='medium', color='2C3E50'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )
        
        # Find the Selected Total Cost and Potential Savings columns
        selected_cost_col = None
        potential_savings_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'Selected Total Cost' in str(header):
                selected_cost_col = col_idx
            elif 'Potential Savings' in str(header):
                potential_savings_col = col_idx
        
        # Add Total Selected Cost
        if selected_cost_col:
            cost_cell = ws_data.cell(row=totals_row, column=selected_cost_col, 
                                   value=f"${self.summary_stats['total_selected_cost']:,.2f}")
            cost_cell.font = Font(size=12, bold=True, color="FFFFFF")
            cost_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            cost_cell.alignment = Alignment(horizontal="center", vertical="center")
            cost_cell.number_format = '"$"#,##0.00'
            cost_cell.border = Border(
                left=Side(style='medium', color='3498DB'),
                right=Side(style='medium', color='3498DB'),
                top=Side(style='medium', color='3498DB'),
                bottom=Side(style='medium', color='3498DB')
            )
        
        # Add Total Potential Savings (most important number)
        if potential_savings_col:
            savings_cell = ws_data.cell(row=totals_row, column=potential_savings_col, 
                                      value=f"${self.summary_stats['total_potential_savings']:,.2f}")
            savings_cell.font = Font(size=14, bold=True, color="FFFFFF")  # Larger font for emphasis
            savings_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            savings_cell.alignment = Alignment(horizontal="center", vertical="center")
            savings_cell.number_format = '"$"#,##0.00'
            savings_cell.border = Border(
                left=Side(style='thick', color='27AE60'),  # Thicker border for emphasis
                right=Side(style='thick', color='27AE60'),
                top=Side(style='thick', color='27AE60'),
                bottom=Side(style='thick', color='27AE60')
            )
        
        # Set height for totals row
        ws_data.row_dimensions[totals_row].height = 25

        
        # Auto-fit column widths on the Processed Data sheet (compact and consistent for all: table, CAL, PI)
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(5, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value)
                        length = len(text)
                        if length > max_length:
                            max_length = length
                # Aggressive compact auto-sizing: minimize width while ensuring content fits
                # Check if this column contains text content
                has_text_content = False
                for check_row in range(5, ws_data.max_row + 1):
                    check_cell = ws_data.cell(row=check_row, column=col_idx)
                    if check_cell.value and any(c.isalpha() for c in str(check_cell.value)):
                        has_text_content = True
                        break
                
                # Very tight padding for maximum compactness
                if has_text_content:
                    padding = 1.0  # Minimal padding for text
                else:
                    padding = 0.5   # Very tight for numbers
                
                # Calculate width with aggressive compacting
                adjusted_width = max_length + padding
                
                # Apply maximum width constraints for compactness
                if has_text_content:
                    max_width = 25  # Cap text columns at 25 characters
                else:
                    max_width = 15  # Cap number columns at 15 characters
                
                final_width = min(adjusted_width, max_width)
                ws_data.column_dimensions[col_letter].width = max(6, final_width)  # Minimum 6 for readability
            
            # Optimize row heights for compact layout
            for rh in [1, 2, 4, 5]:
                if rh <= ws_data.max_row:
                    ws_data.row_dimensions[rh].height = max(ws_data.row_dimensions[rh].height or 0, 20)  # Reduced from 22 to 20
            
            # Ensure gridlines visible
            ws_data.sheet_view.showGridLines = True
        except Exception:
            pass

        # Add thick outside borders around the entire data table
        try:
            # Define thick border styles
            thick_side = Side(style='medium', color='2C3E50')
            
            header_row_idx = 5
            first_row = header_row_idx
            last_row = row
            first_col = 1
            last_col = len(headers)
            
            # Apply thick borders to all edge cells
            for r in range(first_row, last_row + 1):
                for c in range(first_col, last_col + 1):
                    cell = ws_data.cell(row=r, column=c)
                    current_border = cell.border or Border()
                    
                    # Determine which sides need thick borders
                    left_side = thick_side if c == first_col else current_border.left
                    right_side = thick_side if c == last_col else current_border.right
                    top_side = thick_side if r == first_row else current_border.top
                    bottom_side = thick_side if r == last_row else current_border.bottom
                    
                    # Apply the new border
                    cell.border = Border(
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side
                    )
        except Exception:
            pass

        wb.save(output_file)
        wb.close()

class ModernTMSProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("TMS Data Processor Pro")
        self.root.geometry("1200x650")
        self.root.configure(bg='#f8f9fa')
        self.root.minsize(1100, 600)
        self.root.resizable(True, True)
        
        # Initialize processors
        self.basic_processor = ModernTMSProcessor()
        self.detailed_processor = None
        self.template_generator = create_bvc_generator()
        self.input_files = []  # Changed to list for multiple files
        self.output_file = None
        
        # Progress tracking
        self.is_processing = False

        # Savings history tracking
        self.savings_history_file = Path.home() / "Desktop" / "tms_savings_history.json"
        self.savings_history = self.load_savings_history()

        # Configure style
        self.setup_styles()
        
        # Create GUI
        self.create_widgets()
        
        # Center window
        self.center_window()
        
    def setup_styles(self):
        """Setup modern styling for the application"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure colors with modern palette using constants
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 28, 'bold'), 
                       foreground='#1a365d',
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Subtitle.TLabel', 
                       font=('Segoe UI', 12), 
                       foreground=UI_COLORS['TEXT_SECONDARY'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Header.TLabel', 
                       font=('Segoe UI', 13, 'bold'), 
                       foreground=UI_COLORS['TEXT_PRIMARY'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Info.TLabel', 
                       font=('Segoe UI', 10), 
                       foreground=UI_COLORS['TEXT_MUTED'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Success.TLabel', 
                       font=('Segoe UI', 10, 'bold'), 
                       foreground=UI_COLORS['SUCCESS_GREEN'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        
        # Configure modern buttons with hover effects using constants
        style.configure('Primary.TButton', 
                       font=('Segoe UI', 11, 'bold'),
                       background=UI_COLORS['PRIMARY_BLUE'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(20, 10))
        style.map('Primary.TButton',
                 background=[('active', UI_COLORS['PRIMARY_BLUE_HOVER']), ('pressed', UI_COLORS['PRIMARY_BLUE_PRESSED'])])
        
        style.configure('Success.TButton',
                       font=('Segoe UI', 12, 'bold'),
                       background=UI_COLORS['SUCCESS_LIGHT'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(25, 12))
        style.map('Success.TButton',
                 background=[('active', UI_COLORS['SUCCESS_GREEN']), ('pressed', '#2f855a')])
        
        # Enhanced Process Button Style
        style.configure('ProcessButton.TButton',
                       font=('Segoe UI', 14, 'bold'),
                       background='#28a745',
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(35, 15))
        style.map('ProcessButton.TButton',
                 background=[('active', '#218838'), 
                           ('pressed', '#1e7e34'),
                           ('disabled', '#6c757d')])
        
        # Process Button Disabled Style
        style.configure('ProcessButtonDisabled.TButton',
                       font=('Segoe UI', 14, 'bold'),
                       background='#6c757d',
                       foreground='#ffffff',
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(35, 15))
        
        style.configure('Browse.TButton',
                       font=('Segoe UI', 10),
                       background=UI_COLORS['ACCENT_PURPLE'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(15, 8))
        style.map('Browse.TButton',
                 background=[('active', '#5a67d8'), ('pressed', '#4c51bf')])
        
        # Configure modern report type buttons using constants
        style.configure('ReportCard.TButton',
                       font=('Segoe UI', 13, 'bold'),
                       background=UI_COLORS['BACKGROUND_WHITE'],
                       foreground=UI_COLORS['TEXT_SECONDARY'],
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 20))
        style.map('ReportCard.TButton',
                 background=[('active', UI_COLORS['BACKGROUND_LIGHT']), ('pressed', '#edf2f7')])
        
        style.configure('ReportCardActive.TButton',
                       font=('Segoe UI', 13, 'bold'),
                       background=UI_COLORS['PRIMARY_BLUE'],
                       foreground='white',
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(25, 15))
        style.map('ReportCardActive.TButton',
                 background=[('active', UI_COLORS['PRIMARY_BLUE_HOVER']), ('pressed', UI_COLORS['PRIMARY_BLUE_PRESSED'])])
        
        style.configure('ReportCardDisabled.TButton',
                       font=('Segoe UI', 13, 'bold'),
                       background=UI_COLORS['BACKGROUND_WHITE'],
                       foreground=UI_COLORS['TEXT_DISABLED'],
                       borderwidth=1,
                       relief='solid',
                       focuscolor='none',
                       padding=(25, 15))
        style.map('ReportCardDisabled.TButton',
                 background=[('active', '#f1f3f4'), ('pressed', '#e8eaed')])
        
        # Configure frames
        style.configure('Card.TFrame',
                       background='#ffffff',
                       relief='flat',
                       borderwidth=1)
        

    
    def create_card_buttons(self, parent):
        """Create modern card-style buttons"""
        # Create button frame with compact spacing
        button_frame = tk.Frame(parent, bg='#f8f9fa')
        button_frame.grid(row=0, column=0, pady=5)
        
        # Home Button
        self.home_button = ttk.Button(button_frame, 
                                     text="🏠\nHome",
                                     style='ReportCardActive.TButton',
                                     command=lambda: self.select_card('home'))
        self.home_button.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
        
        # Basic Report Button
        self.basic_button = ttk.Button(button_frame, 
                                     text="📊\nBasic Report",
                                     style='ReportCardDisabled.TButton',
                                     command=lambda: self.select_card('basic'))
        self.basic_button.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
        
        # Detailed Report Button
        self.detailed_button = ttk.Button(button_frame,
                                        text="📈\nDetailed Report", 
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('detailed'))
        self.detailed_button.grid(row=0, column=2, padx=10, pady=5, sticky="nsew")
        
        # Template Generator Button
        self.template_button = ttk.Button(button_frame,
                                        text="📋\nBVC Template", 
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('template'))
        self.template_button.grid(row=0, column=3, padx=10, pady=5, sticky="nsew")
        
        # Make columns equal width
        button_frame.grid_columnconfigure(0, weight=1, uniform="card")
        button_frame.grid_columnconfigure(1, weight=1, uniform="card")
        button_frame.grid_columnconfigure(2, weight=1, uniform="card")
        button_frame.grid_columnconfigure(3, weight=1, uniform="card")
        
        # Store references for updating styles
        self.cards = {
            'home': {'button': self.home_button},
            'basic': {'button': self.basic_button},
            'detailed': {'button': self.detailed_button},
            'template': {'button': self.template_button}
        }
        
        # Don't set initial selection here - will be set after sections are created
    
    def select_card(self, card_type):
        """Handle card selection with visual feedback"""
        self.report_type.set(card_type)
        
        # Update navigation button states
        self.update_nav_button_states(card_type)
        
        # Update UI based on selection
        self.update_ui_for_selection(card_type)
    
    def update_nav_button_states(self, active_card):
        """Update navigation button visual states"""
        if not hasattr(self, 'nav_buttons'):
            return
            
        # Reset all buttons to inactive state
        for card_name, button in self.nav_buttons.items():
            if card_name == 'basic':
                if card_name == active_card:
                    # Active basic button
                    button.configure(bg='#4299e1', fg='white', font=('Segoe UI', 11, 'bold'))
                else:
                    # Inactive basic button  
                    button.configure(bg='#f7fafc', fg='#4a5568', font=('Segoe UI', 11))
            else:
                if card_name == active_card:
                    # Active secondary button
                    button.configure(bg='#4299e1', fg='white', font=('Segoe UI', 10, 'bold'))
                else:
                    # Inactive secondary button
                    button.configure(bg='#ffffff', fg='#4a5568', font=('Segoe UI', 10))
    
    def get_report_description(self):
        """Get contextual description text for the current report type"""
        report_type = self.report_type.get()
        if report_type == 'basic':
            return "For M&T and Marmon Reports"
        elif report_type == 'detailed':
            return "For Cast Nylons Reports"
        else:
            return ""
    
    def update_ui_for_selection(self, card_type):
        """Update UI elements based on selected card type"""
        if card_type == 'template':
            # Show date input instead of file input for template generation
            self.show_date_input_ui()
        else:
            # Show normal file input UI
            self.show_file_input_ui()
        
    
    def create_file_input_section(self):
        """Create the file input UI section"""
        self.file_section = tk.Frame(self.input_section, bg='#f8f9fa')
        # Don't grid it initially - let the initial state logic handle visibility
        self.file_section.columnconfigure(0, weight=1)
        
        # Section header with context-specific description
        self.header_container = tk.Frame(self.file_section, bg='#f8f9fa')
        self.header_container.grid(row=0, column=0, pady=(10, 5))
        
        ttk.Label(self.header_container, text="📁 Input File", style='Header.TLabel', background='#f8f9fa').pack()
        
        # Add descriptive text based on report type
        description_text = self.get_report_description()
        if description_text:
            self.desc_label = tk.Label(self.header_container, text=description_text, 
                                 font=('Segoe UI', 9, 'italic'), 
                                 fg='#4a5568', bg='#f8f9fa')
            self.desc_label.pack(pady=(2, 0))
        
        file_frame = tk.Frame(self.file_section, bg='#f8f9fa')
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=15)
        file_frame.columnconfigure(0, weight=1)
        
        # File display with clean styling and drag-drop support - expandable width
        file_display_frame = tk.Frame(file_frame, bg='#ffffff', relief='flat', bd=1)
        file_display_frame.grid(row=0, column=0, padx=(0, 15), sticky=(tk.W, tk.E))
        
        # Create scrollable text widget for multiple file names
        import tkinter.scrolledtext as scrolledtext
        self.file_display = scrolledtext.ScrolledText(file_display_frame,
                                                     height=6,
                                                     width=50,
                                                     font=('Segoe UI', 9),
                                                     fg='#000000',
                                                     bg='#ffffff',
                                                     wrap=tk.WORD,
                                                     state='disabled',
                                                     borderwidth=0,
                                                     highlightthickness=0)
        self.file_display.pack(fill='both', expand=True, padx=6, pady=4)
        
        # Initialize with placeholder text
        self.file_display.config(state='normal')
        self.file_display.insert('1.0', "No files selected")
        self.file_display.config(state='disabled', fg='#6c757d')
        
        file_display_frame.grid_columnconfigure(0, weight=1)
        
        # Store reference to file display frame for updates
        self.file_display_frame = file_display_frame
        
        # Enable drag and drop
        self.setup_drag_drop(file_display_frame)
        
        browse_button = ttk.Button(file_frame, text="📂 Browse", 
                                 command=self.browse_file, style='Browse.TButton')
        browse_button.grid(row=0, column=1)
    
    def create_date_input_section(self):
        """Create the date input UI section for template generation"""
        self.date_section = tk.Frame(self.input_section, bg='#f8f9fa')
        # Don't grid it initially - let the initial state logic handle visibility
        self.date_section.columnconfigure(0, weight=1)
        
        # Section header
        ttk.Label(self.date_section, text="📅 Date Range for Template", style='Header.TLabel', background='#f8f9fa').grid(row=0, column=0, pady=(5, 3))
        
        date_frame = tk.Frame(self.date_section, bg='#f8f9fa')
        date_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 5), padx=8)
        date_frame.columnconfigure(0, weight=1)
        
        # Date selection with calendar widgets
        if CALENDAR_AVAILABLE:
            self._create_calendar_widgets(date_frame)
        else:
            self._create_fallback_date_entry(date_frame)
        
        # No separate button needed - main process button handles template generation
        
        # Initially hide the date section
        self.date_section.grid_remove()
    
   
            
    def show_file_input_ui(self):
        """Show file input UI and hide other sections"""
        if hasattr(self, 'file_section'):
            self.file_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=5)
            # Refresh the description text for the current mode
            self.refresh_file_section_description()
        if hasattr(self, 'date_section'):
            self.date_section.grid_remove()
        # Show process button and navigation bar for file input
        if hasattr(self, 'process_button'):
            self.process_button.grid()
        if hasattr(self, 'nav_bar'):
            self.nav_bar.grid()
        self.update_process_button_state()
    
    def refresh_file_section_description(self):
        """Refresh the description text in the file section header"""
        if not hasattr(self, 'file_section') or not hasattr(self, 'header_container'):
            return
        
        # Remove existing description label if it exists
        if hasattr(self, 'desc_label'):
            self.desc_label.destroy()
        
        # Add new description text based on current report type
        description_text = self.get_report_description()
        if description_text:
            self.desc_label = tk.Label(self.header_container, text=description_text, 
                                     font=('Segoe UI', 9, 'italic'), 
                                     fg='#4a5568', bg='#f8f9fa')
            self.desc_label.pack(pady=(2, 0))
    
    def _create_calendar_widgets(self, parent_frame):
        """Create compact horizontal calendar layout"""
        # Configure parent frame for better alignment
        parent_frame.grid_columnconfigure(0, weight=1)
        parent_frame.grid_rowconfigure(0, weight=1)
        
        # Main horizontal container with improved layout
        main_container = tk.Frame(parent_frame, bg='#ffffff')
        main_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=0, pady=0)
        main_container.grid_columnconfigure(0, weight=5)  # Calendars get more space
        main_container.grid_columnconfigure(1, weight=3)  # Controls get proportional space
        
        # Left side: Calendar container with enhanced layout
        calendar_section = tk.Frame(main_container, bg='#f8f9fa')
        calendar_section.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 15), pady=0)
        calendar_section.grid_columnconfigure(0, weight=1)
        calendar_section.grid_columnconfigure(1, weight=1)
        calendar_section.grid_rowconfigure(0, weight=1)
        
        # Set default dates first
        from datetime import timedelta
        today = datetime.now()
        default_start = today
        default_end = today + timedelta(days=4)
        
        # FROM Calendar (Left side)
        from_shadow = tk.Frame(calendar_section, bg='#e2e8f0')
        from_shadow.grid(row=0, column=0, padx=(8, 4), pady=8, sticky='nsew')
        
        from_frame = tk.Frame(from_shadow, bg='#ffffff', relief='flat', bd=0)
        from_frame.pack(padx=0, pady=0, fill='both', expand=True)
        
        from_header = tk.Frame(from_frame, bg='#4299e1', height=35)
        from_header.pack(fill='both', expand=True, pady=0)
        from_header.pack_propagate(False)
        
        tk.Label(from_header, text="✨ FROM DATE", font=('Segoe UI', 11, 'bold'), 
                fg='white', bg='#4299e1').pack(pady=8)
        
        self.start_calendar = Calendar(from_frame,
                                     selectmode='day',
                                     year=default_start.year,
                                     month=default_start.month,
                                     day=default_start.day,
                                     background='#4299e1',
                                     foreground='white',
                                     selectbackground='#ffd700',
                                     selectforeground='#1a202c',
                                     normalbackground='#ffffff',
                                     normalforeground='#2d3748',
                                     weekendbackground='#ebf8ff',
                                     weekendforeground='#2b6cb0',
                                     othermonthforeground='#a0aec0',
                                     othermonthbackground='#f7fafc',
                                     headersbackground='#bee3f8',
                                     headersforeground='#1a365d',
                                     font=('Segoe UI', 9),
                                     borderwidth=1,
                                     bordercolor='#e2e8f0',
                                     cursor='hand2')
        self.start_calendar.pack(padx=4, pady=(0, 4), fill='both', expand=True)
        self.start_calendar.bind('<<CalendarSelected>>', self.on_start_date_select)
        
        # TO Calendar (Right side)
        to_shadow = tk.Frame(calendar_section, bg='#e2e8f0')
        to_shadow.grid(row=0, column=1, padx=(4, 8), pady=8, sticky='nsew')
        
        to_frame = tk.Frame(to_shadow, bg='#ffffff', relief='flat', bd=0)
        to_frame.pack(padx=0, pady=0, fill='both', expand=True)
        
        to_header = tk.Frame(to_frame, bg='#38a169', height=35)
        to_header.pack(fill='both', expand=True, pady=0)
        to_header.pack_propagate(False)
        
        tk.Label(to_header, text="🎯 TO DATE", font=('Segoe UI', 11, 'bold'), 
                fg='white', bg='#38a169').pack(pady=8)
        
        self.end_calendar = Calendar(to_frame,
                                   selectmode='day',
                                   year=default_end.year,
                                   month=default_end.month,
                                   day=default_end.day,
                                   background='#38a169',
                                   foreground='white',
                                   selectbackground='#ffd700',
                                   selectforeground='#1a202c',
                                   normalbackground='#ffffff',
                                   normalforeground='#2d3748',
                                   weekendbackground='#f0fff4',
                                   weekendforeground='#276749',
                                   othermonthforeground='#a0aec0',
                                   othermonthbackground='#f7fafc',
                                   headersbackground='#9ae6b4',
                                   headersforeground='#1a365d',
                                   font=('Segoe UI', 10, 'bold'),
                                   borderwidth=1,
                                   bordercolor='#e2e8f0',
                                   cursor='hand2')
        self.end_calendar.pack(padx=8, pady=(0, 8), fill='both', expand=True)
        self.end_calendar.bind('<<CalendarSelected>>', self.on_end_date_select)
        
        # Right side: Controls container with proper weights
        controls_section = tk.Frame(main_container, bg='#f8f9fa')
        controls_section.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0), pady=0)
        controls_section.grid_columnconfigure(0, weight=1)
        controls_section.grid_rowconfigure(3, weight=1)  # Spacer row
        
        # Header
        header_frame = tk.Frame(controls_section, bg='#3182ce', height=50)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=0, pady=0)
        header_frame.grid_propagate(False)
        
        tk.Label(header_frame, text="📅 Select Date Range", 
                font=('Segoe UI', 13, 'bold'), fg='white', bg='#3182ce').pack(pady=15)
        
        # Date range text box with enhanced styling
        text_frame = tk.Frame(controls_section, bg='#f8f9fa')
        text_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=15, pady=(25, 15))
        
        # Enhanced label with icon
        label_frame = tk.Frame(text_frame, bg='#f8f9fa')
        label_frame.pack(fill='x', pady=(0, 8))
        
        tk.Label(label_frame, text="📅", font=('Segoe UI', 14), 
                fg='#4299e1', bg='#f8f9fa').pack(side='left', padx=(0, 5))
        
        tk.Label(label_frame, text="Date Range:", font=('Segoe UI', 12, 'bold'), 
                fg='#2d3748', bg='#f8f9fa').pack(side='left')
        
        # Enhanced textbox with better styling - no black border
        textbox_container = tk.Frame(text_frame, bg='#e2e8f0', relief='flat', bd=1)
        textbox_container.pack(fill='x', pady=(0, 5))
        
        self.date_range_entry = tk.Entry(textbox_container,
                                        font=('Segoe UI', 13, 'bold'),
                                        fg='#1a365d',
                                        bg='#ffffff',
                                        justify='center',
                                        relief='flat',
                                        bd=0,
                                        highlightthickness=2,
                                        highlightcolor='#4299e1',
                                        highlightbackground='#e2e8f0',
                                        insertbackground='#4299e1')
        self.date_range_entry.pack(fill='x', padx=6, pady=6)
        self.date_range_entry.bind('<KeyRelease>', self.on_date_range_text_change)
        self.date_range_entry.bind('<Return>', self.on_date_range_enter)
        self.date_range_entry.bind('<FocusIn>', self.on_date_range_focus_in)
        self.date_range_entry.bind('<FocusOut>', self.on_date_range_focus_out)
        
        # Generate button (moved here from bottom)
        button_frame = tk.Frame(controls_section, bg='#f8f9fa')
        button_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), padx=15, pady=(20, 15))
        
        # Create the template button for this layout
        self.template_button = ttk.Button(button_frame, text="📋 GENERATE TEMPLATE", 
                                         command=self.generate_template, style='ProcessButton.TButton', 
                                         state="disabled")
        self.template_button.pack(fill='x', pady=5)
        
        # Spacer row to push everything up
        tk.Frame(controls_section, bg='#f8f9fa').grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Initialize the date range display
        self.update_date_range_display()
    
    def _create_fallback_date_entry(self, parent_frame):
        """Create fallback text entry if calendar widget not available"""
        # Date input with clean styling
        date_input_frame = tk.Frame(parent_frame, bg='#ffffff', relief='flat', bd=1)
        date_input_frame.grid(row=0, column=0, padx=(0, 15), sticky=(tk.W, tk.E))
        date_input_frame.grid_columnconfigure(0, weight=1)
        
        # Date input field
        self.date_entry = tk.Entry(date_input_frame,
                                  font=('Segoe UI', 12),
                                  fg='#2d3748',
                                  bg='#ffffff',
                                  borderwidth=0,
                                  highlightthickness=0)
        self.date_entry.pack(fill='x', expand=True, padx=15, pady=12)
        
        # Add placeholder text
        self.date_entry.insert(0, "Enter date range (e.g., 08.04.25 - 08.08.25)")
        self.date_entry.config(fg='#a0aec0')
        
        # Bind events for placeholder behavior and Enter key
        self.date_entry.bind('<FocusIn>', self.on_date_entry_focus_in)
        self.date_entry.bind('<FocusOut>', self.on_date_entry_focus_out)
        self.date_entry.bind('<KeyRelease>', self.on_date_entry_change)
        self.date_entry.bind('<Return>', self.on_date_entry_enter)

    def show_date_input_ui(self):
        """Show date input UI and hide other sections"""
        if hasattr(self, 'date_section'):
            self.date_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5), padx=3)
        if hasattr(self, 'file_section'):
            self.file_section.grid_remove()
        # Hide main process button (template has its own), but show nav bar
        if hasattr(self, 'process_button'):
            self.process_button.grid_remove()
        if hasattr(self, 'nav_bar'):
            self.nav_bar.grid()
        self.update_process_button_state()
    
    def on_date_entry_focus_in(self, event):
        """Handle date entry focus in"""
        if self.date_entry.get() == "Enter date range (e.g., 08.04.25 - 08.08.25)":
            self.date_entry.delete(0, tk.END)
            self.date_entry.config(fg='#2d3748')
    
    def on_date_entry_focus_out(self, event):
        """Handle date entry focus out"""
        if not self.date_entry.get():
            self.date_entry.insert(0, "Enter date range (e.g., 08.04.25 - 08.08.25)")
            self.date_entry.config(fg='#a0aec0')
    
    def on_start_date_select(self, event=None):
        """Handle start date selection from calendar"""
        self.update_date_range_display()
        self.update_process_button_state()
    
    def on_end_date_select(self, event=None):
        """Handle end date selection from calendar"""
        self.update_date_range_display()
        self.update_process_button_state()
    
    def update_date_range_display(self):
        """Update the date range display in the text box"""
        if hasattr(self, 'start_calendar') and hasattr(self, 'end_calendar') and hasattr(self, 'date_range_entry'):
            try:
                start_date = self.start_calendar.selection_get()
                end_date = self.end_calendar.selection_get()
                
                # Format as MM.DD.YY - MM.DD.YY
                start_str = start_date.strftime('%m.%d.%y')
                end_str = end_date.strftime('%m.%d.%y')
                range_text = f"{start_str} - {end_str}"
                
                # Update text box without triggering events
                self.date_range_entry.delete(0, tk.END)
                self.date_range_entry.insert(0, range_text)
                self.date_range_entry.config(fg='#2d3748')
            except:
                self.date_range_entry.delete(0, tk.END)
                self.date_range_entry.insert(0, "Select dates")
                self.date_range_entry.config(fg='#a0aec0')
    
    def on_date_range_text_change(self, event=None):
        """Handle manual text changes in date range entry"""
        # Only parse if user typed a complete date range
        text = self.date_range_entry.get().strip()
        if " - " in text and len(text) >= 15:  # Basic format check
            try:
                self.parse_and_update_calendars(text)
            except:
                pass  # Invalid format, ignore
        self.update_process_button_state()
    
    def on_date_range_enter(self, event=None):
        """Handle Enter key in date range entry"""
        text = self.date_range_entry.get().strip()
        try:
            self.parse_and_update_calendars(text)
            if self.report_type.get() == 'template':
                self.generate_template()
        except:
            # Reset to calendar values if invalid
            self.update_date_range_display()
    
    def on_date_range_focus_in(self, event=None):
        """Handle focus in on date range entry"""
        if self.date_range_entry.get() == "Select dates":
            self.date_range_entry.delete(0, tk.END)
            self.date_range_entry.config(fg='#2d3748')
    
    def on_date_range_focus_out(self, event=None):
        """Handle focus out on date range entry"""
        if not self.date_range_entry.get().strip():
            self.update_date_range_display()
    
    def parse_and_update_calendars(self, text):
        """Parse date range text and update calendars"""
        if " - " in text:
            parts = text.split(" - ")
            if len(parts) == 2:
                from datetime import datetime
                try:
                    # Try multiple date formats
                    formats = ['%m.%d.%y', '%m/%d/%y', '%m-%d-%y', '%m.%d.%Y', '%m/%d/%Y']
                    start_date = None
                    end_date = None
                    
                    for fmt in formats:
                        try:
                            start_date = datetime.strptime(parts[0].strip(), fmt)
                            end_date = datetime.strptime(parts[1].strip(), fmt)
                            break
                        except:
                            continue
                    
                    if start_date and end_date:
                        # Update calendars
                        self.start_calendar.selection_set(start_date.date())
                        self.end_calendar.selection_set(end_date.date())
                        return True
                except:
                    pass
        raise ValueError("Invalid date format")
    
    def on_date_change(self, event=None):
        """Handle date change from calendar widgets or text entry (fallback)"""
        self.update_process_button_state()
    
    def on_date_entry_change(self, event):
        """Handle date entry text change (fallback mode)"""
        self.update_process_button_state()
    
    def on_date_entry_enter(self, event):
        """Handle Enter key press in date entry field (fallback mode)"""
        # Trigger template generation if valid input
        if self.report_type.get() == 'template':
            date_text = self.date_entry.get().strip()
            if date_text and date_text != "Enter date range (e.g., 08.04.25 - 08.08.25)":
                self.generate_template()
    
    def get_date_range_string(self):
        """Get formatted date range string from text box or calendar widgets"""
        if CALENDAR_AVAILABLE and hasattr(self, 'date_range_entry'):
            # Get from the synced text box
            text = self.date_range_entry.get().strip()
            if text and text != "Select dates":
                return text
            # Fallback to calendar if text box is empty
            elif hasattr(self, 'start_calendar') and hasattr(self, 'end_calendar'):
                try:
                    start = self.start_calendar.selection_get()
                    end = self.end_calendar.selection_get()
                    start_str = start.strftime('%m.%d.%y')
                    end_str = end.strftime('%m.%d.%y')
                    return f"{start_str} - {end_str}"
                except:
                    return ""
        elif hasattr(self, 'date_entry'):
            # Fallback to old text entry
            return self.date_entry.get().strip()
        else:
            return ""
    
    def generate_template(self):
        """Generate BVC template with the specified date range"""
        date_range = self.get_date_range_string()
        
        # Check if valid date range
        if not date_range or date_range == "Enter date range (e.g., 08.04.25 - 08.08.25)":
            messagebox.showwarning("Date Range Required", "Please select or enter a date range for the template.")
            return
        
        # Validate date format
        if not self.template_generator.validate_input(date_range):
            messagebox.showwarning("Invalid Format", "Please enter a valid date range (e.g., 08.04.25 - 08.08.25)")
            return
        
        try:
            # Ask user where to save the template
            default_name = f"MARMON BVC {date_range}.xlsx"
            output_file = filedialog.asksaveasfilename(
                title="Save Template As",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if output_file:
                # Generate the template
                file_path = self.template_generator.generate_template(date_range, output_file)
                messagebox.showinfo("Template Generated", 
                    f"BVC Template created successfully!\n\n"
                    f"File saved as:\n{os.path.basename(file_path)}")
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate template:\n{str(e)}")

    def create_widgets(self):
        """Create the main GUI widgets"""
        # Main container with compact styling
        main_frame = ttk.Frame(self.root, padding="15", style='Card.TFrame')
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Header section
        header_frame = tk.Frame(main_frame, bg='#ffffff', relief='flat')
        header_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)
        
        # Title with modern styling
        title_label = ttk.Label(header_frame, text="🚛 TMS Data Processor", style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(5, 10))
        
        # Initialize report type variable
        self.report_type = tk.StringVar(value="basic")
        
        # Compact Navigation Bar (always visible)
        self.nav_bar = tk.Frame(main_frame, bg='#e2e8f0', height=45)
        self.nav_bar.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5), padx=5)
        self.nav_bar.grid_columnconfigure(0, weight=1)
        self.create_navigation_bar()
        
        # Input Section - Dynamic (File or Date input based on selection)
        self.input_section = tk.Frame(main_frame, bg='#f8f9fa')
        self.input_section.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10), padx=5)
        self.input_section.columnconfigure(0, weight=1)
        
        # Create both file input and date input sections
        self.create_file_input_section()
        self.create_date_input_section()
        
        # Set correct initial state (start with file input visible, date input hidden)
        if hasattr(self, 'file_section'):
            self.file_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10), padx=5)
        if hasattr(self, 'date_section'):
            self.date_section.grid_remove()
        
        # Process Button with enhanced styling
        button_frame = tk.Frame(main_frame, bg='#ffffff')
        button_frame.grid(row=3, column=0, columnspan=3, pady=10)
        
        self.process_button = ttk.Button(button_frame, text="🚀 PROCESS FILE",
                                       command=self.process_file, style='ProcessButton.TButton', state="disabled")
        self.process_button.grid(row=0, column=0)

        # Stats Display Frame (below process button)
        self.stats_display_frame = tk.Frame(main_frame, bg=UI_COLORS['BACKGROUND_WHITE'], relief='ridge', bd=1)
        self.stats_display_frame.grid(row=4, column=0, columnspan=3, pady=(10, 0), padx=10, sticky=(tk.W, tk.E))

        # Initialize stats display
        self.update_savings_display()

        # Create main landing page
        # Set initial selection to basic report page
        self.root.after(1, lambda: self.select_card('basic'))
    
    def create_navigation_bar(self):
        """Create compact navigation bar for switching between pages"""
        nav_container = tk.Frame(self.nav_bar, bg='#e2e8f0')
        nav_container.pack(pady=8, padx=15)
        
        # Store button references for active state management
        self.nav_buttons = {}
        
        # Primary: Basic Report (larger, more prominent)
        self.nav_buttons['basic'] = tk.Button(nav_container, text="📊 Basic", font=('Segoe UI', 11, 'bold'), 
                            bg='#4299e1', fg='white', relief='flat', bd=0,
                            cursor='hand2', command=lambda: self.select_card('basic'),
                            activebackground='#3182ce', padx=15, pady=6)
        self.nav_buttons['basic'].pack(side='left', padx=(0, 8))
        
        # Secondary: Other options (smaller)
        self.nav_buttons['detailed'] = tk.Button(nav_container, text="📈 Detailed", font=('Segoe UI', 10), 
                               bg='#ffffff', fg='#4a5568', relief='flat', bd=1,
                               cursor='hand2', command=lambda: self.select_card('detailed'),
                               activebackground='#f7fafc', padx=10, pady=4)
        self.nav_buttons['detailed'].pack(side='left', padx=(0, 5))
        
        self.nav_buttons['template'] = tk.Button(nav_container, text="📋 Template", font=('Segoe UI', 10), 
                               bg='#ffffff', fg='#4a5568', relief='flat', bd=1,
                               cursor='hand2', command=lambda: self.select_card('template'),
                               activebackground='#f7fafc', padx=10, pady=4)
        self.nav_buttons['template'].pack(side='left')
        
    def setup_drag_drop(self, widget):
        """Setup drag and drop functionality for file selection"""
        def on_drag_enter(event):
            widget.configure(bg='#e6fffa')
            return "copy"
            
        def on_drag_leave(event):
            widget.configure(bg='#ffffff')
            
        def on_drop(event):
            widget.configure(bg='#ffffff')
            if hasattr(event, 'data'):
                files = event.data.split()
                if files:
                    # Filter for valid Excel files
                    excel_files = [f.strip('{}') for f in files if f.strip('{}').lower().endswith(('.xlsx', '.xls'))]
                    
                    if excel_files:
                        self.input_files = excel_files
                        self.update_file_display()
                        self.update_process_button_state()
                    else:
                        messagebox.showwarning("Invalid Files", "Please select Excel files (.xlsx or .xls)")
        
        try:
            # Try to set up tkinter DND if available
            widget.drop_target_register('DND_Files')
            widget.dnd_bind('<<DropEnter>>', on_drag_enter)
            widget.dnd_bind('<<DropLeave>>', on_drag_leave)
            widget.dnd_bind('<<Drop>>', on_drop)
        except:
            # DND not available, continue without it
            pass
            
    def auto_resize_window(self):
        """Dynamically resize window to fit content with padding"""
        self.root.update_idletasks()
        
        # Get the required size of all content
        required_width = self.root.winfo_reqwidth() + 40  # Add padding
        required_height = self.root.winfo_reqheight() + 60  # Add padding
        
        # Set reasonable limits
        min_width = 1100
        max_width = int(self.root.winfo_screenwidth() * 0.9)
        min_height = 600
        max_height = int(self.root.winfo_screenheight() * 0.85)
        
        # Constrain to limits
        width = max(min_width, min(required_width, max_width))
        height = max(min_height, min(required_height, max_height))
        
        # Center on screen
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        
        # Apply new size
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        self.root.minsize(min_width, min_height)
        
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def browse_file(self):
        """Browse for multiple input files"""
        file_paths = filedialog.askopenfilenames(
            title="Select TMS Excel Files (Multiple Selection Allowed)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_paths:
            self.input_files = list(file_paths)
            self.update_file_display()
            self.update_process_button_state()
    
    def update_file_display(self):
        """Update the file display to show all selected filenames with auto-sizing"""
        self.file_display.config(state='normal')
        self.file_display.delete('1.0', tk.END)
        
        if self.input_files:
            file_count = len(self.input_files)
            
            if file_count == 1:
                # Single file - show filename with checkmark
                filename = os.path.basename(self.input_files[0])
                display_text = f"✅ Selected: {filename}"
                self.file_display.insert('1.0', display_text)
                self.file_display.config(fg='#0d9488')
                # Set height for single file
                optimal_height = 3
            else:
                # Multiple files - show count and list all filenames
                header = f"✅ {file_count} files selected:\n\n"
                self.file_display.insert('1.0', header)
                
                # Add each filename on a new line
                for i, file_path in enumerate(self.input_files, 1):
                    filename = os.path.basename(file_path)
                    file_line = f"{i}. {filename}\n"
                    self.file_display.insert(tk.END, file_line)
                
                self.file_display.config(fg='#0d9488')
                # Calculate optimal height: header (2 lines) + files + padding
                optimal_height = min(max(file_count + 3, 5), 15)  # Min 5, max 15 lines
            
            # Auto-resize the display based on content
            self.file_display.config(height=optimal_height)
        else:
            self.file_display.insert('1.0', "No files selected")
            self.file_display.config(fg='#6c757d', height=3)
        
        self.file_display.config(state='disabled')
    
    def update_process_button_state(self):
        """Enable process button based on current selection and input state"""
        if self.report_type.get() == 'template':
            # Template mode - hide main process button completely
            if hasattr(self, 'process_button'):
                self.process_button.grid_remove()  # Hide the button
            
            # Update dedicated template button state
            if hasattr(self, 'template_button'):
                date_range = self.get_date_range_string()
                if CALENDAR_AVAILABLE and hasattr(self, 'date_range_entry'):
                    # Enhanced calendar mode with text box - check for valid input
                    if date_range and date_range not in ["Select dates", "Enter date range (e.g., 08.04.25 - 08.08.25)"]:
                        self.template_button.config(state="normal")
                    else:
                        self.template_button.config(state="disabled")
                elif date_range and date_range != "Enter date range (e.g., 08.04.25 - 08.08.25)":
                    # Fallback text entry mode - check for valid input
                    self.template_button.config(state="normal")
                else:
                    self.template_button.config(state="disabled")
        else:
            # File processing mode - show main process button
            if hasattr(self, 'process_button'):
                self.process_button.grid()  # Show the button
            
            # Check files
            if self.input_files:
                file_count = len(self.input_files)
                button_text = f"🚀 PROCESS {file_count} FILE{'S' if file_count > 1 else ''}"
                self.process_button.config(state="normal", text=button_text)
            else:
                self.process_button.config(state="disabled", text="🚀 PROCESS FILE")
    
    
    def process_file(self):
        """Process the selected files (template generation now handled by dedicated button)"""
        if self.is_processing:
            return
            
        # Only handle file processing mode (template has its own button now)
        if self.report_type.get() != 'template':
            # File processing mode
            if not self.input_files:
                return
                
            # Set processing state
            self.is_processing = True
            
            # Update UI for processing state
            file_count = len(self.input_files)
            self.process_button.config(state="disabled", text=f"⏳ PROCESSING {file_count} FILE{'S' if file_count > 1 else ''}...")
            
            # Start processing in separate thread
            thread = threading.Thread(target=self._process_file_thread)
            thread.daemon = True
            thread.start()
        
    def _process_file_thread(self):
        """Process files in background thread"""
        try:
            # Handle single vs multiple file processing differently
            if len(self.input_files) == 1:
                # Single file - ask for specific output file location
                input_name = os.path.splitext(os.path.basename(self.input_files[0]))[0]
                report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                default_name = f"{input_name}_processed_{report_type}.xlsx"
                
                output_file = filedialog.asksaveasfilename(
                    title="Save Processed File As",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                
                if not output_file:
                    self.root.after(0, self._reset_ui)
                    return
                
                output_folder = None  # Not used for single file
            else:
                # Multiple files - auto-create timestamped folder
                from datetime import datetime
                
                # Get base directory (use Desktop by default, or same location as first input file)
                if self.input_files:
                    base_dir = os.path.dirname(self.input_files[0])
                else:
                    base_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                
                # Create timestamped folder name
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                folder_name = f"TMS_Processed_{report_type}_{timestamp}"
                output_folder = os.path.join(base_dir, folder_name)
                
                # Create the folder
                try:
                    os.makedirs(output_folder, exist_ok=True)
                    print(f"Created output folder: {output_folder}")
                except Exception as e:
                    print(f"Error creating folder: {e}")
                    # Fallback to user selection if auto-creation fails
                    output_folder = filedialog.askdirectory(
                        title="Auto-folder creation failed. Select Folder to Save Processed Files"
                    )
                    if not output_folder:
                        self.root.after(0, self._reset_ui)
                        return
                
                output_file = None  # Will be generated per file
            
            # Select processor based on report type
            if self.report_type.get() == "basic":
                processor = self.basic_processor
            else:
                # Import detailed processor when needed
                if self.detailed_processor is None:
                    try:
                        from tms_detailed_processor import TMSDetailedDataProcessor
                        self.detailed_processor = TMSDetailedDataProcessor()
                    except ImportError as e:
                        self.root.after(0, lambda: messagebox.showerror("Error", 
                            f"Failed to load detailed processor: {e}"))
                        self.root.after(0, self._reset_ui)
                        return
                processor = self.detailed_processor
            
            # Process files
            processed_count = 0
            total_files = len(self.input_files)
            all_stats = []
            processed_files = []
            
            if len(self.input_files) == 1:
                # Single file processing
                try:
                    # Update button text
                    self.root.after(0, lambda: self.process_button.config(
                        text="⏳ PROCESSING FILE..."
                    ))
                    
                    # Process the data
                    processed_data = processor.clean_and_process_data(self.input_files[0])
                    
                    # Save the processed data (output_file already determined above)
                    processor.save_processed_data(output_file)
                    
                    # Collect stats
                    stats = processor.summary_stats.copy()
                    stats['filename'] = os.path.basename(self.input_files[0])
                    stats['output_file'] = os.path.basename(output_file)
                    all_stats.append(stats)
                    processed_files.append(output_file)
                    processed_count = 1
                    
                except Exception as file_error:
                    print(f"Error processing {self.input_files[0]}: {file_error}")
            else:
                # Multiple files processing
                for i, input_file in enumerate(self.input_files, 1):
                    try:
                        # Update button text with progress
                        self.root.after(0, lambda: self.process_button.config(
                            text=f"⏳ PROCESSING {i}/{total_files}..."
                        ))
                        
                        # Process the data
                        processed_data = processor.clean_and_process_data(input_file)
                        
                        # Generate output filename for each file
                        input_name = os.path.splitext(os.path.basename(input_file))[0]
                        report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                        file_output = os.path.join(output_folder, f"{input_name}_processed_{report_type}.xlsx")
                        
                        # Save the processed data
                        processor.save_processed_data(file_output)
                        
                        # Collect stats
                        stats = processor.summary_stats.copy()
                        stats['filename'] = os.path.basename(input_file)
                        stats['output_file'] = os.path.basename(file_output)
                        all_stats.append(stats)
                        processed_files.append(file_output)
                        processed_count += 1
                        
                    except Exception as file_error:
                        print(f"Error processing {input_file}: {file_error}")
                        continue
            
            # Show success message
            if processed_count > 0:
                total_loads = sum(stat['total_loads'] for stat in all_stats)
                total_savings = sum(stat['total_potential_savings'] for stat in all_stats)

                # Save to savings history
                combined_stats = {
                    'total_potential_savings': total_savings,
                    'total_loads': total_loads,
                    'percentage_savings': sum(stat['percentage_savings'] for stat in all_stats) / len(all_stats) if all_stats else 0,
                    'loads_with_savings': sum(stat['loads_with_savings'] for stat in all_stats)
                }
                report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                self.save_savings_history(combined_stats, report_type, len(self.input_files))
                
                if len(self.input_files) == 1:
                    # Single file success message
                    self.root.after(0, lambda: messagebox.showinfo("Processing Complete", 
                        f"🎉 Successfully processed file!\n\n"
                        f"📊 Total Loads: {total_loads:,}\n"
                        f"💵 Total Potential Savings: ${total_savings:,.2f}\n\n"
                        f"📁 File saved as:\n{os.path.basename(output_file)}"))
                else:
                    # Multiple files success message
                    files_list = "\n".join([f"✅ {stat['output_file']}" for stat in all_stats])
                    
                    self.root.after(0, lambda: messagebox.showinfo("Batch Processing Complete", 
                        f"🎉 Successfully processed {processed_count}/{total_files} files!\n\n"
                        f"📊 Total Loads: {total_loads:,}\n"
                        f"💵 Total Potential Savings: ${total_savings:,.2f}\n\n"
                        f"📁 Files saved to:\n{output_folder}\n\n"
                        f"Processed files:\n{files_list}"))
            else:
                self.root.after(0, lambda: messagebox.showerror("Error", 
                    "Failed to process any files. Please check your input files."))
            
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred:\n{error_msg}"))
        
        finally:
            self.root.after(0, self._reset_ui)
            
    def _reset_ui(self):
        """Reset UI to normal state"""
        self.is_processing = False
        self.process_button.config(state="normal", text="🚀 PROCESS FILE")
        self.update_process_button_state()

    def load_savings_history(self):
        """Load savings history from JSON file"""
        try:
            if self.savings_history_file.exists():
                with open(self.savings_history_file, 'r') as f:
                    return json.load(f)
            return []
        except Exception as e:
            print(f"Warning: Could not load savings history: {e}")
            return []

    def save_savings_history(self, stats, report_type, file_count):
        """Save current processing stats to history"""
        try:
            # Create new entry
            entry = {
                'timestamp': datetime.now().isoformat(),
                'report_type': report_type,
                'file_count': file_count,
                'total_potential_savings': stats.get('total_potential_savings', 0),
                'total_loads': stats.get('total_loads', 0),
                'percentage_savings': stats.get('percentage_savings', 0),
                'loads_with_savings': stats.get('loads_with_savings', 0)
            }

            # Add to history
            self.savings_history.insert(0, entry)  # Add to beginning

            # Keep only last 10 entries
            self.savings_history = self.savings_history[:10]

            # Save to file
            with open(self.savings_history_file, 'w') as f:
                json.dump(self.savings_history, f, indent=2)

            # Update UI display
            self.update_savings_display()

        except Exception as e:
            print(f"Warning: Could not save savings history: {e}")

    def update_savings_display(self):
        """Update the savings statistics display"""
        if hasattr(self, 'stats_display_frame'):
            # Clear existing content
            for widget in self.stats_display_frame.winfo_children():
                widget.destroy()

            if self.savings_history:
                # Calculate recent totals
                recent_savings = sum(entry['total_potential_savings'] for entry in self.savings_history)
                recent_loads = sum(entry['total_loads'] for entry in self.savings_history)

                # Create stats display
                stats_label = tk.Label(
                    self.stats_display_frame,
                    text=f"📊 Recent Stats (Last {len(self.savings_history)} uploads)",
                    font=('Segoe UI', 10, 'bold'),
                    bg=UI_COLORS['BACKGROUND_WHITE'],
                    fg=UI_COLORS['TEXT_PRIMARY']
                )
                stats_label.pack(pady=(5, 2))

                savings_label = tk.Label(
                    self.stats_display_frame,
                    text=f"💵 Total Potential Savings: ${recent_savings:,.2f}",
                    font=('Segoe UI', 9),
                    bg=UI_COLORS['BACKGROUND_WHITE'],
                    fg=UI_COLORS['SUCCESS_GREEN']
                )
                savings_label.pack()

                loads_label = tk.Label(
                    self.stats_display_frame,
                    text=f"📦 Total Loads Processed: {recent_loads:,}",
                    font=('Segoe UI', 9),
                    bg=UI_COLORS['BACKGROUND_WHITE'],
                    fg=UI_COLORS['TEXT_SECONDARY']
                )
                loads_label.pack()

                # Show last upload info
                if self.savings_history:
                    last_upload = self.savings_history[0]
                    last_date = datetime.fromisoformat(last_upload['timestamp']).strftime('%m/%d %I:%M%p')
                    last_label = tk.Label(
                        self.stats_display_frame,
                        text=f"🕒 Last Upload: {last_date} (${last_upload['total_potential_savings']:,.2f})",
                        font=('Segoe UI', 8),
                        bg=UI_COLORS['BACKGROUND_WHITE'],
                        fg=UI_COLORS['TEXT_MUTED']
                    )
                    last_label.pack()

def main():
    root = tk.Tk()
    app = ModernTMSProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()