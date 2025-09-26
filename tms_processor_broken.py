import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import time
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# Import UI styles from the new modular UI components
try:
    from ui import COLORS as UI_COLORS
except ImportError:
    # Fallback UI color constants for consistent theming
    UI_COLORS = {
        # Primary colors
        'PRIMARY_BLUE': '#4299e1',
        'PRIMARY_BLUE_HOVER': '#3182ce',
        'PRIMARY_BLUE_PRESSED': '#2c5aa0',

        # Background colors
        'BACKGROUND_WHITE': '#ffffff',
        'BACKGROUND_GRAY': '#f8f9fa',
        'BACKGROUND_LIGHT': '#f7fafc',
        'BACKGROUND_BORDER': '#e2e8f0',

        # Text colors
        'TEXT_PRIMARY': '#2d3748',
        'TEXT_SECONDARY': '#4a5568',
        'TEXT_MUTED': '#718096',
        'TEXT_DISABLED': '#a0aec0',

        # Status colors
        'SUCCESS_GREEN': '#38a169',
        'SUCCESS_LIGHT': '#e6fffa',
        'WARNING_YELLOW': '#ffd700',
        'ERROR_RED': '#e53e3e',

        # Navigation colors
        'NAV_INACTIVE': '#ffffff',
        'NAV_ACTIVE': '#4299e1',

        # Legacy colors for backward compatibility
        'ACCENT_PURPLE': '#667eea'
    }

# Import our new modules
try:
    from config import tms_config
    from logger_config import main_logger, data_logger, gui_logger, ProgressLogger
    from validators import tms_validator
except ImportError as e:
    print(f"Warning: Could not import enhanced modules: {e}")
    print("Falling back to basic functionality...")
    # Create mock objects for backward compatibility
    class MockConfig:
        def get(self, key, default=None):
            defaults = {
                'data_structure.default_header_row': 8,
                'data_structure.default_data_start_row': 11,
                'data_structure.min_data_columns': 5,
                'data_structure.expected_columns': 21,
                'business_rules.min_non_empty_values': 5
            }
            return defaults.get(key, default)

    class MockLogger:
        def info(self, msg, **kwargs): print(f"INFO: {msg}")
        def error(self, msg, **kwargs): print(f"ERROR: {msg}")
        def warning(self, msg, **kwargs): print(f"WARNING: {msg}")
        def debug(self, msg, **kwargs): print(f"DEBUG: {msg}")
        def log_processing_step(self, step, details=None): self.info(f"Processing: {step}")
        def log_data_stats(self, stats, prefix=""): self.info(f"{prefix}: {stats}")
        def log_performance(self, op, duration, records=None): self.info(f"{op}: {duration:.2f}s")

    class MockValidator:
        def run_full_validation(self, file_path):
            return {'overall_valid': True, 'validation_steps': {}}

    tms_config = MockConfig()
    main_logger = data_logger = gui_logger = MockLogger()
    tms_validator = MockValidator()

    def ProgressLogger(logger, total, operation):
        class MockProgress:
            def update(self, inc=1): pass
            def complete(self): pass
        return MockProgress()


class ModernTMSProcessor:
    """Enhanced TMS Processor with comprehensive validation and error handling"""

    def __init__(self):
        self.logger = main_logger
        self.data_logger = data_logger
        self.config = tms_config

        # Data storage
        self.raw_data = None
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}
        self.validation_results = None

        # TL carriers requiring special processing (from Basic_Processor)
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'ONX LOGISTICS INC'
        }

        # Performance tracking
        self.processing_start_time = None
        self.processing_stats = {}

        self.logger.info("ModernTMSProcessor initialized with enhanced features")

    def _extract_title_info(self, df_raw):
        """Extract title and report information from the top rows"""
        title_info = {}

        try:
            # Extract report title with bounds checking
            if len(df_raw) > 1 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[1, 1]):
                title_info['report_title'] = str(df_raw.iloc[1, 1])

            # Extract company name
            if len(df_raw) > 3 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[3, 1]):
                title_info['company_name'] = str(df_raw.iloc[3, 1])

            # Extract date range
            if len(df_raw) > 5 and len(df_raw.columns) > 1 and not pd.isna(df_raw.iloc[5, 1]):
                title_info['date_range'] = str(df_raw.iloc[5, 1])
        except (IndexError, KeyError):
            # If extraction fails, continue with empty title_info
            pass

        return title_info

    def _detect_data_structure(self, df_raw):
        """Intelligently detect header and data start positions"""
        header_row = self.DEFAULT_HEADER_ROW
        data_start_row = self.DEFAULT_DATA_START_ROW

        # Look for header indicators in different rows
        header_indicators = ['Load No.', 'Carrier', 'Service Type', 'Ship Date']

        for row_idx in range(5, min(15, len(df_raw))):
            row_data = df_raw.iloc[row_idx].dropna().astype(str).tolist()
            row_str = ' '.join(row_data).lower()

            # Check if this row contains header-like content
            matches = sum(1 for indicator in header_indicators if indicator.lower() in row_str)
            if matches >= 2:  # Found at least 2 header indicators
                header_row = row_idx
                data_start_row = row_idx + 2  # Skip potential blank row
                break

        return header_row, data_start_row

    def _remove_duplicate_headers(self, df):
        """Remove duplicate header rows that appear in the middle of data"""
        # Look for rows that contain header-like text
        header_indicators = ['Load No.', 'Carrier', 'Service Type']

        rows_to_drop = []
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.dropna().tolist()])
            if any(indicator in row_str for indicator in header_indicators):
                # Check if this looks like a header row (not actual data)
                if not any(str(val).startswith('A') and str(val)[1:].isdigit()
                          for val in row.dropna().tolist()):
                    rows_to_drop.append(idx)

        return df.drop(rows_to_drop)

    def clean_and_process_data(self, file_path: str) -> pd.DataFrame:
        """Main function to clean and process the TMS Excel file with comprehensive validation"""
        self.processing_start_time = time.time()
        self.logger.log_processing_step(
            "Starting TMS data processing",
            {'file': Path(file_path).name})

        try:
            # Step 1: Comprehensive validation
            self.validation_results = tms_validator.run_full_validation(file_path)

            if not self.validation_results['overall_valid']:
                failed_steps = [
                    step for step, result in self.validation_results['validation_steps'].items()
                    if not result.get('valid', False)
                ]
                raise ValueError(f"File validation failed. Issues: {failed_steps}")

            # Step 2: Load and validate Excel data
            self.logger.log_processing_step("Loading Excel file")
            df_raw = pd.read_excel(file_path, sheet_name=0, header=None)

            self.logger.log_data_stats({
                'raw_rows': len(df_raw),
                'raw_columns': len(df_raw.columns),
                'file_size_mb': Path(file_path).stat().st_size / (1024*1024)
            }, "RAW_DATA")

            # Step 3: Extract metadata
            self.logger.log_processing_step("Extracting title information")
            self.title_info = self._extract_title_info(df_raw)

            # Step 4: Use validation results for structure detection
            header_info = (self.validation_results['validation_steps']
                          ['header_detection']['details'])
            header_row = header_info.get(
                'header_row',
                self.config.get('data_structure.default_header_row', 8))
            data_start_row = header_info.get(
                'data_start_row',
                self.config.get('data_structure.default_data_start_row', 11))

            self.logger.log_processing_step("Data structure detected", {
                'header_row': header_row,
                'data_start_row': data_start_row,
                'confidence': header_info.get('confidence_score', 0)
            })

            # Get headers
            headers = df_raw.iloc[header_row].dropna().tolist()

            # Step 5: Extract and clean data with progress tracking
            self.logger.log_processing_step("Extracting data rows")
            data_df = df_raw.iloc[data_start_row:].copy()

            # Remove completely empty rows and duplicate header rows with logging
            initial_rows = len(data_df)
            data_df = data_df.dropna(how='all')
            empty_rows_removed = initial_rows - len(data_df)

            data_df = self._remove_duplicate_headers(data_df)
            duplicate_headers_removed = initial_rows - empty_rows_removed - len(data_df)

            self.logger.log_data_stats({
                'initial_data_rows': initial_rows,
                'empty_rows_removed': empty_rows_removed,
                'duplicate_headers_removed': duplicate_headers_removed,
                'remaining_rows': len(data_df)
            }, "DATA_CLEANING")

            # Reset index after dropping rows
            data_df = data_df.reset_index(drop=True)

            # Step 6: Column extraction with intelligent selection
            self.logger.log_processing_step("Extracting relevant columns")
            max_cols = min(22, len(data_df.columns))
            relevant_columns = list(range(2, max_cols + 1))
            data_df = data_df.iloc[:, relevant_columns]

            self.logger.log_data_stats({
                'total_available_columns': len(df_raw.columns),
                'relevant_columns_selected': len(relevant_columns),
                'max_expected_columns': 22
            }, "COLUMN_EXTRACTION")

            # Set proper column names with full descriptive headers
            base_column_names = [
                'Load No.', 'Ship Date', 'Origin City', 'Origin State', 'Origin Postal',
                'Destination City', 'Destination State', 'Destination Postal',
                'Selected Carrier', 'Selected Service Type', 'Selected Transit Days',
                'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost',
                'Least Cost Carrier', 'Least Cost Service Type', 'Least Cost Transit Days',
                'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost',
                'Potential Savings'
            ]

            # Ensure column names match the actual extracted columns
            if len(base_column_names) != len(data_df.columns):
                print(f"Warning: Column count mismatch. Expected {len(base_column_names)}, "
                      f"got {len(data_df.columns)}")
                # Adjust column names to match actual columns
                if len(data_df.columns) < len(base_column_names):
                    column_names = base_column_names[:len(data_df.columns)]
                else:
                    # Add generic names for extra columns
                    extra_columns = [
                        f'Column_{i}' for i in range(len(base_column_names), len(data_df.columns))
                    ]
                    column_names = base_column_names + extra_columns
            else:
                column_names = base_column_names

            data_df.columns = column_names

            # Step 7: Enhanced data type cleaning with validation
            self.logger.log_processing_step("Cleaning and validating data types")
            cleaning_start = time.time()
            data_df = self._clean_data_types_enhanced(data_df)
            cleaning_time = time.time() - cleaning_start
            self.logger.log_performance("Data type cleaning", cleaning_time, len(data_df))

            # Enhanced row filtering with detailed logging
            self.logger.log_processing_step("Filtering invalid rows")
            pre_filter_count = len(data_df)

            # Remove rows where Load No. is missing or empty
            if 'Load No.' in data_df.columns:
                data_df = data_df.dropna(subset=['Load No.'])
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != '']
                data_df = data_df[data_df['Load No.'].astype(str).str.strip() != 'nan']
            else:
                self.logger.warning("Load No. column not found - skipping Load No. validation")

            # Remove any remaining rows that are mostly empty
            min_values = self.config.get('business_rules.min_non_empty_values', 5)
            data_df = data_df.dropna(thresh=min_values)

            # Reset index and log filtering results
            data_df = data_df.reset_index(drop=True)
            rows_filtered = pre_filter_count - len(data_df)

            self.logger.log_data_stats({
                'rows_before_filtering': pre_filter_count,
                'rows_after_filtering': len(data_df),
                'rows_removed': rows_filtered,
                'filter_rate': (f"{(rows_filtered/pre_filter_count*100):.1f}%"
                               if pre_filter_count > 0 else "0%")
            }, "ROW_FILTERING")

            # Apply business logic rules with enhanced tracking
            self.logger.log_processing_step("Applying business logic rules")
            business_start = time.time()
            data_df = self._apply_business_logic_enhanced(data_df)
            business_time = time.time() - business_start
            self.logger.log_performance("Business logic application", business_time, len(data_df))

            # Sort by Destination City with error handling
            self.logger.log_processing_step("Sorting data")
            if 'Destination City' in data_df.columns:
                data_df = data_df.sort_values('Destination City', na_position='last')
            else:
                self.logger.warning("Destination City column not found - skipping sort")
                # Try alternative column names
                destination_cols = [
                    col for col in data_df.columns
                    if 'destination' in col.lower() and 'city' in col.lower()
                ]
                if destination_cols:
                    data_df = data_df.sort_values(destination_cols[0], na_position='last')
                    self.logger.info(f"Sorted by alternative column: {destination_cols[0]}")

            # Step 8: Calculate summary statistics
            self.logger.log_processing_step("Calculating summary statistics")
            self._calculate_summary_stats(data_df)

            # Step 9: Final processing metrics
            processing_time = time.time() - self.processing_start_time
            self.processing_stats = {
                'total_time': processing_time,
                'records_processed': len(data_df),
                'processing_rate': len(data_df) / processing_time if processing_time > 0 else 0
            }

            self.logger.log_performance(
                "Total TMS processing",
                processing_time,
                len(data_df)
            )

            self.processed_data = data_df
            return data_df

        except (FileNotFoundError, PermissionError) as e:
            self.logger.error("File access error", exception=e, file_path=file_path)
            raise FileNotFoundError(f"Cannot access file: {str(e)}")
        except (pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            self.logger.error("Excel parsing error", exception=e, file_path=file_path)
            raise ValueError(f"Invalid Excel file format: {str(e)}")
        except ValueError as e:
            # Re-raise validation errors
            self.logger.error("Validation error", exception=e)
            raise
        except Exception as e:
            self.logger.error("Unexpected processing error", exception=e, file_path=file_path)
            raise RuntimeError(f"Error processing file: {str(e)}")

    def _clean_data_types_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhanced data type cleaning with comprehensive validation and logging"""
        df = df.copy()
        cleaning_stats = {'columns_processed': 0, 'conversion_failures': 0}

        # Convert numeric columns with enhanced error tracking
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost',
            'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost',
            'Least Cost Accessorial Cost', 'Least Cost Total Cost',
            'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1
                original_nulls = df[col].isnull().sum()
                df[col] = pd.to_numeric(df[col], errors='coerce')
                new_nulls = df[col].isnull().sum()
                conversion_failures = new_nulls - original_nulls
                if conversion_failures > 0:
                    cleaning_stats['conversion_failures'] += conversion_failures
                    self.data_logger.warning(
                        f"Failed to convert {conversion_failures} values "
                        f"in {col} to numeric")

        # Ensure PS column is properly numeric and handle any string values
        if 'PS' in df.columns:
            # First try to convert to numeric, handling any string values
            df['PS'] = pd.to_numeric(df['PS'], errors='coerce')
            # Fill any NaN values with 0
            df['PS'] = df['PS'].fillna(0)

        # Convert date column with enhanced error handling
        if 'Ship Date' in df.columns:
            cleaning_stats['columns_processed'] += 1
            original_nulls = df['Ship Date'].isnull().sum()
            date_series = pd.to_datetime(df['Ship Date'], errors='coerce')
            new_nulls = date_series.isnull().sum()
            date_failures = new_nulls - original_nulls
            if date_failures > 0:
                cleaning_stats['conversion_failures'] += date_failures
                self.data_logger.warning(
                    f"Failed to convert {date_failures} date values in Ship Date"
                )

            date_format = self.config.get('formatting.date_format', '%m/%d/%y')
            df['Ship Date'] = date_series.dt.strftime(date_format)

        # Clean string columns with tracking
        string_columns = [
            'Load No.', 'Origin City', 'Origin State', 'Origin Postal',
            'Destination City', 'Destination State', 'Destination Postal',
            'Selected Carrier', 'Selected Service Type',
            'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                cleaning_stats['columns_processed'] += 1

                # Special handling for Least Cost Carrier and Service Type
                if col in ['Least Cost Carrier', 'Least Cost Service Type']:
                    # Check for numeric zeros that should be converted to empty strings
                    before_cleaning = df[col].copy()
                    df[col] = df[col].astype(str).str.strip()

                    # Convert numeric zeros to empty strings
                    df[col] = df[col].replace(['0', '0.0', 'nan', 'None'], '')

                    # Log the cleaning results for debugging
                    zero_count = (before_cleaning == 0).sum()
                    nan_count = before_cleaning.isna().sum()
                    if zero_count > 0 or nan_count > 0:
                        self.data_logger.info(
                            f"Column {col}: Found {zero_count} zeros and "
                            f"{nan_count} NaN values - converted to empty strings")
                else:
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace('nan', '')

        self.data_logger.log_data_stats(cleaning_stats, "TYPE_CLEANING")
        return df

    def _apply_business_logic_enhanced(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply TMS business logic rules with enhanced tracking and validation"""
        df = df.copy()
        business_stats = {
            'same_carrier_rule_applied': 0,
            'empty_data_rule_applied': 0,
            'negative_savings_rule_applied': 0,
            'tl_carrier_rule_applied': 0,
            'ddi_carrier_rule_applied': 0,
            'total_rows_affected': 0
        }

        try:
            # Ensure PS column is numeric from the start to avoid comparison errors
            if 'PS' in df.columns:
                df['PS'] = pd.to_numeric(df['PS'], errors='coerce').fillna(0)
            else:
                print("Warning: PS column not found in dataframe")
                print(f"Available columns: {df.columns.tolist()}")

            # Rule 1: Same Carriers - Set Potential Savings to 0 (Enhanced)
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                same_carrier_mask = (
                    (df['Selected Carrier'].astype(str) ==
                     df['Least Cost Carrier'].astype(str)) &
                    (df['Selected Carrier'].notna()) &
                    (df['Least Cost Carrier'].notna()) &
                    (df['Selected Carrier'].astype(str) != '') &
                    (df['Least Cost Carrier'].astype(str) != '') &
                    (df['Selected Carrier'].astype(str) != 'nan') &
                    (df['Least Cost Carrier'].astype(str) != 'nan')
                )

                same_carrier_count = same_carrier_mask.sum()
                business_stats['same_carrier_rule_applied'] = same_carrier_count

                if 'Potential Savings' in df.columns and same_carrier_count > 0:
                    default_savings = self.config.get('business_rules.same_carrier_savings', 0.0)
                    df.loc[same_carrier_mask, 'Potential Savings'] = default_savings
                    self.data_logger.info(
                        f"Applied same carrier rule to {same_carrier_count} rows"
                    )
            else:
                self.data_logger.warning(
                    "Cannot apply same carrier rule - required columns missing"
                )

            # Rule 2: Empty Least Cost - Copy Selected data and set savings to 0 (Enhanced)
            if 'Least Cost Carrier' in df.columns:
                empty_least_cost_mask = (
                    df['Least Cost Carrier'].isna() |
                    (df['Least Cost Carrier'].astype(str) == '') |
                    (df['Least Cost Carrier'].astype(str) == 'nan')
                )

                empty_count = empty_least_cost_mask.sum()
                business_stats['empty_data_rule_applied'] = empty_count

                if empty_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, empty_least_cost_mask, column_pairs)

                    if 'Potential Savings' in df.columns:
                        df.loc[empty_least_cost_mask, 'Potential Savings'] = 0

                    self.data_logger.info(f"Applied empty data rule to {empty_count} rows")
            else:
                self.data_logger.warning(
                    "Cannot apply empty data rule - Least Cost Carrier column missing"
                )

            # Rule 3: Negative Savings - Copy Selected data and set savings to 0 (Enhanced)
            if 'Potential Savings' in df.columns:
                # Ensure Potential Savings is numeric before comparison
                ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
                negative_savings_mask = ps_numeric < 0
                negative_count = negative_savings_mask.sum()
                business_stats['negative_savings_rule_applied'] = negative_count

                if negative_count > 0:
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, negative_savings_mask, column_pairs)
                    df.loc[negative_savings_mask, 'Potential Savings'] = 0
                    self.data_logger.info(
                        f"Applied negative savings rule to {negative_count} rows"
                    )
            else:
                self.data_logger.warning(
                    "Cannot apply negative savings rule - Potential Savings column missing"
                )

            # Rule 4: TL Carriers - Copy selected to least cost and zero out savings
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                # Find rows where Selected Carrier or Least Cost Carrier is in TL list
                tl_mask = (
                    df['Selected Carrier'].astype(str).str.upper().isin(
                        [carrier.upper() for carrier in self.TL_CARRIERS]) |
                    df['Least Cost Carrier'].astype(str).str.upper().isin(
                        [carrier.upper() for carrier in self.TL_CARRIERS])
                )

                tl_count = tl_mask.sum()
                business_stats['tl_carrier_rule_applied'] = tl_count

                if tl_count > 0:
                    # Debug: Log which carriers were found
                    tl_carriers_found = df.loc[tl_mask, ['Selected Carrier', 'Least Cost Carrier']].drop_duplicates()
                    self.data_logger.info(
                        f"TL carriers found: {tl_carriers_found.to_dict('records')}")

                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, tl_mask, column_pairs)

                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        # Debug: Log before and after values
                        before_savings = df.loc[tl_mask, 'Potential Savings'].tolist()
                        df.loc[tl_mask, 'Potential Savings'] = 0
                        after_savings = df.loc[tl_mask, 'Potential Savings'].tolist()
                        self.data_logger.info(
                            f"TL Carrier savings - Before: {before_savings}, "
                            f"After: {after_savings}")
                    else:
                        self.data_logger.warning(
                            "Potential Savings column not found for TL carrier rule")

                    self.data_logger.info(
                        f"Applied TL carrier rule to {tl_count} rows "
                        f"(LANDSTAR/SMARTWAY)")
                else:
                    self.data_logger.info("No TL carriers found in data")

            # Rule 5: DALKO DEFENDER INSURANCE pattern matching (from Basic_Processor)
            if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
                dalko_matches = []

                for idx, row in df.iterrows():
                    selected = str(row['Selected Carrier']).strip().upper()
                    least_cost = str(row['Least Cost Carrier']).strip().upper()

                    # Skip empty or nan values
                    if selected in ['', 'NAN', 'NONE'] or least_cost in ['', 'NAN', 'NONE']:
                        continue

                    # Check for DALKO DEFENDER INSURANCE patterns
                    # Pattern 1: "DALKO DEFENDER INSURANCE/XXX" where least cost is "XXX"
                    if 'DALKO DEFENDER INSURANCE/' in selected:
                        carrier_after_slash = selected.split(
                            'DALKO DEFENDER INSURANCE/')[-1].strip()
                        if carrier_after_slash == least_cost:
                            dalko_matches.append(idx)

                    # Pattern 2: "XXX/DALKO DEFENDER INSURANCE" where least cost is "XXX"
                    elif '/DALKO DEFENDER INSURANCE' in selected:
                        carrier_before_slash = selected.split(
                            '/DALKO DEFENDER INSURANCE')[0].strip()
                        if carrier_before_slash == least_cost:
                            dalko_matches.append(idx)

                dalko_match_count = len(dalko_matches)
                business_stats['ddi_carrier_rule_applied'] = dalko_match_count

                if dalko_match_count > 0:
                    dalko_mask = df.index.isin(dalko_matches)

                    # Copy selected carrier data to least cost columns
                    column_pairs = [
                        ('Selected Carrier', 'Least Cost Carrier'),
                        ('Selected Service Type', 'Least Cost Service Type'),
                        ('Selected Transit Days', 'Least Cost Transit Days'),
                        ('Selected Freight Cost', 'Least Cost Freight Cost'),
                        ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
                        ('Selected Total Cost', 'Least Cost Total Cost')
                    ]
                    self._copy_selected_to_least_cost(df, dalko_mask, column_pairs)

                    # Set Potential Savings to 0
                    if 'Potential Savings' in df.columns:
                        df.loc[dalko_mask, 'Potential Savings'] = 0

                    self.data_logger.info(
                        f"Applied DALKO DEFENDER INSURANCE rule to "
                        f"{dalko_match_count} rows")
            else:
                self.data_logger.warning(
                    "Cannot apply DALKO DEFENDER INSURANCE rule - "
                    "required columns missing")

            # Calculate total affected rows
            business_stats['total_rows_affected'] = (
                business_stats['same_carrier_rule_applied'] +
                business_stats['empty_data_rule_applied'] +
                business_stats['negative_savings_rule_applied'] +
                business_stats['tl_carrier_rule_applied'] +
                business_stats['ddi_carrier_rule_applied']
            )

            self.data_logger.log_data_stats(business_stats, "BUSINESS_LOGIC")

        except Exception as e:
            self.data_logger.error("Business logic application failed", exception=e,
                                   df_shape=df.shape, df_columns=df.columns.tolist())
            raise RuntimeError(f"Business logic error: {str(e)}")

        return df

    def _copy_selected_to_least_cost(self, df, mask, column_pairs):
        """Helper method to copy selected carrier data to least cost columns"""
        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]

    def _calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'percentage_savings': 0,
                'loads_with_savings': 0,
                'total_savings_opportunity': 0
            }
            return

        # Basic stats - ensure numeric columns are properly converted
        total_loads = len(df)
        total_selected_cost = pd.to_numeric(
            df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        total_least_cost = pd.to_numeric(
            df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        total_potential_savings = pd.to_numeric(
            df['Potential Savings'], errors='coerce').fillna(0).sum()

        # Advanced stats - optimize by filtering once
        ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
        savings_df = df[ps_numeric > 0]
        loads_with_savings = len(savings_df)
        total_savings_opportunity = pd.to_numeric(
            savings_df['Potential Savings'], errors='coerce').fillna(0).sum()

        # Calculate percentages
        if total_selected_cost > 0:
            percentage_savings = (total_potential_savings / total_selected_cost) * 100
        else:
            percentage_savings = 0

        if total_loads > 0:
            average_savings_per_load = total_potential_savings / total_loads
        else:
            average_savings_per_load = 0

        self.summary_stats = {
            'total_loads': total_loads,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'percentage_savings': percentage_savings,
            'loads_with_savings': loads_with_savings,
            'total_savings_opportunity': total_savings_opportunity
        }

    def save_processed_data(self, output_file):
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Processed Data"

        # Add company and date info (no big title rows)
        row = 1
        if self.title_info:
            # Style company and date range with expanded width
            header_style_border = Border(
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                bottom=Side(style='medium', color='1F4E79')
            )

            if 'company_name' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                company_cell = ws_data[f'A{row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="FFFFFF")
                company_cell.fill = PatternFill(
                    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                )
                company_cell.border = header_style_border
                company_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(
                        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                    )
                ws_data.row_dimensions[row].height = 30
                row += 1

            if 'date_range' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                date_cell = ws_data[f'A{row}']
                date_cell.value = f"Date Range: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="FFFFFF")
                date_cell.fill = PatternFill(
                    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                )
                date_cell.border = header_style_border
                date_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(
                        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                    )
                ws_data.row_dimensions[row].height = 30
                row += 1

            # Add section headers row with color coding
            row = 4
            # Selected Carrier section (columns I-N, which are 9-14) - Light Blue
            selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
            # Reduced from 11 to 10
            selected_header.font = Font(size=10, bold=True, color="FFFFFF")
            selected_header.fill = PatternFill(
                start_color="4A90E2", end_color="4A90E2", fill_type="solid"
            )
            selected_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('I4:N4')
            for col in range(9, 15):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(
                    start_color="4A90E2", end_color="4A90E2", fill_type="solid"
                )

            # Least Cost Carrier section (columns O-T, which are 15-20) - Light Orange
            least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
            # Reduced from 11 to 10
            least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")
            least_cost_header.fill = PatternFill(
                start_color="FF8C42", end_color="FF8C42", fill_type="solid"
            )
            least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('O4:T4')
            for col in range(15, 21):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(
                    start_color="FF8C42", end_color="FF8C42", fill_type="solid"
                )

            row = 5  # Headers will be on row 5

        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Color code headers based on section
            if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                # Light Blue
                cell.fill = PatternFill(
                    start_color="87CEEB", end_color="87CEEB", fill_type="solid"
                )
            elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                # Light Orange
                cell.fill = PatternFill(
                    start_color="FFB366", end_color="FFB366", fill_type="solid"
                )
            elif header == 'Potential Savings':  # Potential Savings column - Green
                # Light Green
                cell.fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
            else:
                # Default blue
                cell.fill = PatternFill(
                    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                )

        # Add data with alternating row colors and comprehensive borders
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Filter out any remaining empty rows before writing to Excel
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']

        # Ensure all data is properly typed before processing
        for col in clean_data.columns:
            if col in ['Selected Transit Days', 'Selected Freight Cost',
                      'Selected Accessorial Cost', 'Selected Total Cost',
                      'Least Cost Transit Days', 'Least Cost Freight Cost',
                      'Least Cost Accessorial Cost', 'Least Cost Total Cost',
                      'Potential Savings']:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')

        for data_idx, data_row in enumerate(
            dataframe_to_rows(clean_data, index=False, header=False)
        ):
            # Skip rows that are mostly empty
            # Ensure we're comparing integers by converting the sum result
            try:
                non_empty_count = sum(
                    1 for val in data_row
                    if val is not None and str(val).strip() != '' and str(val) != 'nan'
                )
                if non_empty_count < 3:
                    continue
            except Exception as e:
                print(f"Error processing row {data_idx}: {e}")
                print(f"Row data: {data_row}")
                continue

            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"

            # First pass: collect all content lengths to determine optimal row height
            max_content_length = 0
            for col_idx, value in enumerate(data_row, 1):
                content_length = len(str(value)) if value else 0
                max_content_length = max(max_content_length, content_length)

            # Enhanced dynamic height calculation for long carrier names
            # Check if this row contains carrier information that might wrap
            has_carrier_data = any('TRANSPORT' in str(val).upper() or
                                 'LOGISTICS' in str(val).upper() or
                                 'FREIGHT' in str(val).upper() or
                                 len(str(val)) > 25 for val in data_row if val)

            if has_carrier_data and max_content_length > 25:
                # For carrier names, be more generous with height to prevent cutoff
                optimal_height = min(50, max(30, max_content_length * 1.2))
            elif max_content_length > 30:  # Very long content
                optimal_height = min(45, max(25, max_content_length * 1.0))
            elif max_content_length > 20:  # Long content
                optimal_height = min(35, max(22, max_content_length * 0.8))
            elif max_content_length > 15:  # Medium content
                optimal_height = 25
            else:
                optimal_height = 20  # Default height with a bit more room

            # Set the row height once for the entire row
            ws_data.row_dimensions[row].height = optimal_height

            # Second pass: apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                header_name = headers[col_idx-1] if col_idx-1 < len(headers) else ""
                # Center all cell contents and enable text wrapping for compactness
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border

                # Apply color coding to data cells based on section
                if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(
                        start_color=light_blue_bg,
                        end_color=light_blue_bg,
                        fill_type="solid"
                    )
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(
                        start_color=light_orange_bg,
                        end_color=light_orange_bg,
                        fill_type="solid"
                    )
                else:
                    cell.fill = PatternFill(
                        start_color=row_color,
                        end_color=row_color,
                        fill_type="solid"
                    )

                # Format currency columns and apply green color
                # for positive Potential Savings values
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost',
                                   'Selected Freight Cost', 'Least Cost Freight Cost',
                                   'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if (headers[col_idx-1] in currency_columns or
                        headers[col_idx-1] == 'Potential Savings'):
                    cell.number_format = '"$"#,##0.00'
                    # Apply light green background for positive Potential Savings values
                    if headers[col_idx-1] == 'Potential Savings':
                        try:
                            # Convert value to float for comparison, handle None and string values
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                # Handle different value types safely
                                if isinstance(value, (int, float)):
                                    numeric_value = float(value)
                                else:
                                    numeric_value = float(
                                        str(value).replace('$', '').replace(',', '')
                                    )
                                if numeric_value > 0:
                                    cell.fill = PatternFill(
                                        start_color="C6EFCE",
                                        end_color="C6EFCE",
                                        fill_type="solid"
                                    )
                        except (ValueError, TypeError, AttributeError):
                            pass  # Skip coloring if value can't be converted
                    cell.font = Font(size=10, bold=False, color="2C3E50")
            else:
                cell.font = Font(size=10, color="495057")

        # Enable auto-filter over header and data range (no freeze panes)
        try:
            header_row_idx = 5
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Add totals row with key financial metrics
        totals_row = row + 2

        # Add "TOTALS" label
        totals_label = ws_data.cell(row=totals_row, column=1, value="TOTALS")
        totals_label.font = Font(size=12, bold=True, color="FFFFFF")
        totals_label.fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid"
        )
        totals_label.alignment = Alignment(horizontal="center", vertical="center")
        totals_label.border = Border(
            left=Side(style='medium', color='2C3E50'),
            right=Side(style='medium', color='2C3E50'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )

        # Find the Selected Total Cost and Potential Savings columns
        selected_cost_col = None
        potential_savings_col = None
        for col_idx, header in enumerate(headers, 1):
            if 'Selected Total Cost' in str(header):
                selected_cost_col = col_idx
            elif 'Potential Savings' in str(header):
                potential_savings_col = col_idx

        # Add Total Selected Cost
        if selected_cost_col:
            cost_cell = ws_data.cell(row=totals_row, column=selected_cost_col,
                                   value=f"${self.summary_stats['total_selected_cost']:,.2f}")
            cost_cell.font = Font(size=12, bold=True, color="FFFFFF")
            cost_cell.fill = PatternFill(
                start_color="3498DB", end_color="3498DB", fill_type="solid"
            )
            cost_cell.alignment = Alignment(horizontal="center", vertical="center")
            cost_cell.number_format = '"$"#,##0.00'
            cost_cell.border = Border(
                left=Side(style='medium', color='3498DB'),
                right=Side(style='medium', color='3498DB'),
                top=Side(style='medium', color='3498DB'),
                bottom=Side(style='medium', color='3498DB')
            )

        # Add Total Potential Savings (most important number)
        if potential_savings_col:
            savings_cell = ws_data.cell(row=totals_row, column=potential_savings_col,
                value=f"${self.summary_stats['total_potential_savings']:,.2f}"
            )
            # Larger font for emphasis
            savings_cell.font = Font(size=14, bold=True, color="FFFFFF")
            savings_cell.fill = PatternFill(
                start_color="27AE60", end_color="27AE60", fill_type="solid"
            )
            savings_cell.alignment = Alignment(horizontal="center", vertical="center")
            savings_cell.number_format = '"$"#,##0.00'
            savings_cell.border = Border(
                left=Side(style='thick', color='27AE60'),  # Thicker border for emphasis
                right=Side(style='thick', color='27AE60'),
                top=Side(style='thick', color='27AE60'),
                bottom=Side(style='thick', color='27AE60')
            )

        # Set height for totals row
        ws_data.row_dimensions[totals_row].height = 25


        # Auto-fit column widths on the Processed Data sheet
        # (compact and consistent for all: table, CAL, PI)
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(5, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value)
                        length = len(text)
                        if length > max_length:
                            max_length = length
                # Aggressive compact auto-sizing: minimize width while ensuring content fits
                # Check if this column contains text content
                has_text_content = False
                for check_row in range(5, ws_data.max_row + 1):
                    check_cell = ws_data.cell(row=check_row, column=col_idx)
                    if check_cell.value and any(c.isalpha() for c in str(check_cell.value)):
                        has_text_content = True
                        break

                # Very tight padding for maximum compactness
                if has_text_content:
                    padding = 1.0  # Minimal padding for text
                else:
                    padding = 0.5   # Very tight for numbers

                # Calculate width with aggressive compacting
                adjusted_width = max_length + padding

                # Apply maximum width constraints for compactness
                if has_text_content:
                    max_width = 25  # Cap text columns at 25 characters
                else:
                    max_width = 15  # Cap number columns at 15 characters

                final_width = min(adjusted_width, max_width)
                # Minimum 6 for readability
                ws_data.column_dimensions[col_letter].width = max(6, final_width)

            # Optimize row heights for compact layout
            for rh in [1, 2, 4, 5]:
                if rh <= ws_data.max_row:
                    # Reduced from 22 to 20
                    ws_data.row_dimensions[rh].height = max(
                        ws_data.row_dimensions[rh].height or 0, 20
                    )

            # Ensure gridlines visible
            ws_data.sheet_view.showGridLines = True
        except Exception:
            pass

        # Add thick outside borders around the entire data table
        try:
            # Define thick border styles
            thick_side = Side(style='medium', color='2C3E50')

            header_row_idx = 5
            first_row = header_row_idx
            last_row = row
            first_col = 1
            last_col = len(headers)

            # Apply thick borders to all edge cells
            for r in range(first_row, last_row + 1):
                for c in range(first_col, last_col + 1):
                    cell = ws_data.cell(row=r, column=c)
                    current_border = cell.border or Border()

                    # Determine which sides need thick borders
                    left_side = thick_side if c == first_col else current_border.left
                    right_side = thick_side if c == last_col else current_border.right
                    top_side = thick_side if r == first_row else current_border.top
                    bottom_side = thick_side if r == last_row else current_border.bottom

                    # Apply the new border
                    cell.border = Border(
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side
                    )
        except Exception:
            pass

        wb.save(output_file)
        wb.close()

        # Results storage
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}

    def process_excel_file(self, file_path):
        """Complete processing pipeline using Basic_Processor logic"""
        print(f"Processing UTC Main report: {os.path.basename(file_path)}")

        # Step 1: Load data
        df = self.load_data(file_path)

        # Step 2: Apply business rules
        df = self.apply_business_rules(df)

        # Step 3: Calculate summary stats
        self.calculate_summary_stats(df)

        # Step 4: Sort by destination city (new requirement)
        if 'Destination City' in df.columns:
            df = df.sort_values('Destination City', na_position='last')
            print(f"Sorted by Destination City - {len(df)} rows")

        self.processed_data = df
        return df

    def load_data(self, file_path):
        """Load and clean Excel data using Basic_Processor logic"""
        # Read Excel file
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)

        # Extract title information
        self.title_info = self._extract_title_info(df_raw)

        # Get headers and data
        headers = df_raw.iloc[self.HEADER_ROW].fillna('').astype(str).tolist()
        data_df = df_raw.iloc[self.DATA_START_ROW:].copy()
        data_df.columns = headers

        # Map UTC Main specific column names to expected format
        data_df = self._map_utc_columns(data_df)

        # Handle PS -> Potential Savings column mapping
        if 'PS' in data_df.columns:
            data_df = data_df.rename(columns={'PS': 'Potential Savings'})

        # Clean the data
        data_df = self._clean_data(data_df)

        # Calculate Potential Savings if missing
        data_df = self._calculate_potential_savings(data_df)

        return data_df

    def apply_business_rules(self, df):
        """Apply TMS business rules using Basic_Processor logic"""
        df = df.copy()
        print("Applying UTC Main business rules...")

        # Rule 1: Same Carrier Rule
        df = self._apply_same_carrier_rule(df)

        # Rule 2: Empty Least Cost Data Rule
        df = self._apply_empty_data_rule(df)

        # Rule 3: Negative Savings Rule
        df = self._apply_negative_savings_rule(df)

        # Rule 4: TL Carriers Rule
        df = self._apply_tl_carriers_rule(df)

        # Rule 5: DALKO DEFENDER INSURANCE Rule
        df = self._apply_dalko_rule(df)

        # Rule 6: Evanston City Rule (UTC Main specific)
        df = self._apply_evanston_rule(df)

        return df

    def _apply_same_carrier_rule(self, df):
        """Rule 1: Set PS to 0 when selected carrier = least cost carrier"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                (df['Selected Carrier'].notna()) &
                (df['Least Cost Carrier'].notna()) &
                (df['Selected Carrier'].astype(str) != '') &
                (df['Least Cost Carrier'].astype(str) != '')
            )

            count = mask.sum()
            if count > 0 and 'Potential Savings' in df.columns:
                df.loc[mask, 'Potential Savings'] = 0
                print(f"Same Carrier Rule: {count} rows")

        return df

    def _apply_empty_data_rule(self, df):
        """Rule 2: Copy selected data when least cost data is missing"""
        if 'Least Cost Carrier' in df.columns:
            mask = (
                df['Least Cost Carrier'].isna() |
                (df['Least Cost Carrier'].astype(str) == '') |
                (df['Least Cost Carrier'].astype(str) == 'nan')
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"Empty Data Rule: {count} rows")

        return df

    def _apply_negative_savings_rule(self, df):
        """Rule 3: Fix negative potential savings"""
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            mask = ps_numeric < 0
            count = mask.sum()

            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                df.loc[mask, 'Potential Savings'] = 0
                print(f"Negative Savings Rule: {count} rows")

        return df

    def _apply_tl_carriers_rule(self, df):
        """Rule 4: Special handling for TL carriers"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                df['Selected Carrier'].astype(str).str.upper().isin(
                    [c.upper() for c in self.TL_CARRIERS]
                ) |
                df['Least Cost Carrier'].astype(str).str.upper().isin(
                    [c.upper() for c in self.TL_CARRIERS]
                )
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"TL Carriers Rule: {count} rows")

        return df

    def _apply_dalko_rule(self, df):
        """Rule 5: DALKO DEFENDER INSURANCE pattern matching"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            dalko_matches = []

            for idx, row in df.iterrows():
                selected = str(row['Selected Carrier']).strip().upper()
                least_cost = str(row['Least Cost Carrier']).strip().upper()

                # Skip empty or nan values
                if selected in ['', 'NAN', 'NONE'] or least_cost in ['', 'NAN', 'NONE']:
                    continue

                # Check for DALKO DEFENDER INSURANCE patterns
                # Pattern 1: "DALKO DEFENDER INSURANCE/XXX" where least cost is "XXX"
                if 'DALKO DEFENDER INSURANCE/' in selected:
                    carrier_after_slash = selected.split('DALKO DEFENDER INSURANCE/')[-1].strip()
                    if carrier_after_slash == least_cost:
                        dalko_matches.append(idx)

                # Pattern 2: "XXX/DALKO DEFENDER INSURANCE" where least cost is "XXX"
                elif '/DALKO DEFENDER INSURANCE' in selected:
                    carrier_before_slash = selected.split('/DALKO DEFENDER INSURANCE')[0].strip()
                    if carrier_before_slash == least_cost:
                        dalko_matches.append(idx)

            count = len(dalko_matches)
            if count > 0:
                dalko_mask = df.index.isin(dalko_matches)
                self._copy_selected_to_least_cost(df, dalko_mask)
                if 'Potential Savings' in df.columns:
                    df.loc[dalko_mask, 'Potential Savings'] = 0
                print(f"DALKO DEFENDER INSURANCE Rule: {count} rows")

        return df

    def _apply_evanston_rule(self, df):
        """Rule 6: Evanston City Rule - Zero out PS for any Evanston origin or destination"""
        evanston_count = 0

        # Check if city columns exist
        origin_col_found = 'Origin City' in df.columns
        dest_col_found = 'Destination City' in df.columns

        if not (origin_col_found or dest_col_found):
            print("Evanston Rule: No city columns found, skipping rule")
            return df

        # Create mask for Evanston cities
        evanston_mask = pd.Series([False] * len(df), index=df.index)

        # Check Origin City for "Evanston"
        if origin_col_found:
            origin_evanston = df['Origin City'].astype(str).str.upper().str.contains(
                'EVANSTON', na=False
            )
            evanston_mask |= origin_evanston
            origin_count = origin_evanston.sum()
            if origin_count > 0:
                print(f"Found {origin_count} rows with Evanston in Origin City")

        # Check Destination City for "Evanston"
        if dest_col_found:
            dest_evanston = df['Destination City'].astype(str).str.upper().str.contains(
                'EVANSTON', na=False
            )
            evanston_mask |= dest_evanston
            dest_count = dest_evanston.sum()
            if dest_count > 0:
                print(f"Found {dest_count} rows with Evanston in Destination City")

        evanston_count = evanston_mask.sum()

        if evanston_count > 0 and 'Potential Savings' in df.columns:
            # Zero out Potential Savings for Evanston rows
            df.loc[evanston_mask, 'Potential Savings'] = 0
            print(f"Evanston City Rule: Zeroed out PS for {evanston_count} rows")
        elif evanston_count > 0:
            print(
                f"Evanston City Rule: Found {evanston_count} Evanston rows "
                f"but no Potential Savings column"
            )
        else:
            print("Evanston City Rule: No Evanston cities found")

        return df

    def _copy_selected_to_least_cost(self, df, mask):
        """Helper: Copy selected carrier data to least cost columns"""
        column_pairs = [
            ('Selected Carrier', 'Least Cost Carrier'),
            ('Selected Service Type', 'Least Cost Service Type'),
            ('Selected Transit Days', 'Least Cost Transit Days'),
            ('Selected Freight Cost', 'Least Cost Freight Cost'),
            ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
            ('Selected Total Cost', 'Least Cost Total Cost')
        ]

        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]

    def _map_utc_columns(self, df):
        """Map UTC Main column names to expected format"""
        df = df.copy()

        # Find carrier columns - there should be two "Carrier" columns
        new_columns = list(df.columns)
        carrier_positions = []

        for i, col in enumerate(new_columns):
            if col == 'Carrier':
                carrier_positions.append(i)

        # Rename the two "Carrier" columns
        if len(carrier_positions) >= 2:
            new_columns[carrier_positions[0]] = 'Selected Carrier'
            new_columns[carrier_positions[1]] = 'Least Cost Carrier'

        # Map other column names to standardized format
        column_mapping = {
            'Service Type': 'Selected Service Type',  # First occurrence
            'Transit\nDays': 'Selected Transit Days',  # First occurrence
            'Freight + Fuel': 'Selected Freight Cost',  # First occurrence
            'Total Acc.': 'Selected Accessorial Cost',  # First occurrence
            'Total Cost ': 'Selected Total Cost'  # First occurrence (note the trailing space)
        }

        # Apply mapping to first occurrence of each column
        for old_name, new_name in column_mapping.items():
            if old_name in new_columns:
                # Find first occurrence and rename it
                idx = new_columns.index(old_name)
                new_columns[idx] = new_name

        # Now handle the second set (Least Cost columns)
        least_cost_mapping = {
            'Service Type': 'Least Cost Service Type',
            'Transit\nDays': 'Least Cost Transit Days',
            'Freight + Fuel': 'Least Cost Freight Cost',
            'Total Acc.': 'Least Cost Accessorial Cost',
            'Total Cost ': 'Least Cost Total Cost'
        }

        # Find and rename second occurrences
        for old_name, new_name in least_cost_mapping.items():
            # Count occurrences and rename the second one
            occurrences = [i for i, col in enumerate(new_columns) if col == old_name]
            if len(occurrences) >= 2:
                new_columns[occurrences[1]] = new_name

        df.columns = new_columns
        return df

    def _clean_data(self, df):
        """Clean and validate data types"""
        df = df.copy()

        # Remove rows with missing Load No.
        if 'Load No.' in df.columns:
            df = df.dropna(subset=['Load No.'])
            df = df[df['Load No.'].astype(str).str.strip() != '']
            df = df[df['Load No.'].astype(str) != 'nan']

        # Clean numeric columns
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost',
            'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost',
            'Least Cost Accessorial Cost', 'Least Cost Total Cost',
            'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Clean string columns
        string_columns = [
            'Selected Carrier', 'Selected Service Type',
            'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')

        # Remove mostly empty rows
        df = df.dropna(thresh=5).reset_index(drop=True)

        return df

    def _calculate_potential_savings(self, df):
        """Calculate Potential Savings if missing"""
        df = df.copy()

        # If PS already exists and has valid data, keep it
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce')
            if ps_numeric.notna().sum() > 0:
                return df

        # Calculate PS from cost difference
        if 'Selected Total Cost' in df.columns and 'Least Cost Total Cost' in df.columns:
            selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0)
            least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0)
            df['Potential Savings'] = selected_cost - least_cost

            # Handle cases where least cost is 0
            mask_zero_least = least_cost == 0
            df.loc[mask_zero_least, 'Potential Savings'] = 0
        else:
            df['Potential Savings'] = 0

        return df

    def _extract_title_info(self, df_raw):
        """Extract title and company information from Excel header"""
        title_info = {}

        try:
            # Company name (row 4, column B)
            if len(df_raw) > 3 and len(df_raw.columns) > 1:
                company = df_raw.iloc[3, 1]
                if pd.notna(company):
                    title_info['company_name'] = str(company)

            # Date range (row 6, column B)
            if len(df_raw) > 5 and len(df_raw.columns) > 1:
                date_range = df_raw.iloc[5, 1]
                if pd.notna(date_range):
                    title_info['date_range'] = str(date_range)
        except Exception:
            pass

        return title_info

    def calculate_summary_stats(self, df):
        """Calculate comprehensive summary statistics"""
        if df.empty:
            self.summary_stats = {
                'total_loads': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'loads_with_savings': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'percentage_savings': 0
            }
            return

        # Calculate all stats
        total_loads = len(df)
        if 'Selected Total Cost' in df.columns:
            total_selected_cost = pd.to_numeric(
                df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        else:
            total_selected_cost = 0

        if 'Least Cost Total Cost' in df.columns:
            total_least_cost = pd.to_numeric(
                df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        else:
            total_least_cost = 0

        if 'Potential Savings' in df.columns:
            ps_series = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            total_potential_savings = ps_series.sum()
            ps_numeric = ps_series
        else:
            total_potential_savings = 0
            ps_numeric = pd.Series([0] * len(df))
        loads_with_savings = (ps_numeric > 0).sum()

        percentage_savings = (
            (total_potential_savings / total_selected_cost * 100)
            if total_selected_cost > 0 else 0
        )
        average_savings_per_load = total_potential_savings / total_loads if total_loads > 0 else 0

        self.summary_stats = {
            'total_loads': total_loads,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'loads_with_savings': loads_with_savings,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'percentage_savings': percentage_savings
        }


class ModernTMSProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("TMS Data Processor Pro")

        # Auto-size window to content
        self.root.configure(bg='#f8f9fa')
        self.root.resizable(True, True)
        # Let content determine size initially

        # Initialize processors
        self.basic_processor = ModernTMSProcessor()
        self.detailed_processor = None

        # Import and initialize city processors
        try:
            from city_processors import UTCMainProcessor, UTCFSProcessor, TranscoProcessor
            self.utc_main_processor = UTCMainProcessor()
            self.utc_fs_processor = UTCFSProcessor()
            self.transco_processor = TranscoProcessor()
            print("City processors loaded successfully")
        except ImportError as e:
            print(f"Warning: Could not import city processors: {e}")
            self.utc_main_processor = None
            self.utc_fs_processor = None
            self.transco_processor = None
        self.input_files = []  # Changed to list for multiple files
        self.output_file = None

        # Progress tracking
        self.is_processing = False

        # Savings history tracking
        self.savings_history_file = Path.home() / "Desktop" / "tms_savings_history.json"
        self.savings_history = self.load_savings_history()

        # Configure style
        self.setup_styles()

        # Create GUI
        self.create_widgets()

    def setup_styles(self):
        """Setup modern styling for the application"""
        style = ttk.Style()
        style.theme_use('clam')

        # Configure colors with modern palette using constants
        style.configure('Title.TLabel',
                       font=('Segoe UI', 20, 'bold'),
                       foreground='#1a365d',
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Subtitle.TLabel',
                       font=('Segoe UI', 10),
                       foreground=UI_COLORS['TEXT_SECONDARY'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Header.TLabel',
                       font=('Segoe UI', 11, 'bold'),
                       foreground=UI_COLORS['TEXT_PRIMARY'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Info.TLabel',
                       font=('Segoe UI', 10),
                       foreground=UI_COLORS['TEXT_MUTED'],
                       background=UI_COLORS['BACKGROUND_GRAY'])
        style.configure('Success.TLabel',
                       font=('Segoe UI', 10, 'bold'),
                       foreground=UI_COLORS['SUCCESS_GREEN'],
                       background=UI_COLORS['BACKGROUND_GRAY'])

        # Configure modern buttons with hover effects using constants
        style.configure('Primary.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background=UI_COLORS['PRIMARY_BLUE'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(20, 10))
        style.map('Primary.TButton',
                 background=[
                     ('active', UI_COLORS['PRIMARY_BLUE_HOVER']),
                     ('pressed', UI_COLORS['PRIMARY_BLUE_PRESSED'])
                 ])

        style.configure('Success.TButton',
                       font=('Segoe UI', 12, 'bold'),
                       background=UI_COLORS['SUCCESS_LIGHT'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(25, 12))
        style.map('Success.TButton',
                 background=[('active', UI_COLORS['SUCCESS_GREEN']), ('pressed', '#2f855a')])

        # Enhanced Process Button Style
        style.configure('ProcessButton.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background='#28a745',
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(35, 15))
        style.map('ProcessButton.TButton',
                 background=[('active', '#218838'),
                           ('pressed', '#1e7e34'),
                           ('disabled', '#6c757d')])

        # Process Button Disabled Style
        style.configure('ProcessButtonDisabled.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background='#6c757d',
                       foreground='#ffffff',
                       borderwidth=0,
                       focuscolor='none',
                       relief='flat',
                       padding=(35, 15))

        style.configure('Browse.TButton',
                       font=('Segoe UI', 10),
                       background=UI_COLORS['ACCENT_PURPLE'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(15, 8))
        style.map('Browse.TButton',
                 background=[('active', '#5a67d8'), ('pressed', '#4c51bf')])

        # Configure modern report type buttons using constants
        style.configure('ReportCard.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background=UI_COLORS['BACKGROUND_WHITE'],
                       foreground=UI_COLORS['TEXT_SECONDARY'],
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 20))
        style.map('ReportCard.TButton',
                 background=[('active', UI_COLORS['BACKGROUND_LIGHT']), ('pressed', '#edf2f7')])

        style.configure('ReportCardActive.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background=UI_COLORS['PRIMARY_BLUE'],
                       foreground='white',
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(25, 15))
        style.map('ReportCardActive.TButton',
                 background=[
                     ('active', UI_COLORS['PRIMARY_BLUE_HOVER']),
                     ('pressed', UI_COLORS['PRIMARY_BLUE_PRESSED'])
                 ])

        style.configure('ReportCardDisabled.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       background=UI_COLORS['BACKGROUND_WHITE'],
                       foreground=UI_COLORS['TEXT_DISABLED'],
                       borderwidth=1,
                       relief='solid',
                       focuscolor='none',
                       padding=(25, 15))
        style.map('ReportCardDisabled.TButton',
                 background=[('active', '#f1f3f4'), ('pressed', '#e8eaed')])

        # Configure frames with modern styling
        style.configure('Card.TFrame',
                       background='#ffffff',
                       relief='solid',
                       borderwidth=1,
                       lightcolor='#e2e8f0',
                       darkcolor='#e2e8f0')

        # Add modern input styling
        style.configure('Modern.TEntry',
                       fieldbackground='#ffffff',
                       borderwidth=2,
                       relief='solid',
                       lightcolor='#e2e8f0',
                       darkcolor='#e2e8f0',
                       focuscolor=UI_COLORS['PRIMARY_BLUE'],
                       font=('Segoe UI', 10))


    def create_card_buttons(self, parent):
        """Create modern card-style buttons"""
        # Create button frame with compact spacing and minimum width
        button_frame = tk.Frame(parent, bg='#f8f9fa', width=800)  # Force minimum width
        button_frame.grid(row=0, column=0, pady=1, sticky="ew")
        button_frame.grid_propagate(False)  # Don't shrink frame
        # Ensure button frame expands
        parent.grid_columnconfigure(0, weight=1)

        # Basic Report Button
        self.basic_button = ttk.Button(button_frame,
                                     text="📊\nBasic Report",
                                     style='ReportCardActive.TButton',
                                     command=lambda: self.select_card('basic'))
        self.basic_button.grid(row=0, column=0, padx=2, pady=2, sticky="nsew")

        # Detailed Report Button
        self.detailed_button = ttk.Button(button_frame,
                                        text="📈\nDetailed Report",
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('detailed'))
        self.detailed_button.grid(row=0, column=1, padx=2, pady=2, sticky="nsew")

        # UTC Main Button
        self.utc_main_button = ttk.Button(button_frame,
                                        text="🏢\nUTC Main",
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('utc_main'))
        self.utc_main_button.grid(row=0, column=2, padx=2, pady=2, sticky="nsew")

        # UTC FS Button
        self.utc_fs_button = ttk.Button(button_frame,
                                        text="🏭\nUTC FS",
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('utc_fs'))
        self.utc_fs_button.grid(row=0, column=3, padx=2, pady=2, sticky="nsew")
        print("UTC FS button created and placed in column 3")

        # Transco Button
        self.transco_button = ttk.Button(button_frame,
                                        text="🚛\nTransco",
                                        style='ReportCardDisabled.TButton',
                                        command=lambda: self.select_card('transco'))
        self.transco_button.grid(row=0, column=4, padx=2, pady=2, sticky="nsew")
        print("Transco button created and placed in column 4")

        # Make columns equal width (5 buttons total)
        button_frame.grid_columnconfigure(0, weight=1, uniform="card")
        button_frame.grid_columnconfigure(1, weight=1, uniform="card")
        button_frame.grid_columnconfigure(2, weight=1, uniform="card")
        button_frame.grid_columnconfigure(3, weight=1, uniform="card")
        button_frame.grid_columnconfigure(4, weight=1, uniform="card")

        # Force layout update
        button_frame.update_idletasks()

        # Debug: Print button info
        print("="*50)
        print("BUTTON CREATION DEBUG:")
        print(f"Button frame width: {button_frame.winfo_reqwidth()}")
        print(f"Total buttons created: 5 (Home, Basic, Detailed, UTC Main, UTC FS)")
        print("Buttons in grid:")
        print("  Column 0: Home")
        print("  Column 1: Basic")
        print("  Column 2: Detailed")
        print("  Column 3: UTC Main")
        print("  Column 4: UTC FS")
        print("="*50)

        # Store references for updating styles
        self.cards = {
            'basic': {'button': self.basic_button},
            'detailed': {'button': self.detailed_button},
            'utc_main': {'button': self.utc_main_button},
            'utc_fs': {'button': self.utc_fs_button},
            'transco': {'button': self.transco_button}
        }

        # Don't set initial selection here - will be set after sections are created

    def select_card(self, card_type):
        """Handle card selection with visual feedback"""
        self.report_type.set(card_type)

        # Update navigation button states
        self.update_nav_button_states(card_type)

        # Update UI based on selection
        self.update_ui_for_selection(card_type)

        # Update browse button text for single/multiple file selection
        self.update_browse_button_text()

        # Update file display placeholder text
        self.update_file_display_placeholder()

    def update_nav_button_states(self, active_card):
        """Update navigation button visual states"""
        if not hasattr(self, 'nav_buttons'):
            return

        # Reset all buttons to inactive state
        for card_name, button in self.nav_buttons.items():
            if card_name == 'basic':
                if card_name == active_card:
                    # Active basic button
                    button.configure(bg='#4299e1', fg='white', font=('Segoe UI', 11, 'bold'))
                else:
                    # Inactive basic button
                    button.configure(bg='#f7fafc', fg='#4a5568', font=('Segoe UI', 11))
            else:
                if card_name == active_card:
                    # Active secondary button
                    button.configure(bg='#4299e1', fg='white', font=('Segoe UI', 10, 'bold'))
                else:
                    # Inactive secondary button
                    button.configure(bg='#ffffff', fg='#4a5568', font=('Segoe UI', 10))

    def get_report_description(self):
        """Get contextual description text for the current report type"""
        report_type = self.report_type.get()
        if report_type == 'basic':
            return "For M&T Reports"
        elif report_type == 'detailed':
            return "For Cast Nylons Reports"
        elif report_type == 'utc_main':
            return "For UTC Main Reports with Basic Logic"
        elif report_type == 'utc_fs':
            return "For UTC FS Reports with Green River Rule"
        elif report_type == 'transco':
            return "For Transco Reports with Miles City Rule"
        else:
            return ""

    def update_ui_for_selection(self, card_type):
        """Update UI elements based on selected card type"""
        if card_type == 'utc_main':
            # Show normal file input UI for UTC Main
            self.show_file_input_ui()
            # Show stats display for processing pages
            if hasattr(self, 'stats_outer_frame'):
                self.stats_outer_frame.grid(
                    row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S),
                    padx=(8, 5), pady=5
                )
        elif card_type == 'utc_fs':
            # Show normal file input UI for UTC FS
            self.show_file_input_ui()
            # Show stats display for processing pages
            if hasattr(self, 'stats_outer_frame'):
                self.stats_outer_frame.grid(
                    row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S),
                    padx=(8, 5), pady=5
                )
        elif card_type == 'transco':
            # Show normal file input UI for Transco
            self.show_file_input_ui()
            # Show stats display for processing pages
            if hasattr(self, 'stats_outer_frame'):
                self.stats_outer_frame.grid(
                    row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S),
                    padx=(8, 5), pady=5
                )
        else:
            # Show normal file input UI
            self.show_file_input_ui()
            # Show stats display for processing pages
            if hasattr(self, 'stats_outer_frame'):
                self.stats_outer_frame.grid(
                    row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S),
                    padx=(8, 5), pady=5
                )


    def create_file_input_section(self):
        """Create the file input UI section"""
        self.file_section = tk.Frame(self.input_section, bg='#f8f9fa')
        # Don't grid it initially - let the initial state logic handle visibility
        self.file_section.columnconfigure(0, weight=1)

        # Section header with context-specific description
        self.header_container = tk.Frame(self.file_section, bg='#f8f9fa')
        self.header_container.grid(row=0, column=0, pady=(2, 1))

        self.header_label = ttk.Label(
            self.header_container, text="📁 Input File",
            style='Header.TLabel', background='#f8f9fa'
        )
        self.header_label.pack()

        # Add descriptive text based on report type
        description_text = self.get_report_description()
        if description_text:
            self.desc_label = tk.Label(self.header_container, text=description_text,
                                 font=('Segoe UI', 9, 'italic'),
                                 fg='#4a5568', bg='#f8f9fa')
            self.desc_label.pack(pady=(1, 0))

        file_frame = tk.Frame(self.file_section, bg='#f8f9fa')
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=3)
        file_frame.columnconfigure(0, weight=1)

        # File display with clean styling and drag-drop support - expandable width
        file_display_frame = tk.Frame(file_frame, bg='#ffffff', relief='flat', bd=0)
        file_display_frame.grid(row=0, column=0, padx=(0, 3), sticky=(tk.W, tk.E))

        # Create custom scrollable text widget with cooler scroller
        scroll_container = tk.Frame(file_display_frame, bg='#ffffff')
        scroll_container.pack(fill='both', expand=True, padx=2, pady=1)

        # Text widget
        self.file_display = tk.Text(scroll_container,
                                   height=3,
                                   width=30,
                                   font=('Segoe UI', 10),
                                   fg='#000000',
                                   bg='#ffffff',
                                   wrap=tk.WORD,
                                   state='disabled',
                                   borderwidth=0,
                                   highlightthickness=0,
                                   yscrollcommand=lambda *args: self._on_scroll(*args))

        # Modern custom scrollbar with sleek design
        self.cool_scrollbar = tk.Canvas(scroll_container, width=12, bg='#f1f3f4',
                                       highlightthickness=0, bd=0)

        # Configure grid layout
        scroll_container.grid_columnconfigure(0, weight=1)
        scroll_container.grid_rowconfigure(0, weight=1)

        self.file_display.grid(row=0, column=0, sticky='nsew')
        self.cool_scrollbar.grid(row=0, column=1, sticky='ns')

        # Initialize scrollbar
        self._setup_cool_scrollbar()

        # Initialize with placeholder text
        self.file_display.config(state='normal')
        self.file_display.insert('1.0', "No files selected")
        self.file_display.config(state='disabled', fg='#6c757d')

        # Update scrollbar
        self._update_scrollbar()

        file_display_frame.grid_columnconfigure(0, weight=1)

        # Store reference to file display frame for updates
        self.file_display_frame = file_display_frame

        # Enable drag and drop
        self.setup_drag_drop(file_display_frame)

        # Browse Button Frame (separate from file display)
        browse_frame = tk.Frame(self.file_section, bg='#f8f9fa')
        browse_frame.grid(row=2, column=0, pady=(2, 1))

        self.browse_button_main = ttk.Button(browse_frame, text="📂 Browse Files",
                                       command=self.browse_file, style='Browse.TButton')
        self.browse_button_main.pack()

        # Process Button added to input section (below browse)
        process_frame = tk.Frame(self.file_section, bg='#f8f9fa')
        process_frame.grid(row=3, column=0, pady=(1, 1))

        self.process_button = ttk.Button(process_frame, text="🚀 PROCESS",
            command=self.process_file,
            style='ProcessButton.TButton',
            state="disabled"
        )
        self.process_button.pack()


    def show_file_input_ui(self):
        """Show file input UI and hide other sections"""
        if hasattr(self, 'file_section'):
            self.file_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=1)
            # Refresh the description text for the current mode
            self.refresh_file_section_description()
        # Show navigation bar for file input (process button is now integrated)
        if hasattr(self, 'nav_bar'):
            self.nav_bar.grid()
        self.update_process_button_state()

    def show_marmon_special_ui(self):
        """Show Marmon special clients UI and hide other sections"""
        if hasattr(self, 'marmon_section'):
            self.marmon_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=1)
            # Refresh the description text for the current mode
            self.refresh_marmon_section_description()
        else:
            # Create the marmon section if it doesn't exist
            self.create_marmon_special_section()
            self.marmon_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=1)

        if hasattr(self, 'file_section'):
            self.file_section.grid_remove()
        # Show navigation bar
        if hasattr(self, 'nav_bar'):
            self.nav_bar.grid()
        self.update_process_button_state()

    def refresh_marmon_section_description(self):
        """Refresh the description text in the marmon section header"""
        if not hasattr(self, 'marmon_section') or not hasattr(self, 'marmon_header_container'):
            return

        # Remove existing description label if it exists
        if hasattr(self, 'marmon_desc_label'):
            self.marmon_desc_label.destroy()

        # Add new description label
        description_text = self.get_report_description()
        if description_text:
            self.marmon_desc_label = tk.Label(self.marmon_header_container, text=description_text,
                                             font=('Segoe UI', 9, 'italic'),
                                             fg='#4a5568', bg='#f8f9fa')
            self.marmon_desc_label.pack(pady=(1, 0))

    def refresh_file_section_description(self):
        """Refresh the description text in the file section header"""
        if not hasattr(self, 'file_section') or not hasattr(self, 'header_container'):
            return

        # Remove existing description label if it exists
        if hasattr(self, 'desc_label'):
            self.desc_label.destroy()

        # Add new description text based on current report type
        description_text = self.get_report_description()
        if description_text:
            self.desc_label = tk.Label(self.header_container, text=description_text,
                                     font=('Segoe UI', 9, 'italic'),
                                     fg='#4a5568', bg='#f8f9fa')
            self.desc_label.pack(pady=(1, 0))

    def update_browse_button_text(self):
        """Update browse button text based on report type"""
        if hasattr(self, 'browse_button_main'):
            if self.report_type.get() in ["utc_main", "utc_fs", "transco"]:
                self.browse_button_main.config(text="📂 Browse File")
            else:
                self.browse_button_main.config(text="📂 Browse Files")

    def update_file_display_placeholder(self):
        """Update file display placeholder text based on report type"""
        if hasattr(self, 'file_display') and not self.input_files:
            self.file_display.config(state='normal')
            self.file_display.delete('1.0', tk.END)
            if self.report_type.get() == "utc_main":
                self.file_display.insert('1.0', "No UTC Main file selected")
            elif self.report_type.get() == "utc_fs":
                self.file_display.insert('1.0', "No UTC FS file selected")
            elif self.report_type.get() == "transco":
                self.file_display.insert('1.0', "No Transco file selected")
            else:
                self.file_display.insert('1.0', "No files selected")
            self.file_display.config(state='disabled', fg='#6c757d')


    def create_marmon_special_section(self):
        """Create the Marmon special clients UI section"""
        self.marmon_section = tk.Frame(self.input_section, bg='#f8f9fa')
        # Don't grid it initially - let the initial state logic handle visibility
        self.marmon_section.columnconfigure(0, weight=1)

        # Section header with context-specific description
        self.marmon_header_container = tk.Frame(self.marmon_section, bg='#f8f9fa')
        self.marmon_header_container.grid(row=0, column=0, pady=(2, 1))

        ttk.Label(
            self.marmon_header_container, text="🏭 Marmon Special Clients",
            style='Header.TLabel', background='#f8f9fa'
        ).pack()

        # Add descriptive text
        description_text = self.get_report_description()
        if description_text:
            self.marmon_desc_label = tk.Label(self.marmon_header_container, text=description_text,
                                             font=('Segoe UI', 9, 'italic'),
                                             fg='#4a5568', bg='#f8f9fa')
            self.marmon_desc_label.pack(pady=(1, 0))

        # Client selection frame
        client_frame = tk.Frame(self.marmon_section, bg='#f8f9fa')
        client_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=3)
        client_frame.columnconfigure(0, weight=1)

        # Client selection label
        client_label_frame = tk.Frame(client_frame, bg='#f8f9fa')
        client_label_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(5, 3))

        tk.Label(client_label_frame, text="Select Client:", font=('Segoe UI', 11, 'bold'),
                fg='#2d3748', bg='#f8f9fa').pack(side='left')

        # Client selection dropdown
        self.marmon_client_var = tk.StringVar()
        client_options = ["UTC MAIN", "UTC FS", "TRANSCO"]
        self.marmon_client_dropdown = ttk.Combobox(
            client_frame, textvariable=self.marmon_client_var,
            values=client_options, state="readonly", width=15
        )
        self.marmon_client_dropdown.grid(row=1, column=0, sticky=(tk.W), pady=(0, 5))
        self.marmon_client_dropdown.bind('<<ComboboxSelected>>', self.on_marmon_client_select)

        # File display area
        file_display_frame = tk.Frame(client_frame, bg='#ffffff', relief='flat', bd=1)
        file_display_frame.grid(row=2, column=0, padx=(0, 3), sticky=(tk.W, tk.E), pady=(5, 0))
        file_display_frame.columnconfigure(0, weight=1)

        # Create custom scrollable text widget with cool scroller for Marmon
        scroll_container = tk.Frame(file_display_frame, bg='#ffffff')
        scroll_container.pack(fill='both', expand=True, padx=2, pady=1)

        # Text widget for Marmon files
        self.marmon_file_display = tk.Text(scroll_container,
                                          height=3,
                                          width=30,
                                          font=('Segoe UI', 10),
                                          fg='#000000',
                                          bg='#ffffff',
                                          wrap=tk.WORD,
                                          state='disabled',
                                          borderwidth=0,
                                          highlightthickness=0,
            yscrollcommand=lambda *args: self._on_marmon_scroll(*args)
        )

        # Modern custom scrollbar for Marmon section
        self.marmon_cool_scrollbar = tk.Canvas(scroll_container, width=12, bg='#f1f3f4',
                                              highlightthickness=0, bd=0)

        # Configure grid layout
        scroll_container.grid_columnconfigure(0, weight=1)
        scroll_container.grid_rowconfigure(0, weight=1)

        self.marmon_file_display.grid(row=0, column=0, sticky='nsew')
        self.marmon_cool_scrollbar.grid(row=0, column=1, sticky='ns')

        # Initialize scrollbar for Marmon
        self._setup_marmon_cool_scrollbar()

        # Initialize with placeholder text
        self.marmon_file_display.config(state='normal')
        self.marmon_file_display.insert('1.0', "Select a client and upload raw data file")
        self.marmon_file_display.config(state='disabled', fg='#6c757d')

        # Update scrollbar
        self._update_marmon_scrollbar()

        # Store reference to file display frame for updates
        self.marmon_file_display_frame = file_display_frame

        # Enable drag and drop for Marmon section
        self.setup_drag_drop(file_display_frame)

        # Browse Button Frame for Marmon
        marmon_browse_frame = tk.Frame(self.marmon_section, bg='#f8f9fa')
        marmon_browse_frame.grid(row=2, column=0, pady=(2, 1))

        marmon_browse_button = ttk.Button(marmon_browse_frame, text="📂 Browse",
                                         command=self.browse_marmon_file, style='Browse.TButton')
        marmon_browse_button.pack()

        # Process Button Frame for Marmon
        marmon_process_frame = tk.Frame(self.marmon_section, bg='#f8f9fa')
        marmon_process_frame.grid(row=3, column=0, pady=(1, 1))

        self.marmon_process_button = ttk.Button(marmon_process_frame, text="🏭 PROCESS",
            command=self.process_marmon_file,
            style='ProcessButton.TButton',
            state="disabled"
        )
        self.marmon_process_button.pack()


    def create_widgets(self):
        """Create the main GUI widgets"""
        # Main container with compact styling
        main_frame = ttk.Frame(self.root, padding="15", style='Card.TFrame')
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        # Header section
        header_frame = tk.Frame(main_frame, bg='#ffffff', relief='flat')
        header_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)

        # Title with modern styling
        title_label = ttk.Label(header_frame, text="🚛 TMS Data Processor", style='Title.TLabel')
        title_label.grid(row=0, column=0, pady=(5, 10))

        # Initialize report type variable
        self.report_type = tk.StringVar(value="basic")

        # Modern Navigation Bar with enhanced styling
        nav_container = tk.Frame(main_frame, bg='#ffffff', relief='flat', bd=0)
        nav_container.grid(
            row=1, column=0, columnspan=3, sticky=(tk.W, tk.E),
            pady=(0, 2), padx=1
        )

        self.nav_bar = tk.Frame(nav_container, bg='#f8fafc')
        self.nav_bar.pack(fill='both', expand=True, padx=0, pady=0)
        self.nav_bar.grid_columnconfigure(0, weight=1)
        self.create_navigation_bar()

        # Create modern horizontal layout container for input and recent uploads
        self.content_container = tk.Frame(main_frame, bg='#ffffff', relief='flat', bd=0)
        self.content_container.grid(
            row=2, column=0, columnspan=3,
            sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 8), padx=3
        )

        # Inner container with subtle background
        self.content_inner = tk.Frame(self.content_container, bg='#f8fafc')
        self.content_inner.pack(fill='both', expand=True, padx=1, pady=1)
        self.content_inner.grid_columnconfigure(0, weight=1)  # Input section (left) - 25%
        self.content_inner.grid_columnconfigure(1, weight=3)  # Recent uploads (right) - 75%

        # Input Section - Dynamic (File or Date input based on selection) - LEFT SIDE
        self.input_section = tk.Frame(self.content_inner, bg='#f8fafc')
        self.input_section.grid(
            row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S),
            pady=5, padx=(5, 8)
        )
        self.input_section.columnconfigure(0, weight=1)

        # Create both file input and date input sections
        self.create_file_input_section()

        # Set correct initial state (start with file input visible, date input hidden)
        if hasattr(self, 'file_section'):
            self.file_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 2), padx=1)

        # Process button is now integrated into the file input section

        # Stats Display Frame (right side of content) - Modern styling
        self.stats_outer_frame = tk.Frame(
            self.content_inner,
            bg=UI_COLORS['BACKGROUND_WHITE'],
            relief='flat', bd=0
        )
        self.stats_outer_frame.grid(
            row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S),
            padx=(8, 5), pady=5
        )

        # Create scrollable canvas for stats
        self.stats_canvas = tk.Canvas(
            self.stats_outer_frame, bg=UI_COLORS['BACKGROUND_WHITE'],
            highlightthickness=0
        )
        stats_scrollbar = ttk.Scrollbar(
            self.stats_outer_frame, orient="vertical",
            command=self.stats_canvas.yview
        )
        self.stats_display_frame = tk.Frame(self.stats_canvas, bg=UI_COLORS['BACKGROUND_WHITE'])

        # Configure canvas scrolling
        self.stats_display_frame.bind(
            "<Configure>",
            lambda e: self.stats_canvas.configure(
                scrollregion=self.stats_canvas.bbox("all")
            )
        )
        self.stats_canvas.create_window((0, 0), window=self.stats_display_frame, anchor="nw")
        self.stats_canvas.configure(yscrollcommand=stats_scrollbar.set)

        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            self.stats_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        def _bind_mousewheel(event):
            self.stats_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(event):
            self.stats_canvas.unbind_all("<MouseWheel>")

        # Bind mouse wheel events when mouse enters/leaves the canvas
        self.stats_canvas.bind('<Enter>', _bind_mousewheel)
        self.stats_canvas.bind('<Leave>', _unbind_mousewheel)

        # Pack scrollable components
        self.stats_canvas.pack(side="left", fill="both", expand=True)
        stats_scrollbar.pack(side="right", fill="y")

        # Initialize stats display
        self.update_savings_display()

        # Create main landing page
        # Set initial selection to basic report page and auto-resize
        self.root.after(1, lambda: self.select_card('basic'))
        self.root.after(100, self.auto_resize_window)

    def create_navigation_bar(self):
        """Create compact navigation bar for switching between pages"""
        nav_container = tk.Frame(self.nav_bar, bg='#e2e8f0')
        nav_container.pack(pady=2, padx=3)

        # Store button references for active state management
        self.nav_buttons = {}

        # Primary: Basic Report (larger, more prominent)
        self.nav_buttons['basic'] = tk.Button(
            nav_container, text="📊 Basic", font=('Segoe UI', 10, 'bold'),
                            bg='#4299e1', fg='white', relief='flat', bd=0,
                            cursor='hand2', command=lambda: self.select_card('basic'),
                            activebackground='#3182ce', padx=18, pady=10,
                            highlightthickness=0, width=12)
        self.nav_buttons['basic'].pack(side='left', padx=(0, 5))

        # Secondary: Other options (improved styling)
        self.nav_buttons['detailed'] = tk.Button(
            nav_container, text="📈 Detailed", font=('Segoe UI', 10),
                               bg='#e2e8f0', fg='#4a5568', relief='flat', bd=0,
                               cursor='hand2', command=lambda: self.select_card('detailed'),
                               activebackground='#cbd5e0', padx=18, pady=10,
                               highlightthickness=0, width=12)
        self.nav_buttons['detailed'].pack(side='left', padx=(0, 5))


        self.nav_buttons['utc_main'] = tk.Button(
            nav_container, text="🏢 UTC Main", font=('Segoe UI', 10),
                               bg='#e2e8f0', fg='#4a5568', relief='flat', bd=0,
                               cursor='hand2', command=lambda: self.select_card('utc_main'),
                               activebackground='#cbd5e0', padx=18, pady=10,
                               highlightthickness=0, width=12)
        self.nav_buttons['utc_main'].pack(side='left', padx=(0, 5))

        self.nav_buttons['utc_fs'] = tk.Button(
            nav_container, text="🏭 UTC FS", font=('Segoe UI', 10),
                               bg='#e2e8f0', fg='#4a5568', relief='flat', bd=0,
                               cursor='hand2', command=lambda: self.select_card('utc_fs'),
                               activebackground='#cbd5e0', padx=18, pady=10,
                               highlightthickness=0, width=12)
        self.nav_buttons['utc_fs'].pack(side='left', padx=(0, 5))

        self.nav_buttons['transco'] = tk.Button(
            nav_container, text="🚛 Transco", font=('Segoe UI', 10),
                               bg='#e2e8f0', fg='#4a5568', relief='flat', bd=0,
                               cursor='hand2', command=lambda: self.select_card('transco'),
                               activebackground='#cbd5e0', padx=18, pady=10,
                               highlightthickness=0, width=12)
        self.nav_buttons['transco'].pack(side='left')

    def setup_drag_drop(self, widget):
        """Setup drag and drop functionality for file selection"""
        def on_drag_enter(event):
            widget.configure(bg='#e6fffa')
            return "copy"

        def on_drag_leave(event):
            widget.configure(bg='#ffffff')

        def on_drop(event):
            widget.configure(bg='#ffffff')
            if hasattr(event, 'data'):
                files = event.data.split()
                if files:
                    # Filter for valid Excel files
                    excel_files = [
                        f.strip('{}') for f in files
                        if f.strip('{}').lower().endswith(('.xlsx', '.xls'))
                    ]

                    if excel_files:
                        self.input_files = excel_files
                        self.update_file_display()
                        self.update_process_button_state()
                    else:
                        messagebox.showwarning(
                            "Invalid Files",
                            "Please select Excel files (.xlsx or .xls)"
                        )

        try:
            # Try to set up tkinter DND if available
            widget.drop_target_register('DND_Files')
            widget.dnd_bind('<<DropEnter>>', on_drag_enter)
            widget.dnd_bind('<<DropLeave>>', on_drag_leave)
            widget.dnd_bind('<<Drop>>', on_drop)
        except Exception:
            # DND not available, continue without it
            pass

    def auto_resize_window(self):
        """Dynamically resize window to fit content with minimal padding"""
        self.root.update_idletasks()

        # Get the required size of all content with minimal padding
        required_width = self.root.winfo_reqwidth() + 10  # Minimal padding
        required_height = self.root.winfo_reqheight() + 80  # Minimal padding

        # Set more aggressive limits for minimalism - need space for 5 buttons
        min_width = 750
        max_width = int(self.root.winfo_screenwidth() * 0.8)
        min_height = 550
        max_height = int(self.root.winfo_screenheight() * 0.9)

        # Constrain to limits
        width = max(min_width, min(required_width, max_width))
        height = max(min_height, min(required_height, max_height))

        # Center on screen
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)

        # Apply new size
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        self.root.minsize(min_width, min_height)

    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def _setup_cool_scrollbar(self):
        """Setup the cool custom scrollbar"""
        self.scrollbar_thumb = None
        self.scrollbar_track_height = 0
        self.scrollbar_thumb_height = 0
        self.scrollbar_thumb_pos = 0

        # Bind scrollbar events
        self.cool_scrollbar.bind('<Button-1>', self._on_scrollbar_click)
        self.cool_scrollbar.bind('<B1-Motion>', self._on_scrollbar_drag)
        self.cool_scrollbar.bind('<Configure>', self._on_scrollbar_configure)

        # Bind text widget scroll events
        self.file_display.bind('<MouseWheel>', self._on_mousewheel)
        self.file_display.bind('<Configure>', self._update_scrollbar)

    def _on_scroll(self, *args):
        """Handle text widget scroll events"""
        self._update_scrollbar()

    def _update_scrollbar(self, event=None):
        """Update the cool scrollbar appearance"""
        try:
            # Get scroll info from text widget
            top, bottom = self.file_display.yview()

            # Clear scrollbar
            self.cool_scrollbar.delete('all')

            # Get scrollbar dimensions
            sb_height = self.cool_scrollbar.winfo_height()
            sb_width = self.cool_scrollbar.winfo_width()

            if sb_height <= 1:  # Not configured yet
                self.cool_scrollbar.after(10, self._update_scrollbar)
                return

            # Draw track (subtle background)
            track_x = sb_width // 2
            self.cool_scrollbar.create_rectangle(track_x - 1, 0, track_x + 1, sb_height,
                                               fill='#e8eaed', outline='')

            # Calculate thumb dimensions
            visible_ratio = bottom - top
            if visible_ratio >= 1.0:  # No scrolling needed
                return

            thumb_height = max(20, int(sb_height * visible_ratio))
            thumb_top = int(top * (sb_height - thumb_height))
            thumb_bottom = thumb_top + thumb_height

            # Draw modern thumb with rounded appearance
            thumb_x1 = track_x - 3
            thumb_x2 = track_x + 3

            # Main thumb body
            self.scrollbar_thumb = self.cool_scrollbar.create_rectangle(
                thumb_x1, thumb_top + 2, thumb_x2, thumb_bottom - 2,
                fill='#5f6368', outline='', tags='thumb'
            )

            # Rounded ends
            self.cool_scrollbar.create_oval(thumb_x1, thumb_top, thumb_x2, thumb_top + 4,
                                          fill='#5f6368', outline='', tags='thumb')
            self.cool_scrollbar.create_oval(thumb_x1, thumb_bottom - 4, thumb_x2, thumb_bottom,
                                          fill='#5f6368', outline='', tags='thumb')

            # Store thumb info for dragging
            self.scrollbar_thumb_pos = thumb_top
            self.scrollbar_thumb_height = thumb_height
            self.scrollbar_track_height = sb_height

        except Exception:
            pass  # Ignore errors during updates

    def _on_scrollbar_click(self, event):
        """Handle scrollbar click"""
        try:
            click_y = event.y
            sb_height = self.cool_scrollbar.winfo_height()

            # Calculate target scroll position
            target_ratio = click_y / sb_height

            # Scroll to target
            self.file_display.yview_moveto(target_ratio)
            self._update_scrollbar()
        except Exception:
            pass

    def _on_scrollbar_drag(self, event):
        """Handle scrollbar drag"""
        try:
            drag_y = event.y
            sb_height = self.cool_scrollbar.winfo_height()

            # Calculate scroll position
            scroll_ratio = max(0, min(1, drag_y / sb_height))

            # Scroll to position
            self.file_display.yview_moveto(scroll_ratio)
            self._update_scrollbar()
        except Exception:
            pass

    def _on_scrollbar_configure(self, event):
        """Handle scrollbar resize"""
        self._update_scrollbar()

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling"""
        try:
            self.file_display.yview_scroll(int(-1 * (event.delta / 120)), 'units')
            self._update_scrollbar()
        except Exception:
            pass

    # Marmon scrollbar methods (similar to main scrollbar but for marmon section)
    def _setup_marmon_cool_scrollbar(self):
        """Setup the cool custom scrollbar for Marmon section"""
        self.marmon_scrollbar_thumb = None
        self.marmon_scrollbar_track_height = 0
        self.marmon_scrollbar_thumb_height = 0
        self.marmon_scrollbar_thumb_pos = 0

        # Bind scrollbar events
        self.marmon_cool_scrollbar.bind('<Button-1>', self._on_marmon_scrollbar_click)
        self.marmon_cool_scrollbar.bind('<B1-Motion>', self._on_marmon_scrollbar_drag)
        self.marmon_cool_scrollbar.bind('<Configure>', self._on_marmon_scrollbar_configure)

        # Bind text widget scroll events
        self.marmon_file_display.bind('<MouseWheel>', self._on_marmon_mousewheel)
        self.marmon_file_display.bind('<Configure>', self._update_marmon_scrollbar)

    def _on_marmon_scroll(self, *args):
        """Handle Marmon text widget scroll events"""
        self._update_marmon_scrollbar()

    def _update_marmon_scrollbar(self, event=None):
        """Update the Marmon scrollbar appearance"""
        try:
            top, bottom = self.marmon_file_display.yview()
            self.marmon_cool_scrollbar.delete('all')

            sb_height = self.marmon_cool_scrollbar.winfo_height()
            sb_width = self.marmon_cool_scrollbar.winfo_width()

            if sb_height <= 1:
                self.marmon_cool_scrollbar.after(10, self._update_marmon_scrollbar)
                return

            track_x = sb_width // 2
            self.marmon_cool_scrollbar.create_rectangle(track_x - 1, 0, track_x + 1, sb_height,
                                                       fill='#e8eaed', outline='')

            visible_ratio = bottom - top
            if visible_ratio >= 1.0:
                return

            thumb_height = max(20, int(sb_height * visible_ratio))
            thumb_top = int(top * (sb_height - thumb_height))
            thumb_bottom = thumb_top + thumb_height

            thumb_x1 = track_x - 3
            thumb_x2 = track_x + 3

            self.marmon_scrollbar_thumb = self.marmon_cool_scrollbar.create_rectangle(
                thumb_x1, thumb_top + 2, thumb_x2, thumb_bottom - 2,
                fill='#5f6368', outline='', tags='thumb'
            )

            self.marmon_cool_scrollbar.create_oval(thumb_x1, thumb_top, thumb_x2, thumb_top + 4,
                                                  fill='#5f6368', outline='', tags='thumb')
            self.marmon_cool_scrollbar.create_oval(
                thumb_x1, thumb_bottom - 4, thumb_x2, thumb_bottom,
                                                  fill='#5f6368', outline='', tags='thumb')

            self.marmon_scrollbar_thumb_pos = thumb_top
            self.marmon_scrollbar_thumb_height = thumb_height
            self.marmon_scrollbar_track_height = sb_height
        except Exception:
            pass

    def _on_marmon_scrollbar_click(self, event):
        """Handle Marmon scrollbar click"""
        try:
            click_y = event.y
            sb_height = self.marmon_cool_scrollbar.winfo_height()
            target_ratio = click_y / sb_height
            self.marmon_file_display.yview_moveto(target_ratio)
            self._update_marmon_scrollbar()
        except Exception:
            pass

    def _on_marmon_scrollbar_drag(self, event):
        """Handle Marmon scrollbar drag"""
        try:
            drag_y = event.y
            sb_height = self.marmon_cool_scrollbar.winfo_height()
            scroll_ratio = max(0, min(1, drag_y / sb_height))
            self.marmon_file_display.yview_moveto(scroll_ratio)
            self._update_marmon_scrollbar()
        except Exception:
            pass

    def _on_marmon_scrollbar_configure(self, event):
        """Handle Marmon scrollbar resize"""
        self._update_marmon_scrollbar()

    def _on_marmon_mousewheel(self, event):
        """Handle Marmon mouse wheel scrolling"""
        try:
            self.marmon_file_display.yview_scroll(int(-1 * (event.delta / 120)), 'units')
            self._update_marmon_scrollbar()
        except Exception:
            pass

    def browse_file(self):
        """Browse for input files - single file for UTC Main, multiple for others"""
        try:
            # Ensure the parent window is active
            self.root.focus_force()
            self.root.lift()

            # UTC Main and UTC FS only allow single file selection
            if self.report_type.get() == "utc_main":
                file_path = filedialog.askopenfilename(
                    parent=self.root,
                    title="Select UTC Main Excel File",
                    initialdir=os.path.expanduser("~/Documents/"),
                    filetypes=[
                        ("Excel files", "*.xlsx *.xls"),
                        ("All files", "*.*")
                    ]
                )
                file_paths = (file_path,) if file_path else ()
            elif self.report_type.get() == "utc_fs":
                file_path = filedialog.askopenfilename(
                    parent=self.root,
                    title="Select UTC FS Excel File",
                    initialdir=os.path.expanduser("~/Documents/"),
                    filetypes=[
                        ("Excel files", "*.xlsx *.xls"),
                        ("All files", "*.*")
                    ]
                )
                file_paths = (file_path,) if file_path else ()
            elif self.report_type.get() == "transco":
                file_path = filedialog.askopenfilename(
                    parent=self.root,
                    title="Select Transco Excel File",
                    initialdir=os.path.expanduser("~/Documents/"),
                    filetypes=[
                        ("Excel files", "*.xlsx *.xls"),
                        ("All files", "*.*")
                    ]
                )
                file_paths = (file_path,) if file_path else ()
            else:
                # Other report types allow multiple files
                file_paths = filedialog.askopenfilenames(
                    parent=self.root,
                    title="Select TMS Excel Files (Hold Ctrl+Click for Multiple Selection)",
                    initialdir=os.path.expanduser("~/Documents/"),
                    filetypes=[
                        ("Excel files", "*.xlsx *.xls"),
                        ("All files", "*.*")
                    ]
                )
        except Exception as e:
            print(f"File dialog error: {e}")
            # Fallback to single file selection
            file_path = filedialog.askopenfilename(
                parent=self.root,
                title="Select TMS Excel File (Single Selection Only - Dialog Issue)",
                initialdir=os.path.expanduser("~/Documents"),
                filetypes=[
                    ("Excel files", "*.xlsx *.xls"),
                    ("All files", "*.*")
                ]
            )
            file_paths = (file_path,) if file_path else ()

        # Debug: Log the selection results
        if file_paths:
            print(f"Selected {len(file_paths)} files:")
            for i, path in enumerate(file_paths):
                print(f"  {i+1}. {os.path.basename(path)}")

            self.input_files = list(file_paths)
            self.update_file_display()
            self.update_process_button_state()

            # Show success message for multi-file selection
            if len(file_paths) > 1:
                messagebox.showinfo(
                    "Multi-File Selection Successful",
                    f"✅ Successfully selected {len(file_paths)} files for processing:\n\n" +
                    "\n".join([f"• {os.path.basename(f)}" for f in file_paths])
                )
        else:
            print("No files selected")

    def update_file_display(self):
        """Update the file display to show all selected filenames with auto-sizing"""
        self.file_display.config(state='normal')
        self.file_display.delete('1.0', tk.END)

        if self.input_files:
            file_count = len(self.input_files)

            if file_count == 1:
                # Single file - show filename with checkmark
                filename = os.path.basename(self.input_files[0])
                display_text = f"✅ Selected: {filename}"
                self.file_display.insert('1.0', display_text)
                self.file_display.config(fg='#0d9488')
                # Set height for single file
                optimal_height = 2
            else:
                # Multiple files - show count and list all filenames
                header = f"✅ {file_count} files selected:\n\n"
                self.file_display.insert('1.0', header)

                # Add each filename on a new line
                for i, file_path in enumerate(self.input_files, 1):
                    filename = os.path.basename(file_path)
                    file_line = f"{i}. {filename}\n"
                    self.file_display.insert(tk.END, file_line)

                self.file_display.config(fg='#0d9488')
                # Calculate optimal height: header (2 lines) + files + padding
                # Min 3, max 6 lines (compact for max 10 files)
                optimal_height = min(max(file_count + 1, 3), 6)

            # Auto-resize the display based on content
            self.file_display.config(height=optimal_height)
        else:
            self.file_display.insert('1.0', "No files selected")
            self.file_display.config(fg='#6c757d', height=2)

        self.file_display.config(state='disabled')

    def update_process_button_state(self):
        """Enable process button based on current selection and input state"""
        # File processing mode - all report types use the main process button
        if True:
            # File processing mode - show main process button
            if hasattr(self, 'process_button'):
                self.process_button.grid()  # Show the button

            # Check files
            if self.input_files:
                file_count = len(self.input_files)
                button_text = f"🚀 PROCESS {file_count} FILE{'S' if file_count > 1 else ''}"
                self.process_button.config(state="normal", text=button_text)
            else:
                self.process_button.config(state="disabled", text="🚀 PROCESS FILE")


    def process_file(self):
        """Process the selected files"""
        if self.is_processing:
            return

        # File processing mode
        if not self.input_files:
            return

        # Set processing state
        self.is_processing = True

        # Update UI for processing state
        file_count = len(self.input_files)
        self.process_button.config(
            state="disabled",
            text=f"⏳ PROCESSING {file_count} FILE{'S' if file_count > 1 else ''}..."
        )

        # Start processing in separate thread
        thread = threading.Thread(target=self._process_file_thread)
        thread.daemon = True
        thread.start()

    def _process_file_thread(self):
        """Process files in background thread"""
        try:
            # Handle single vs multiple file processing differently
            if len(self.input_files) == 1:
                # Single file - ask for specific output file location
                input_name = os.path.splitext(os.path.basename(self.input_files[0]))[0]
                report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                default_name = f"{input_name}_processed_{report_type}.xlsx"

                output_file = filedialog.asksaveasfilename(
                    title="Save Processed File As",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )

                if not output_file:
                    self.root.after(0, self._reset_ui)
                    return

                output_folder = None  # Not used for single file
            else:
                # Multiple files - auto-create timestamped folder
                from datetime import datetime

                # Get base directory (use Desktop by default,
                # or same location as first input file)
                if self.input_files:
                    base_dir = os.path.dirname(self.input_files[0])
                else:
                    base_dir = os.path.join(os.path.expanduser("~"), "Desktop")

                # Create timestamped folder name
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                report_type = "basic" if self.report_type.get() == "basic" else "detailed"
                folder_name = f"TMS_Processed_{report_type}_{timestamp}"
                output_folder = os.path.join(base_dir, folder_name)

                # Create the folder
                try:
                    os.makedirs(output_folder, exist_ok=True)
                    print(f"Created output folder: {output_folder}")
                except Exception as e:
                    print(f"Error creating folder: {e}")
                    # Fallback to user selection if auto-creation fails
                    output_folder = filedialog.askdirectory(
                        title="Auto-folder creation failed. Select Folder to Save Processed Files"
                    )
                    if not output_folder:
                        self.root.after(0, self._reset_ui)
                        return

                output_file = None  # Will be generated per file

            # Select processor based on report type
            if self.report_type.get() == "basic":
                processor = self.basic_processor
            elif self.report_type.get() == "utc_main":
                processor = self.utc_main_processor
            elif self.report_type.get() == "utc_fs":
                processor = self.utc_fs_processor
            elif self.report_type.get() == "transco":
                processor = self.transco_processor
            else:
                # Import detailed processor when needed
                if self.detailed_processor is None:
                    try:
                        from tms_detailed_processor import TMSDetailedDataProcessor
                        self.detailed_processor = TMSDetailedDataProcessor()
                    except ImportError as e:
                        error_msg = f"Failed to load detailed processor: {e}"
                        self.root.after(0, lambda: messagebox.showerror("Error",
                                                                      error_msg))
                        self.root.after(0, self._reset_ui)
                        return
                processor = self.detailed_processor

            # Process files
            processed_count = 0
            total_files = len(self.input_files)
            all_stats = []
            processed_files = []

            if len(self.input_files) == 1:
                # Single file processing
                try:
                    # Update button text
                    self.root.after(0, lambda: self.process_button.config(
                        text="⏳ PROCESSING FILE..."
                    ))

                    # Process the data based on processor type
                    if self.report_type.get() == "utc_main":
                        processed_data = processor.process_excel_file(self.input_files[0])
                        # UTC Main processor needs custom save logic (Basic_Processor style)
                        self._save_utc_main_data(
                            processed_data, processor.title_info, output_file
                        )
                    elif self.report_type.get() == "utc_fs":
                        processed_data = processor.process_excel_file(self.input_files[0])
                        # UTC FS processor needs custom save logic (Basic_Processor style)
                        self._save_utc_main_data(
                            processed_data, processor.title_info, output_file
                        )
                    elif self.report_type.get() == "transco":
                        processed_data = processor.process_excel_file(self.input_files[0])
                        # Transco processor needs custom save logic (Basic_Processor style)
                        self._save_utc_main_data(
                            processed_data, processor.title_info, output_file
                        )
                    else:
                        processed_data = processor.clean_and_process_data(self.input_files[0])
                        # Save using the standard method
                        processor.save_processed_data(output_file)

                    # Collect stats
                    stats = processor.summary_stats.copy()
                    stats['filename'] = os.path.basename(self.input_files[0])
                    stats['output_file'] = os.path.basename(output_file)
                    all_stats.append(stats)
                    processed_files.append(output_file)
                    processed_count = 1

                except Exception as file_error:
                    print(f"Error processing {self.input_files[0]}: {file_error}")
            else:
                # Multiple files processing
                for i, input_file in enumerate(self.input_files, 1):
                    try:
                        # Update button text with progress
                        self.root.after(0, lambda: self.process_button.config(
                            text=f"⏳ PROCESSING {i}/{total_files}..."
                        ))

                        # Process the data based on processor type
                        if self.report_type.get() == "utc_main":
                            processed_data = processor.process_excel_file(input_file)
                            # Generate output filename for each file
                            input_name = os.path.splitext(os.path.basename(input_file))[0]
                            file_output = os.path.join(
                                output_folder, f"{input_name}_processed_utc_main.xlsx"
                            )
                            # UTC Main processor needs custom save logic
                            self._save_utc_main_data(
                                processed_data, processor.title_info, file_output
                            )
                        elif self.report_type.get() == "utc_fs":
                            processed_data = processor.process_excel_file(input_file)
                            # Generate output filename for each file
                            input_name = os.path.splitext(os.path.basename(input_file))[0]
                            file_output = os.path.join(
                                output_folder, f"{input_name}_processed_utc_fs.xlsx"
                            )
                            # UTC FS processor needs custom save logic
                            self._save_utc_main_data(
                                processed_data, processor.title_info, file_output
                            )
                        elif self.report_type.get() == "transco":
                            processed_data = processor.process_excel_file(input_file)
                            # Generate output filename for each file
                            input_name = os.path.splitext(os.path.basename(input_file))[0]
                            file_output = os.path.join(
                                output_folder, f"{input_name}_processed_transco.xlsx"
                            )
                            # Transco processor needs custom save logic
                            self._save_utc_main_data(
                                processed_data, processor.title_info, file_output
                            )
                        else:
                            processed_data = processor.clean_and_process_data(input_file)
                            # Generate output filename for each file
                            input_name = os.path.splitext(os.path.basename(input_file))[0]
                            report_type = (
                                "basic" if self.report_type.get() == "basic"
                                else "detailed"
                            )
                            file_output = os.path.join(
                                output_folder,
                                f"{input_name}_processed_{report_type}.xlsx"
                            )
                            # Save using the standard method
                            processor.save_processed_data(file_output)

                        # Collect stats
                        stats = processor.summary_stats.copy()
                        stats['filename'] = os.path.basename(input_file)
                        stats['output_file'] = os.path.basename(file_output)
                        all_stats.append(stats)
                        processed_files.append(file_output)
                        processed_count += 1

                    except Exception as file_error:
                        print(f"Error processing {input_file}: {file_error}")
                        continue

            # Show success message
            if processed_count > 0:
                total_loads = sum(stat['total_loads'] for stat in all_stats)
                total_savings = sum(stat['total_potential_savings'] for stat in all_stats)

                # Save individual file stats to history
                current_type = self.report_type.get()
                if current_type == "basic":
                    report_type = "basic"
                elif current_type == "utc_main":
                    report_type = "utc_main"
                elif current_type == "utc_fs":
                    report_type = "utc_fs"
                else:
                    report_type = "detailed"
                for i, stat in enumerate(all_stats):
                    file_stats = {
                        'total_potential_savings': stat['total_potential_savings'],
                        'total_loads': stat['total_loads'],
                        'percentage_savings': stat['percentage_savings'],
                        'loads_with_savings': stat['loads_with_savings']
                    }
                    file_name = stat.get('filename', f'file_{i+1}')
                    self.save_savings_history(file_stats, report_type, 1, file_names=[file_name],
                                           file_index=i+1, total_files=len(all_stats))

                if len(self.input_files) == 1:
                    # Single file success message
                    self.root.after(0, lambda: messagebox.showinfo("Processing Complete",
                        f"🎉 Successfully processed file!\n\n"
                        f"📊 Total Loads: {total_loads:,}\n"
                        f"💵 Total Potential Savings: ${total_savings:,.2f}\n\n"
                        f"📁 File saved as:\n{os.path.basename(output_file)}"))
                else:
                    # Multiple files success message
                    files_list = "\n".join([f"✅ {stat['output_file']}" for stat in all_stats])

                    self.root.after(0, lambda: messagebox.showinfo("Batch Processing Complete",
                        f"🎉 Successfully processed {processed_count}/{total_files} files!\n\n"
                        f"📊 Total Loads: {total_loads:,}\n"
                        f"💵 Total Potential Savings: ${total_savings:,.2f}\n\n"
                        f"📁 Files saved to:\n{output_folder}\n\n"
                        f"Processed files:\n{files_list}"))
            else:
                self.root.after(0, lambda: messagebox.showerror("Error",
                    "Failed to process any files. Please check your input files."))

        except Exception as e:
            error_msg = str(e)
            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Error", f"An error occurred:\n{error_msg}"
                )
            )

        finally:
            self.root.after(0, self._reset_ui)

    def _reset_ui(self):
        """Reset UI to normal state"""
        self.is_processing = False
        self.process_button.config(state="normal", text="🚀 PROCESS FILE")
        self.update_process_button_state()

    def load_savings_history(self):
        """Load savings history from JSON file"""
        try:
            if self.savings_history_file.exists():
                with open(self.savings_history_file, 'r') as f:
                    return json.load(f)
            return []
        except Exception as e:
            print(f"Warning: Could not load savings history: {e}")
            return []

    def save_savings_history(
            self, stats, report_type, file_count, file_names=None,
            file_index=None, total_files=None
    ):
        """Save current processing stats to history"""
        try:
            # Get file names from the current input files
            if file_names is None and hasattr(self, 'input_files') and self.input_files:
                file_names = [os.path.basename(f) for f in self.input_files]
            elif file_names is None:
                file_names = []

            # Create new entry with proper type conversion
            entry = {
                'timestamp': datetime.now().isoformat(),
                'report_type': str(report_type),
                'file_count': int(file_count),
                'file_names': file_names if file_names else [],
                'file_index': int(file_index) if file_index is not None else None,
                'total_files': int(total_files) if total_files is not None else None,
                'total_potential_savings': float(stats.get('total_potential_savings', 0)),
                'total_loads': int(stats.get('total_loads', 0)),
                'percentage_savings': float(stats.get('percentage_savings', 0)),
                'loads_with_savings': int(stats.get('loads_with_savings', 0))
            }

            # Add to history
            self.savings_history.insert(0, entry)  # Add to beginning

            # Keep only last 20 entries
            self.savings_history = self.savings_history[:20]

            # Save to file
            with open(self.savings_history_file, 'w') as f:
                json.dump(self.savings_history, f, indent=2)

            # Update UI display
            self.update_savings_display()

        except Exception as e:
            print(f"Warning: Could not save savings history: {e}")

    def update_savings_display(self):
        """Update the savings statistics display"""
        if hasattr(self, 'stats_display_frame'):
            # Clear existing content
            for widget in self.stats_display_frame.winfo_children():
                widget.destroy()

            if self.savings_history:
                # Header with clear button
                header_frame = tk.Frame(
                    self.stats_display_frame, bg=UI_COLORS['BACKGROUND_WHITE']
                )
                header_frame.pack(fill='x', pady=(1, 1), padx=1)

                header_label = tk.Label(
                    header_frame,
                    text=f"📊 Recent Uploads (Last {len(self.savings_history)})",
                    font=('Segoe UI', 12, 'bold'),
                    bg=UI_COLORS['BACKGROUND_WHITE'],
                    fg=UI_COLORS['TEXT_PRIMARY']
                )
                header_label.pack(side='left')

                clear_button = tk.Button(
                    header_frame,
                    text="🗑️ Clear",
                    font=('Segoe UI', 10),
                    fg=UI_COLORS['ERROR_RED'],
                    bg=UI_COLORS['BACKGROUND_WHITE'],
                    relief='flat',
                    cursor='hand2',
                    command=self.clear_savings_history
                )
                clear_button.pack(side='right')

                # Show individual upload records
                for i, entry in enumerate(self.savings_history):
                    upload_date = datetime.fromisoformat(
                        entry['timestamp']
                    ).strftime('%m/%d %I:%M%p')
                    report_type = entry['report_type'].title()
                    file_index = entry.get('file_index')
                    total_files = entry.get('total_files')

                    # Build date text with file index if available
                    if file_index and total_files and total_files > 1:
                        date_text = (
                            f"🕒 {upload_date} • {report_type} • "
                            f"{file_index} of {total_files}"
                        )
                    else:
                        date_text = f"🕒 {upload_date} • {report_type}"

                    # Create frame for each upload record
                    record_frame = tk.Frame(
                        self.stats_display_frame, bg=UI_COLORS['BACKGROUND_WHITE']
                    )
                    record_frame.pack(fill='x', padx=1, pady=0)

                    # Date and type
                    date_label = tk.Label(
                        record_frame,
                        text=date_text,
                        font=('Segoe UI', 10),
                        bg=UI_COLORS['BACKGROUND_WHITE'],
                        fg=UI_COLORS['TEXT_MUTED'],
                        anchor='w'
                    )
                    date_label.pack(fill='x')

                    # File names (if available)
                    file_names = entry.get('file_names', [])
                    if file_names:
                        if len(file_names) == 1:
                            file_text = f"📄 {file_names[0]}"
                        elif len(file_names) <= 3:
                            file_text = f"📄 {', '.join(file_names)}"
                        else:
                            file_text = (
                                f"📄 {', '.join(file_names[:2])}, "
                                f"+{len(file_names)-2} more"
                            )

                        file_label = tk.Label(
                            record_frame,
                            text=file_text,
                            font=('Segoe UI', 10),
                            bg=UI_COLORS['BACKGROUND_WHITE'],
                            fg=UI_COLORS['TEXT_SECONDARY'],
                            anchor='w'
                        )
                        file_label.pack(fill='x', padx=(10, 0))

                    # Savings amount
                    savings_label = tk.Label(
                        record_frame,
                        text=f"💵 Potential Savings: ${entry['total_potential_savings']:,.2f}",
                        font=('Segoe UI', 9),
                        bg=UI_COLORS['BACKGROUND_WHITE'],
                        fg=UI_COLORS['SUCCESS_GREEN'],
                        anchor='w'
                    )
                    savings_label.pack(fill='x', padx=(10, 0))

                    # Add separator line if not the last item
                    if i < len(self.savings_history) - 1:
                        separator = tk.Frame(
                            record_frame, height=1,
                            bg=UI_COLORS['BACKGROUND_BORDER']
                        )
                        separator.pack(fill='x', pady=(3, 0))

    def clear_savings_history(self):
        """Clear all savings history with confirmation"""
        if not self.savings_history:
            return

        result = messagebox.askyesno(
            "Clear Recent Uploads",
            f"Are you sure you want to clear all {len(self.savings_history)} "
            f"recent upload records?\n\n"
            "This action cannot be undone.",
            icon='warning'
        )

        if result:
            self.savings_history = []
            # Delete the history file
            try:
                if self.savings_history_file.exists():
                    self.savings_history_file.unlink()
            except Exception as e:
                print(f"Warning: Could not delete history file: {e}")

            # Update the display
            self.update_savings_display()

            messagebox.showinfo("Cleared", "Recent uploads history has been cleared.")

    def on_marmon_client_select(self, event=None):
        """Handle Marmon client selection"""
        selected_client = self.marmon_client_var.get()
        if selected_client:
            # Update file display placeholder
            if hasattr(self, 'marmon_file_display'):
                self.marmon_file_display.config(state='normal')
                self.marmon_file_display.delete('1.0', tk.END)
                self.marmon_file_display.insert(
                    '1.0', f"Upload raw data file for {selected_client}")
                self.marmon_file_display.config(state='disabled', fg='#6c757d')
                self._update_marmon_scrollbar()

    def browse_marmon_file(self):
        """Browse for Marmon raw data file"""
        try:
            self.root.focus_force()
            self.root.lift()

            selected_client = self.marmon_client_var.get()
            if not selected_client:
                messagebox.showwarning(
                    "No Client Selected", "Please select a Marmon client first.")
                return

            filetypes = [
                ('Excel files', '*.xlsx *.xls'),
                ('All files', '*.*')
            ]

            file_path = filedialog.askopenfilename(
                title=f"Select {selected_client} Raw Data File",
                filetypes=filetypes,
                initialdir=os.path.expanduser("~")
            )

            if file_path:
                self.marmon_input_files = [file_path]
                self.update_marmon_file_display([file_path])
                self.update_marmon_process_button_state()

        except Exception as e:
            messagebox.showerror("Browse Error", f"Error browsing for file: {str(e)}")

    def update_marmon_file_display(self, files):
        """Update the Marmon file display with selected files"""
        if not hasattr(self, 'marmon_file_display'):
            return

        try:
            self.marmon_file_display.config(state='normal')
            self.marmon_file_display.delete('1.0', tk.END)

            if not files:
                self.marmon_file_display.insert('1.0', "Select a client and upload raw data file")
                self.marmon_file_display.config(fg='#6c757d')
            else:
                # Display selected files
                file_list = []
                for file_path in files:
                    filename = os.path.basename(file_path)
                    file_size = os.path.getsize(file_path)
                    size_mb = file_size / (1024 * 1024)
                    file_list.append(f"📄 {filename} ({size_mb:.1f} MB)")

                self.marmon_file_display.insert('1.0', '\n'.join(file_list))
                self.marmon_file_display.config(fg='#000000')

            self.marmon_file_display.config(state='disabled')
            self._update_marmon_scrollbar()

        except Exception as e:
            print(f"Error updating Marmon file display: {e}")

    def update_marmon_process_button_state(self):
        """Enable/disable Marmon process button based on selections"""
        if not hasattr(self, 'marmon_process_button'):
            return

        selected_client = self.marmon_client_var.get()
        has_files = hasattr(self, 'marmon_input_files') and self.marmon_input_files

        if selected_client and has_files:
            self.marmon_process_button.config(state="normal",
                                             text=f"🏭 PROCESS {selected_client} SPECIAL")
        else:
            self.marmon_process_button.config(state="disabled",
                                             text="🏭 PROCESS")

    def process_marmon_file(self):
        """Process Marmon special client files with special city logic"""
        try:
            selected_client = self.marmon_client_var.get()
            if not selected_client:
                messagebox.showwarning(
                    "No Client Selected", "Please select a Marmon client first.")
                return

            if not hasattr(self, 'marmon_input_files') or not self.marmon_input_files:
                messagebox.showwarning("No File Selected", "Please select a raw data file first.")
                return

            # Use Basic processing first, then apply special city rules
            input_file = self.marmon_input_files[0]

            print(f"[MARMON] Step 1: Running Basic processing for {selected_client}")

            # Run the same Basic processing logic
            processed_df = self.basic_processor.process_file(input_file)

            print(f"[MARMON] Step 2: Applying {selected_client} special city rules")

            # Define special cities for Marmon clients
            special_cities = ['EVANSTON', 'GREEN RIVER', 'MILES CITY']

            # Apply special city modifications on top of Basic results
            processed_df = self._apply_marmon_special_modifications(
                processed_df, selected_client, special_cities
            )

            # Generate output filename
            timestamp = datetime.now().strftime("%m.%d.%y_%H.%M")
            base_name = os.path.splitext(os.path.basename(input_file))[0]
            output_filename = f"{selected_client}_{base_name}_SPECIAL_{timestamp}.xlsx"

            # Ask user where to save
            output_file = filedialog.asksaveasfilename(
                title=f"Save {selected_client} Special Processing Results",
                defaultextension=".xlsx",
                initialfile=output_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if output_file:
                # Save the processed data
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    processed_df.to_excel(writer, sheet_name='Processed Data', index=False)

                # Show success message
                # Count special rows (will be calculated in the modification method)
                special_rows = getattr(self, '_last_special_rows_count', 0)

                messagebox.showinfo("Marmon Special Processing Complete",
                                   f"✅ {selected_client} special processing completed!\n\n"
                                   f"📁 File: {os.path.basename(output_file)}\n"
                                   f"📊 Total rows: {len(processed_df)}\n"
                                   f"🏭 Special city rows: {special_rows}\n"
                                   f"🎯 Special cities: {', '.join(special_cities)}")
            else:
                messagebox.showinfo("Cancelled", "Processing cancelled by user.")

        except Exception as e:
            messagebox.showerror("Processing Error", f"Error processing Marmon file: {str(e)}")

    def _apply_marmon_special_modifications(self, df, client_name, special_cities):
        """Apply special city modifications to already-processed Basic results"""

        # Find city column in the processed data
        city_column = None
        for col in df.columns:
            if ('city' in col.lower() and
                    ('dest' in col.lower() or 'to' in col.lower() or
                     'delivery' in col.lower())):
                city_column = col
                break

        if not city_column:
            print(f"[MARMON] No city column found - skipping special modifications")
            self._last_special_rows_count = 0
            return df

        print(f"[MARMON] Using column '{city_column}' for special city modifications")

        # Find special city rows
        special_mask = df[city_column].str.upper().isin([city.upper() for city in special_cities])
        special_count = special_mask.sum()
        self._last_special_rows_count = special_count

        if special_count > 0:
            print(f"[MARMON] Modifying PS for {special_count} {client_name} special city rows")

            # Apply simple PS modifications based on client
            if 'Potential Savings' in df.columns:
                if client_name == "UTC MAIN":
                    # UTC MAIN: Increase PS by 15% for remote locations
                    df.loc[special_mask, 'Potential Savings'] *= 1.15
                elif client_name == "UTC FS":
                    # UTC FS: Increase PS by 20% for priority routes
                    df.loc[special_mask, 'Potential Savings'] *= 1.20
                elif client_name == "TRANSCO":
                    # TRANSCO: Decrease PS by 10% for consolidated shipping
                    df.loc[special_mask, 'Potential Savings'] *= 0.90

                print(f"[MARMON] Applied {client_name} PS modifications to {special_count} rows")
            else:
                print(f"[MARMON] Warning: 'Potential Savings' column not found")
        else:
            print(f"[MARMON] No special cities found in data")

        return df

    def _save_utc_main_data(self, df, title_info, output_file):
        """Save processed UTC Main data to Excel with professional formatting"""
        if df is None or df.empty:
            raise Exception("No data to save")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UTC Main Report"

        # Add title information
        row = 1
        if title_info:
            if 'company_name' in title_info:
                cell = ws.cell(row=row, column=1, value=f"Company: {title_info['company_name']}")
                cell.font = Font(size=12, bold=True)
                row += 1

            if 'date_range' in title_info:
                cell = ws.cell(
                    row=row, column=1,
                    value=f"Report Period: {title_info['date_range']}"
                )
                cell.font = Font(size=11)
                row += 1

            row += 1  # Empty row

        # Add headers with color coding
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add data rows
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Format currency columns
                header_name = headers[col_idx-1]
                if any(cost_term in header_name for cost_term in ['Cost', 'Savings']):
                    cell.number_format = '"$"#,##0.00'

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        print(f"Saved UTC Main report: {output_file}")


def main():
    root = tk.Tk()
    app = ModernTMSProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()

