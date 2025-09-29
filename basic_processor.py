#!/usr/bin/env python3
"""
Basic TMS Processor - Core processing logic shared by all processors
This contains the fundamental TMS business rules that apply to all clients
"""

import pandas as pd
import os
from typing import Dict, Any
from processor_interface import TMSProcessorInterface


class BasicTMSProcessor(TMSProcessorInterface):
    """
    Core TMS processing logic shared by UTC Main, UTC FS, and Transco
    Contains all the fundamental business rules without city-specific customizations
    """

    def __init__(self):
        # File structure settings
        self.HEADER_ROW = 8          # Row 9 in Excel (0-indexed) - actual header row
        self.DATA_START_ROW = 10     # Row 11 in Excel (0-indexed) - first data row

        # TL carriers requiring special processing
        self.TL_CARRIERS = {
            'LANDSTAR RANGER INC',
            'SMARTWAY TRANSPORTATION INC',
            'ONX LOGISTICS INC'
        }

        # Results storage
        self.processed_data = None
        self.summary_stats = {}
        self.title_info = {}

    def load_data(self, file_path: str) -> pd.DataFrame:
        """Load and clean Excel data using Basic_Processor logic"""
        # Read Excel file
        df_raw = pd.read_excel(file_path, sheet_name=0, header=None)

        # Extract title information
        self.title_info = self._extract_title_info(df_raw)

        # Get headers and data
        headers = df_raw.iloc[self.HEADER_ROW].fillna('').astype(str).tolist()
        data_df = df_raw.iloc[self.DATA_START_ROW:].copy()
        data_df.columns = headers

        # Remove columns with empty/nan headers
        valid_columns = []
        for i, header in enumerate(headers):
            if header.strip() != '' and header.lower() not in ['nan', 'unnamed']:
                valid_columns.append(i)

        if valid_columns:
            data_df = data_df.iloc[:, valid_columns]

        # Map column names to expected format
        data_df = self._map_columns(data_df)

        # Handle PS -> Potential Savings column mapping
        if 'PS' in data_df.columns:
            data_df = data_df.rename(columns={'PS': 'Potential Savings'})

        # Clean the data
        data_df = self._clean_data(data_df)

        # Calculate Potential Savings if missing
        data_df = self._calculate_potential_savings(data_df)

        return data_df

    def apply_basic_business_rules(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply the 5 core TMS business rules (excluding city-specific rules)"""
        df = df.copy()
        print("Applying Basic TMS business rules...")

        # Rule 1: Same Carrier Rule
        df = self._apply_same_carrier_rule(df)

        # Rule 2: Empty Least Cost Data Rule
        df = self._apply_empty_data_rule(df)

        # Rule 3: Negative Savings Rule
        df = self._apply_negative_savings_rule(df)

        # Rule 4: TL Carriers Rule
        df = self._apply_tl_carriers_rule(df)

        # Rule 5: DALKO DEFENDER INSURANCE Rule
        df = self._apply_dalko_rule(df)

        return df

    def _apply_same_carrier_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """Rule 1: Set PS to 0 when selected carrier = least cost carrier"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                (df['Selected Carrier'].astype(str) == df['Least Cost Carrier'].astype(str)) &
                (df['Selected Carrier'].notna()) &
                (df['Least Cost Carrier'].notna()) &
                (df['Selected Carrier'].astype(str) != '') &
                (df['Least Cost Carrier'].astype(str) != '')
            )

            count = mask.sum()
            if count > 0 and 'Potential Savings' in df.columns:
                df.loc[mask, 'Potential Savings'] = 0
                print(f"Same Carrier Rule: {count} rows")

        return df

    def _apply_empty_data_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """Rule 2: Copy selected data when least cost data is missing"""
        if 'Least Cost Carrier' in df.columns:
            mask = (
                df['Least Cost Carrier'].isna() |
                (df['Least Cost Carrier'].astype(str) == '') |
                (df['Least Cost Carrier'].astype(str) == 'nan')
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"Empty Data Rule: {count} rows")

        return df

    def _apply_negative_savings_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """Rule 3: Fix negative potential savings"""
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            mask = ps_numeric < 0
            count = mask.sum()

            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                df.loc[mask, 'Potential Savings'] = 0
                print(f"Negative Savings Rule: {count} rows")

        return df

    def _apply_tl_carriers_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """Rule 4: Special handling for TL carriers"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            mask = (
                df['Selected Carrier'].astype(str).str.upper().isin(
                    [c.upper() for c in self.TL_CARRIERS]
                ) |
                df['Least Cost Carrier'].astype(str).str.upper().isin(
                    [c.upper() for c in self.TL_CARRIERS]
                )
            )

            count = mask.sum()
            if count > 0:
                self._copy_selected_to_least_cost(df, mask)
                if 'Potential Savings' in df.columns:
                    df.loc[mask, 'Potential Savings'] = 0
                print(f"TL Carriers Rule: {count} rows")

        return df

    def _apply_dalko_rule(self, df: pd.DataFrame) -> pd.DataFrame:
        """Rule 5: DALKO DEFENDER INSURANCE pattern matching"""
        if 'Selected Carrier' in df.columns and 'Least Cost Carrier' in df.columns:
            dalko_matches = []

            for idx, row in df.iterrows():
                selected = str(row['Selected Carrier']).strip().upper()
                least_cost = str(row['Least Cost Carrier']).strip().upper()

                # Skip empty or nan values
                if selected in ['', 'NAN', 'NONE'] or least_cost in ['', 'NAN', 'NONE']:
                    continue

                # Check for DALKO DEFENDER INSURANCE patterns
                # Pattern 1: "DALKO DEFENDER INSURANCE/XXX" where least cost is "XXX"
                if 'DALKO DEFENDER INSURANCE/' in selected:
                    carrier_after_slash = selected.split('DALKO DEFENDER INSURANCE/')[-1].strip()
                    if carrier_after_slash == least_cost:
                        dalko_matches.append(idx)

                # Pattern 2: "XXX/DALKO DEFENDER INSURANCE" where least cost is "XXX"
                elif '/DALKO DEFENDER INSURANCE' in selected:
                    carrier_before_slash = selected.split('/DALKO DEFENDER INSURANCE')[0].strip()
                    if carrier_before_slash == least_cost:
                        dalko_matches.append(idx)

            count = len(dalko_matches)
            if count > 0:
                dalko_mask = df.index.isin(dalko_matches)
                self._copy_selected_to_least_cost(df, dalko_mask)
                if 'Potential Savings' in df.columns:
                    df.loc[dalko_mask, 'Potential Savings'] = 0
                print(f"DALKO DEFENDER INSURANCE Rule: {count} rows")

        return df

    def _copy_selected_to_least_cost(self, df: pd.DataFrame, mask: pd.Series) -> None:
        """Helper: Copy selected carrier data to least cost columns"""
        column_pairs = [
            ('Selected Carrier', 'Least Cost Carrier'),
            ('Selected Service Type', 'Least Cost Service Type'),
            ('Selected Transit Days', 'Least Cost Transit Days'),
            ('Selected Freight Cost', 'Least Cost Freight Cost'),
            ('Selected Accessorial Cost', 'Least Cost Accessorial Cost'),
            ('Selected Total Cost', 'Least Cost Total Cost')
        ]

        for selected_col, least_cost_col in column_pairs:
            if selected_col in df.columns and least_cost_col in df.columns:
                df.loc[mask, least_cost_col] = df.loc[mask, selected_col]

    def _map_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Map column names to standardized format"""
        df = df.copy()

        # Find carrier columns - there should be two "Carrier" columns
        new_columns = list(df.columns)
        carrier_positions = []

        for i, col in enumerate(new_columns):
            if col == 'Carrier':
                carrier_positions.append(i)

        # Rename the two "Carrier" columns
        if len(carrier_positions) >= 2:
            new_columns[carrier_positions[0]] = 'Selected Carrier'
            new_columns[carrier_positions[1]] = 'Least Cost Carrier'

        # Map other column names to standardized format
        column_mapping = {
            'Service Type': 'Selected Service Type',  # First occurrence
            'Transit\nDays': 'Selected Transit Days',  # First occurrence
            'Freight + Fuel': 'Selected Freight Cost',  # First occurrence
            'Total Acc.': 'Selected Accessorial Cost',  # First occurrence
            'Total Cost ': 'Selected Total Cost'  # First occurrence (note the trailing space)
        }

        # Apply mapping to first occurrence of each column
        for old_name, new_name in column_mapping.items():
            if old_name in new_columns:
                # Find first occurrence and rename it
                idx = new_columns.index(old_name)
                new_columns[idx] = new_name

        # Now handle the second set (Least Cost columns)
        least_cost_mapping = {
            'Service Type': 'Least Cost Service Type',
            'Transit\nDays': 'Least Cost Transit Days',
            'Freight + Fuel': 'Least Cost Freight Cost',
            'Total Acc.': 'Least Cost Accessorial Cost',
            'Total Cost ': 'Least Cost Total Cost'
        }

        # Find and rename second occurrences
        for old_name, new_name in least_cost_mapping.items():
            # Count occurrences and rename the second one
            occurrences = [i for i, col in enumerate(new_columns) if col == old_name]
            if len(occurrences) >= 2:
                new_columns[occurrences[1]] = new_name

        df.columns = new_columns
        return df

    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and validate data types"""
        df = df.copy()

        # Remove rows with missing Load No.
        if 'Load No.' in df.columns:
            df = df.dropna(subset=['Load No.'])
            df = df[df['Load No.'].astype(str).str.strip() != '']
            df = df[df['Load No.'].astype(str) != 'nan']

        # Clean numeric columns
        numeric_columns = [
            'Selected Transit Days', 'Selected Freight Cost',
            'Selected Accessorial Cost', 'Selected Total Cost',
            'Least Cost Transit Days', 'Least Cost Freight Cost',
            'Least Cost Accessorial Cost', 'Least Cost Total Cost',
            'Potential Savings'
        ]

        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Clean string columns
        string_columns = [
            'Selected Carrier', 'Selected Service Type',
            'Least Cost Carrier', 'Least Cost Service Type'
        ]

        for col in string_columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')

        # Remove mostly empty rows
        df = df.dropna(thresh=5).reset_index(drop=True)

        return df

    def _calculate_potential_savings(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate Potential Savings if missing"""
        df = df.copy()

        # If PS already exists and has valid data, keep it
        if 'Potential Savings' in df.columns:
            ps_numeric = pd.to_numeric(df['Potential Savings'], errors='coerce')
            if ps_numeric.notna().sum() > 0:
                return df

        # Calculate PS from cost difference
        if 'Selected Total Cost' in df.columns and 'Least Cost Total Cost' in df.columns:
            selected_cost = pd.to_numeric(df['Selected Total Cost'], errors='coerce').fillna(0)
            least_cost = pd.to_numeric(df['Least Cost Total Cost'], errors='coerce').fillna(0)
            df['Potential Savings'] = selected_cost - least_cost

            # Handle cases where least cost is 0
            mask_zero_least = least_cost == 0
            df.loc[mask_zero_least, 'Potential Savings'] = 0
        else:
            df['Potential Savings'] = 0

        return df

    def _extract_title_info(self, df_raw: pd.DataFrame) -> Dict[str, Any]:
        """Extract title and company information from Excel header"""
        title_info = {}

        try:
            # Company name (row 4, column B)
            if len(df_raw) > 3 and len(df_raw.columns) > 1:
                company = df_raw.iloc[3, 1]
                if pd.notna(company):
                    title_info['company_name'] = str(company)

            # Date range (row 6, column B)
            if len(df_raw) > 5 and len(df_raw.columns) > 1:
                date_range = df_raw.iloc[5, 1]
                if pd.notna(date_range):
                    title_info['date_range'] = str(date_range)
        except Exception:
            pass

        return title_info

    def calculate_summary_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate comprehensive summary statistics"""
        if df.empty:
            stats = {
                'total_loads': 0,
                'total_potential_savings': 0,
                'average_savings_per_load': 0,
                'loads_with_savings': 0,
                'total_selected_cost': 0,
                'total_least_cost': 0,
                'percentage_savings': 0
            }
            self.summary_stats = stats
            return stats

        # Calculate all stats
        total_loads = len(df)

        if 'Selected Total Cost' in df.columns:
            total_selected_cost = pd.to_numeric(
                df['Selected Total Cost'], errors='coerce').fillna(0).sum()
        else:
            total_selected_cost = 0

        if 'Least Cost Total Cost' in df.columns:
            total_least_cost = pd.to_numeric(
                df['Least Cost Total Cost'], errors='coerce').fillna(0).sum()
        else:
            total_least_cost = 0

        if 'Potential Savings' in df.columns:
            ps_series = pd.to_numeric(df['Potential Savings'], errors='coerce').fillna(0)
            total_potential_savings = ps_series.sum()
            ps_numeric = ps_series
        else:
            total_potential_savings = 0
            ps_numeric = pd.Series([0] * len(df))

        loads_with_savings = (ps_numeric > 0).sum()

        percentage_savings = (
            (total_potential_savings / total_selected_cost * 100)
            if total_selected_cost > 0 else 0
        )
        average_savings_per_load = total_potential_savings / total_loads if total_loads > 0 else 0

        stats = {
            'total_loads': total_loads,
            'total_potential_savings': total_potential_savings,
            'average_savings_per_load': average_savings_per_load,
            'loads_with_savings': loads_with_savings,
            'total_selected_cost': total_selected_cost,
            'total_least_cost': total_least_cost,
            'percentage_savings': percentage_savings
        }

        self.summary_stats = stats
        self.stats = stats  # For interface compatibility
        return stats

    def process_excel_file(self, file_path: str) -> pd.DataFrame:
        """Complete processing pipeline for basic processor"""
        print(f"Processing basic report: {os.path.basename(file_path)}")

        # Step 1: Load data
        df = self.load_data(file_path)

        # Step 2: Apply basic business rules
        df = self.apply_basic_business_rules(df)

        # Step 3: Calculate summary stats
        self.calculate_summary_stats(df)

        self.processed_data = df
        return df

    def clean_and_process_data(self, file_path: str) -> pd.DataFrame:
        """Compatibility method for GUI - calls process_excel_file"""
        return self.process_excel_file(file_path)

    def save_processed_data(self, output_file: str) -> None:
        """Save processed data to Excel with professional formatting"""
        if self.processed_data is None:
            raise Exception("No processed data to save")

        # Import required openpyxl modules
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter

        # Create a new workbook
        wb = openpyxl.Workbook()

        # Create processed data sheet
        ws_data = wb.active
        ws_data.title = "Processed Data"

        # Add company and date info (no big title rows)
        row = 1
        if hasattr(self, 'title_info') and self.title_info:
            # Style company and date range with expanded width
            header_style_border = Border(
                left=Side(style='medium', color='1F4E79'),
                right=Side(style='medium', color='1F4E79'),
                top=Side(style='medium', color='1F4E79'),
                bottom=Side(style='medium', color='1F4E79')
            )

            if 'company_name' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                company_cell = ws_data[f'A{row}']
                company_cell.value = f"Company: {self.title_info['company_name']}"
                company_cell.font = Font(size=12, bold=True, color="FFFFFF")
                company_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                company_cell.border = header_style_border
                company_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1

            if 'date_range' in self.title_info:
                last_col_letter = get_column_letter(len(self.processed_data.columns.tolist()))
                date_cell = ws_data[f'A{row}']
                date_cell.value = f"Date Range: {self.title_info['date_range']}"
                date_cell.font = Font(size=12, bold=True, color="FFFFFF")
                date_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                date_cell.border = header_style_border
                date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_data.merge_cells(f'A{row}:{last_col_letter}{row}')
                for col in range(1, len(self.processed_data.columns.tolist()) + 1):
                    cell = ws_data.cell(row=row, column=col)
                    cell.border = header_style_border
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws_data.row_dimensions[row].height = 30
                row += 1

            # Add section headers row with color coding
            row = 4
            # Selected Carrier section (columns I-N, which are 9-14) - Light Blue
            selected_header = ws_data.cell(row=row, column=9, value="Selected Carrier")
            selected_header.font = Font(size=10, bold=True, color="FFFFFF")
            selected_header.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            selected_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('I4:N4')
            for col in range(9, 15):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

            # Least Cost Carrier section (columns O-T, which are 15-20) - Light Orange
            least_cost_header = ws_data.cell(row=row, column=15, value="Least Cost Carrier")
            least_cost_header.font = Font(size=10, bold=True, color="FFFFFF")
            least_cost_header.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")
            least_cost_header.alignment = Alignment(horizontal="center", vertical="center")
            ws_data.merge_cells('O4:T4')
            for col in range(15, 21):
                cell = ws_data.cell(row=4, column=col)
                cell.fill = PatternFill(start_color="FF8C42", end_color="FF8C42", fill_type="solid")

            row = 5  # Headers will be on row 5

        # Add headers with enhanced styling and color coding
        headers = self.processed_data.columns.tolist()
        header_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=row, column=col_idx, value=header)
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Color code headers based on section
            if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light Blue
            elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Light Orange
            elif header == 'Potential Savings':  # Potential Savings column - Green
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
            else:
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Default blue

        # Add data with alternating row colors and comprehensive borders
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # Filter out any remaining empty rows before writing to Excel
        clean_data = self.processed_data.dropna(subset=['Load No.'])
        clean_data = clean_data[clean_data['Load No.'].astype(str).str.strip() != '']

        # Ensure all data is properly typed before processing
        for col in clean_data.columns:
            if col in ['Selected Transit Days', 'Selected Freight Cost', 'Selected Accessorial Cost', 'Selected Total Cost', 'Least Cost Transit Days', 'Least Cost Freight Cost', 'Least Cost Accessorial Cost', 'Least Cost Total Cost', 'Potential Savings']:
                clean_data[col] = pd.to_numeric(clean_data[col], errors='coerce').fillna(0)
            else:
                clean_data[col] = clean_data[col].astype(str).fillna('')

        for data_idx, data_row in enumerate(dataframe_to_rows(clean_data, index=False, header=False)):
            # Skip rows that are mostly empty
            try:
                non_empty_count = sum(1 for val in data_row if val is not None and str(val).strip() != '' and str(val) != 'nan')
                if non_empty_count < 3:
                    continue
            except Exception as e:
                print(f"Error processing row {data_idx}: {e}")
                continue

            row += 1
            row_color = "F8F9FA" if data_idx % 2 == 0 else "FFFFFF"

            # First pass: collect all content lengths to determine optimal row height
            max_content_length = 0
            for col_idx, value in enumerate(data_row, 1):
                content_length = len(str(value)) if value else 0
                max_content_length = max(max_content_length, content_length)

            # Enhanced dynamic height calculation for long carrier names
            has_carrier_data = any('TRANSPORT' in str(val).upper() or
                                 'LOGISTICS' in str(val).upper() or
                                 'FREIGHT' in str(val).upper() or
                                 len(str(val)) > 25 for val in data_row if val)

            if has_carrier_data and max_content_length > 25:
                optimal_height = min(50, max(30, max_content_length * 1.2))
            elif max_content_length > 30:
                optimal_height = min(45, max(25, max_content_length * 1.0))
            elif max_content_length > 20:
                optimal_height = min(35, max(22, max_content_length * 0.8))
            elif max_content_length > 15:
                optimal_height = 25
            else:
                optimal_height = 20

            ws_data.row_dimensions[row].height = optimal_height

            # Second pass: apply formatting to all cells in the row
            for col_idx, value in enumerate(data_row, 1):
                cell = ws_data.cell(row=row, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = data_border

                # Apply color coding to data cells based on section
                if 9 <= col_idx <= 14:  # Selected Carrier columns (I-N)
                    light_blue_bg = "E6F3FF" if data_idx % 2 == 0 else "F0F8FF"
                    cell.fill = PatternFill(start_color=light_blue_bg, end_color=light_blue_bg, fill_type="solid")
                elif 15 <= col_idx <= 20:  # Least Cost Carrier columns (O-T)
                    light_orange_bg = "FFF2E6" if data_idx % 2 == 0 else "FFF8F0"
                    cell.fill = PatternFill(start_color=light_orange_bg, end_color=light_orange_bg, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

                # Format currency columns and apply green color for positive Potential Savings values
                currency_columns = ['Selected Total Cost', 'Least Cost Total Cost', 'Selected Freight Cost', 'Least Cost Freight Cost', 'Selected Accessorial Cost', 'Least Cost Accessorial Cost']
                if headers[col_idx-1] in currency_columns or headers[col_idx-1] == 'Potential Savings':
                    cell.number_format = '"$"#,##0.00'
                    if headers[col_idx-1] == 'Potential Savings':
                        try:
                            if value is not None and value != '' and str(value).lower() != 'nan':
                                if isinstance(value, (int, float)):
                                    numeric_value = float(value)
                                else:
                                    numeric_value = float(str(value).replace('$', '').replace(',', ''))
                                if numeric_value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        except (ValueError, TypeError, AttributeError):
                            pass
                    cell.font = Font(size=10, bold=False, color="2C3E50")
                else:
                    cell.font = Font(size=10, color="495057")

        # Enable auto-filter over header and data range
        try:
            header_row_idx = 5
            ws_data.auto_filter.ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{row}"
        except Exception:
            pass

        # Add totals row with key financial metrics (if summary_stats available)
        if hasattr(self, 'summary_stats') and self.summary_stats:
            totals_row = row + 2

            # Add "TOTALS" label
            totals_label = ws_data.cell(row=totals_row, column=1, value="TOTALS")
            totals_label.font = Font(size=12, bold=True, color="FFFFFF")
            totals_label.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            totals_label.alignment = Alignment(horizontal="center", vertical="center")
            totals_label.border = Border(
                left=Side(style='medium', color='2C3E50'),
                right=Side(style='medium', color='2C3E50'),
                top=Side(style='medium', color='2C3E50'),
                bottom=Side(style='medium', color='2C3E50')
            )

            # Find the Selected Total Cost and Potential Savings columns
            selected_cost_col = None
            potential_savings_col = None
            for col_idx, header in enumerate(headers, 1):
                if 'Selected Total Cost' in str(header):
                    selected_cost_col = col_idx
                elif 'Potential Savings' in str(header):
                    potential_savings_col = col_idx

            # Add Total Selected Cost
            if selected_cost_col and 'total_selected_cost' in self.summary_stats:
                cost_cell = ws_data.cell(row=totals_row, column=selected_cost_col,
                                       value=f"${self.summary_stats['total_selected_cost']:,.2f}")
                cost_cell.font = Font(size=12, bold=True, color="FFFFFF")
                cost_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                cost_cell.alignment = Alignment(horizontal="center", vertical="center")
                cost_cell.number_format = '"$"#,##0.00'
                cost_cell.border = Border(
                    left=Side(style='medium', color='3498DB'),
                    right=Side(style='medium', color='3498DB'),
                    top=Side(style='medium', color='3498DB'),
                    bottom=Side(style='medium', color='3498DB')
                )

            # Add Total Potential Savings (most important number)
            if potential_savings_col and 'total_potential_savings' in self.summary_stats:
                savings_cell = ws_data.cell(row=totals_row, column=potential_savings_col,
                                          value=f"${self.summary_stats['total_potential_savings']:,.2f}")
                savings_cell.font = Font(size=14, bold=True, color="FFFFFF")
                savings_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                savings_cell.alignment = Alignment(horizontal="center", vertical="center")
                savings_cell.number_format = '"$"#,##0.00'
                savings_cell.border = Border(
                    left=Side(style='thick', color='27AE60'),
                    right=Side(style='thick', color='27AE60'),
                    top=Side(style='thick', color='27AE60'),
                    bottom=Side(style='thick', color='27AE60')
                )

            # Set height for totals row
            ws_data.row_dimensions[totals_row].height = 25

        # Auto-fit column widths on the Processed Data sheet
        try:
            for col_idx in range(1, ws_data.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(5, ws_data.max_row + 1):
                    cell = ws_data.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        text = str(cell.value)
                        length = len(text)
                        if length > max_length:
                            max_length = length

                # Check if this column contains text content
                has_text_content = False
                for check_row in range(5, ws_data.max_row + 1):
                    check_cell = ws_data.cell(row=check_row, column=col_idx)
                    if check_cell.value and any(c.isalpha() for c in str(check_cell.value)):
                        has_text_content = True
                        break

                # Very tight padding for maximum compactness
                if has_text_content:
                    padding = 1.0
                else:
                    padding = 0.5

                # Calculate width with aggressive compacting
                adjusted_width = max_length + padding

                # Apply maximum width constraints for compactness
                if has_text_content:
                    max_width = 25
                else:
                    max_width = 15

                final_width = min(adjusted_width, max_width)
                ws_data.column_dimensions[col_letter].width = max(6, final_width)

            # Optimize row heights for compact layout
            for rh in [1, 2, 4, 5]:
                if rh <= ws_data.max_row:
                    ws_data.row_dimensions[rh].height = max(ws_data.row_dimensions[rh].height or 0, 20)

            # Ensure gridlines visible
            ws_data.sheet_view.showGridLines = True
        except Exception:
            pass

        wb.save(output_file)